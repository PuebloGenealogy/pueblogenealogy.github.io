"""
Regenerate the genealogy chart.

    python3 scripts/make_chart.py            private build -> build/
    python3 scripts/make_chart.py --public   published build -> docs/

TWO SOURCES, TWO DESTINATIONS. The private build reads the workbook, including
anything typed into the green english_name / census_name columns, and lands in
build/, which is git-ignored. The public build reads scripts/transcription.py --
the 1923 baseline, which has no research columns to begin with -- and lands in
docs/, which is what GitHub Pages serves.

That is the privacy boundary, and it is structural rather than a matter of
remembering to strip columns: there is no code path from the workbook to docs/.
The two builds share the same Chart class and the same person_line(), because
in baseline mode the research keys are simply empty strings, so the chips that
render them never appear. See the README.
"""

import argparse
import base64
import datetime as dt
import html
import importlib
import json
import re
import shutil
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "data" / "parsons_genealogy_I.xlsx"

# The workbook holds research columns (english_name, census_name) that are NOT
# for publication, so this build lands in build/ -- which is git-ignored -- and
# never in docs/. Nothing under build/ reaches the internet.
OUT = ROOT / "build" / "genealogy-i-private.html"

DOCS = ROOT / "docs"
FONT_DIR = ROOT / "vendor" / "gentium"
# The finding aid's build output, copied in rather than generated here. See
# vendor/search/SOURCE.md for where it comes from and how to refresh it, and
# write_search() below for what this build does to it on the way to docs/.
SEARCH_DIR = ROOT / "vendor" / "search"

# The misprint annotation, one row under the line it annotates. Table-agnostic:
# the note it points at is the plate note, whose id every table that declares a
# misprint must carry.
SIC_ROW = ('<a class="sic" href="#note-misprint">'
           "(misprint, click here to see notes)</a>")


def _p(n, text=None):
    """
    An apparatus reference to a person on the same table: 8 -> <a href="#p8">8</a>.

    Written explicitly at every call site, never by regex over the prose. The
    apparatus is full of numbers that are not people -- 1923, vol. 19,
    pp. 133-292, U23, d. 1908 -- and a pattern loose enough to catch "58+59"
    would link those too. `text` carries a range or a phrase: _p(36, "36-43")
    points a printed range at its first member, the same rule linkify_xref uses
    on the chart.
    """
    return f'<a href="#p{n}">{n if text is None else text}</a>'


# One entry per transcribed plate. Adding a table is an entry here plus its
# transcription module -- nothing else in this file should need editing, which
# is the point: Table 1's numbering must not leak into the renderer.
TABLES = {
    "i": {
        "numeral": "I",
        "plate": "Table 1",
        "module": "transcription",
        "roots": [1, 54],          # the founding women, in plate order
        "slug": "genealogy-i",
        "couples": f"{_p(1)}+{_p(2)} and {_p(54)}+{_p(55)}",
        # child -> father, where the PLATE leaves paternity unassigned and this
        # edition supplies a reading. Deliberately not in transcription.py: that
        # module is the plate, and the plate does not say this. Applied to the
        # apparatus only -- the chart still draws the single bracket the plate
        # draws -- and every row it produces is marked and linked to its note.
        "paternity": {83: 69, 84: 69, 85: 70},
        "notes": f"""
    <li>{_p(8, "Person 8")} (Yu&#729;si) appears twice on the plate &mdash; as husband in
        the upper half and as a son of {_p(58)}+{_p(59)} in the lower half. Here he is
        drawn once, in the lower half, with a cross-reference standing in for the
        repeated sibling group.</li>
    <li id="note-misprint">The &lsquo;+&rsquo; line under {_p(76)} is numbered
        <strong>68</strong> on the plate but names Shuwai&#700;&#7590;ri, Turkey &mdash;
        {_p(67, "person 67")}. The chart prints <strong>68</strong>, as the plate does,
        and links it to 67; the misprint is recorded on union U23. Person 67&rsquo;s own
        cross-reference &mdash; &ldquo;For second wife and offspring see below,
        76, 90-3&rdquo; &mdash; is what identifies him.
        <strong>Table&nbsp;2 corroborates it independently</strong>: its own person 60
        is Shuwai&#700;&#7590;ri, Turkey, and the line beneath his name reads
        &ldquo;See Gen. I, 68&rdquo; &mdash; naming this same man 68 again, on a
        different plate. It is part of a pattern there, not an isolated slip:
        Table&nbsp;2&rsquo;s references into this genealogy are exact through
        person&nbsp;53 and one too high from person&nbsp;66 onward, checked by name,
        sex and clan on twenty-nine of them. So <strong>68 is Parsons&rsquo;s own
        number</strong>, carried from a numbering of Genealogy&nbsp;I that ran one
        ahead of the one finally printed &mdash; which is the reason this edition
        reproduces it rather than correcting it to 67. See
        <a href="../genealogy-ii/#note-crossref">Table&nbsp;2&rsquo;s note on the
        displacement</a>.</li>
    <li id="note-paternity">The plate brackets {_p(83)}, {_p(84)} and {_p(85)} under
        {_p(68)} alone. She has two husbands on the plate, {_p(69)} and {_p(70)}, and
        the bracket does not say which marriage the children belong to &mdash; which
        is why the transcription records their father as unassigned. <strong>The
        chart draws them as the plate does</strong>, in one group under 68. In the
        register and the person cards this edition attributes 83 and 84 to 68&rsquo;s
        marriage with 69, and 85 to her marriage with 70, marked &dagger; &mdash; on
        the parents&rsquo; entries and on each child&rsquo;s own, so the reading is
        visible from either end. <strong>The &dagger; marks the pairing, not the
        mother</strong>: that {_p(68)} is their mother is the plate&rsquo;s own
        bracket and is not in question. That
        attribution is an editorial reading resting on documentary evidence from
        outside the plate. The plate does not state it, and the supporting
        records are not reproduced in this edition.</li>
""",
    },
    "ii": {
        "numeral": "II",
        "plate": "Table 2",
        "module": "transcription_ii",
        # THREE blocks, and they are one genealogy, not three. 14 is a child of
        # 154+155 and a husband in the first block; 54 is a child of 232+233 and
        # a husband in the first. That is why the derived generations of those
        # two are 2 and 3 rather than 1: they are the parents of people drawn
        # above them.
        #
        # 31 WAS a fourth root here until 2026-07-30, and that was wrong -- not
        # about the genealogy but about the page. He has no leader stub, so
        # nothing links him to a bracket and rooting him was the only way he
        # got drawn at all; but a root is drawn at generation 1, at the far
        # left, and the plate sets him at the children's indent inside 9+10's
        # bracket, four columns over. The block was correct and its position
        # was not. He is now spliced into that column by UNATTACHED_BLOCKS in
        # transcription_ii.py, which withholds the stub, so the page says what
        # the plate says: printed here, descent not drawn.
        "roots": [1, 154, 232],
        # A root is drawn at generation 1 unless it says otherwise, and on this
        # plate two of the three say otherwise. Parsons does not set the lower
        # block at the sheet's left margin: measured on the scan 2026-07-30,
        # person 1 sits at x 225 and person 3 -- generation 2 -- at x 1425,
        # while the lower block's 154 sits at x 1340 (person 3's column) and
        # 232 at x 2690, the same column as 164, who is 154+155's child. So the
        # plate prints 154+155 one column in and 232+233 two columns in, and
        # the generation field already stored for them, 2 and 3, is that same
        # reading arrived at independently by walking the tree.
        #
        # This is an INDENT, not a splice. UNATTACHED_BLOCKS puts a couple
        # inside somebody's child column, which is right for 31 -- the plate
        # prints him among 9+10's children with the vertical passing his row.
        # It is wrong here: the lower block is not descended from the upper
        # one, it is a separate block that simply starts further in, and the
        # bracket-column strip at x 2480, y 9900 shows 154+155's vertical
        # ending on 164 with nothing beside 232 at all. Splicing would also run
        # into self_check()'s last-child rule for exactly that reason.
        # Expressed in the same GEOM tokens as the grid, so generation d still
        # lands at d x (--col + --stub) and column drift stays 0.
        "root_columns": {154: 2, 232: 3},
        "slug": "genealogy-ii",
        "couples": (f"{_p(1)}+{_p(2)}, {_p(154)}+{_p(155)} "
                    f"and {_p(232)}+{_p(233)}"),
        # The second editorial attribution in the edition, and the first resting
        # on a source that can be CITED. The plate brackets 116-118 under 48
        # alone; she has two husbands, 47 and 49. Parsons's own text, p. 195,
        # says of "Gen. II, 47" that he died offspring lacking -- so the children
        # are not his, and 49 is the only other husband. This was deliberately
        # NOT encoded on 2026-07-30, because METHOD.md requires every such row to
        # be daggered to a footnote and no source had been found in her text.
        # The source is now found, quoted in note-paternity below, so the rule
        # that blocked it -- an attribution that cannot be footnoted is not
        # made -- is satisfied rather than waived.
        "paternity": {116: 49, 117: 49, 118: 49},
        "notes": f"""
    <li id="note-duplicate-101">The plate <strong>numbers two different people
        101</strong> &mdash; a woman, Naaʼd&#7590;ityʼi of the Water clan, and a man
        whose name it prints as a dash. Both rows print <strong>101</strong> here,
        because the chart prints what the plate prints. They are separate entries in
        the register and separate person cards; the numbering, not the edition, is
        what is ambiguous. One consequence a reader will meet: typing
        &ldquo;101&rdquo; into Find goes to the first of the two, since the number
        alone cannot say which is meant. Searching the name, or opening either row
        from the register, reaches each of them exactly.</li>
    <li id="note-unattached">{_p(31)} and {_p(32)} sit inside the bracket holding
        {_p(9)} and {_p(10)}&rsquo;s children, at the children&rsquo;s indent,
        <strong>with no rule joining them to it</strong> &mdash; and that gap is the
        plate&rsquo;s, reproduced here rather than tidied away. Every other person in
        a bracket is reached by a short horizontal stub off its vertical rule;
        {_p(31)}&rsquo;s row is the one the rule passes without stopping. So Parsons
        prints him among {_p(9)}&rsquo;s children while declining to say he is one,
        and this edition prints him where she does and says no more. His clan is
        Water, as {_p(9)} and {_p(10)} are, so the matrilineal check that tests every
        other bracket on this plate can neither confirm nor deny it; the absent stub
        is the whole of the evidence. His wife {_p(32)} and their child {_p(97)}
        follow from his line as usual.</li>
    <li>Six people are drawn twice on the plate and are drawn <strong>once</strong>
        here, with the plate&rsquo;s own &ldquo;For descendants, see above&rdquo;
        standing in for the repeat: {_p(13)}, {_p(14)}, {_p(53)}, {_p(54)}, {_p(125)}
        and {_p(126)}. Person {_p(169)} repeats within the lower block in the same way.</li>
    <li id="note-repeat-names">Where a person is drawn twice, <strong>the plate
        sometimes sets the name differently in the two places</strong> &mdash;
        {_p(13)} (Dzia&#729;&#700;yots&#700;a / Tsiaiutsa), {_p(54)}
        (Ma&#729;&#700;rani / Ma&#729;&#700;ran&#729;i) and {_p(125)}
        (Gowa&#700;k&#729;&#700;d&#729;y&#259;i&#700; / Gowa&#700;k&#700;ad&#729;z&#259;i&#700;).
        Both settings were read at high magnification and both are what the plate
        prints, so this edition carries both rather than choosing between them.
        A <strong>parenthesis after a name</strong> on this table therefore means
        one of three things: the second setting of a repeat person&rsquo;s name, as
        here; an English name the plate itself prints in parentheses, at {_p(27)},
        {_p(42)}, {_p(43)} and {_p(140)}; or the second half of a pair the plate
        joins with a brace, at {_p(14)}.</li>
    <li id="note-crossref">Parsons&rsquo;s cross-references from this plate into
        Genealogy&nbsp;I <strong>do not all use the numbering Genealogy&nbsp;I was
        finally printed with</strong>. Checked by name, sex and clan, they are exact
        through Genealogy&nbsp;I&rsquo;s person&nbsp;53 &mdash; twenty-two matches
        &mdash; and <strong>one too high</strong> from its person&nbsp;66 onward, on
        seven more. Nothing between its 54 and 65 is referenced from here, so where
        the displacement begins cannot be settled from these two plates.
        The references are reproduced exactly as printed and are deliberately
        <strong>not</strong> linked: a link would have to resolve the displacement,
        and resolving it silently would send a reader to the wrong person.
        This is the same phenomenon as Genealogy&nbsp;I&rsquo;s own misprint at
        <a href="../genealogy-i/#note-misprint">person&nbsp;67</a>, which
        Table&nbsp;2 independently numbers 68 at {_p(60)}.</li>
    <li>One cross-reference carries <strong>Parsons&rsquo;s own question mark</strong>
        &mdash; {_p(199)} prints &ldquo;See Gen. I, 43 (?)&rdquo;. It is the only
        reference on the plate with one, and the only one in the exact range that does
        not match by name. The &ldquo;(?)&rdquo; is hers and is reproduced.</li>
    <li>Three lines print a word in the clan position that is <strong>not a
        clan</strong> &mdash; White at {_p(179)}, Mexican at {_p(183)}, Mohave at
        {_p(243)}. Each is recorded where the plate sets it, not reinterpreted.
        Person {_p(161)} is marked &ldquo;M.-F.&rdquo;, a notation used nowhere else
        on the plate; the sex is left as printed rather than resolved.</li>
    <li>Persons {_p(160)} and {_p(163)} carry cross-references into
        <strong><a href="../genealogy-iii/">Genealogy&nbsp;III</a></strong>, which is
        now transcribed and published here. The references themselves are printed as
        the plate has them, without links on the numbers: those numbers are Genealogy
        III&rsquo;s own numbering, and every numbered link on this page is an anchor
        within this plate. Following one means opening that table and looking the
        number up there.</li>
    <li>The plate draws a short rule from {_p(235)}+{_p(236)}&rsquo;s sibling bracket
        to the &lsquo;+&rsquo; line of {_p(255)}, which it does for no other spouse
        line. It cannot mean 255 is their child: every child of 236 is Water, her own
        clan, and 255 is Eagle. It is recorded as an observation of the plate, not as
        descent.</li>
    <li id="note-paternity">The plate brackets {_p(116)}, {_p(117)} and {_p(118)}
        under {_p(48)} alone. She has two husbands on the plate, {_p(47)} and
        {_p(49)}, and the bracket does not say which marriage the children belong to
        &mdash; which is why the transcription records their father as unassigned.
        <strong>The chart draws them as the plate does</strong>, in the one group
        under 48. In the register and the person cards this edition attributes all
        three to 48&rsquo;s marriage with {_p(49)}, marked &dagger; &mdash; on 48
        and 49&rsquo;s entries and on each child&rsquo;s own, so the reading is
        visible from either end. <strong>The &dagger; marks the pairing, not the
        mother</strong>: that these children belong to 48&rsquo;s marriage with 49
        rather than with 47. That {_p(48)} is their mother is the plate&rsquo;s own
        bracket and is not in question.
        <strong>Parsons&rsquo;s own text is what settles it.</strong> Writing on
        inheritance, she records that &ldquo;in one instance noted (Gen.&nbsp;II, 47),
        offspring lacking, sheep and fields were inherited
        by the widow, not by the sister of the deceased or his brothers&rdquo;
        (p.&nbsp;195) &mdash; naming this man, on this genealogy, as having died
        <strong>without offspring</strong>. So 116&ndash;118 are not his, and
        {_p(49)} is the only other husband the plate gives {_p(48)}. Her sentence
        also agrees with the plate independently: it has him dead and 48 surviving
        him, which is what the plate&rsquo;s &ldquo;d.&rdquo; on 47 records.
        This differs from the attribution on
        <a href="../genealogy-i/#note-paternity">Genealogy&nbsp;I</a> in the one way
        that matters to a reader checking it: that one rests on evidence this
        edition does not reproduce, and this one on a published source, quoted and
        cited here, which anyone can weigh for themselves.</li>
""",
    },
    "iii": {
        "numeral": "III",
        "plate": "Table 3",
        "module": "transcription_iii",
        "roots": [1, 230],         # the founding women, in plate order
        # 230's block is NOT descended from 1+2 -- the plate simply sets it in
        # the generation-2 column rather than at the sheet's left edge. Same
        # case as Genealogy II's 154 and 232: an indent, not a splice.
        "root_columns": {230: 2},
        "slug": "genealogy-iii",
        "couples": f"{_p(1)}+{_p(2)} and {_p(230)}+{_p(231)}",
        # No "paternity" key, and that is a finding about the plate rather than
        # an omission. This plate marks paternity itself: the leader rule
        # reaching a sibling bracket sits on the line of the parent whose
        # marriage the group belongs to, so a spouse with no leader had no
        # recorded issue. 85/86/87 is Genealogy I's 83-85 shape and still needs
        # nothing, because 86's leader is on her own line and 87 has none.
        "notes": f"""
    <li id="note-paternity-rule">This plate <strong>says which marriage a sibling
        group belongs to</strong>, where Genealogies&nbsp;I and&nbsp;II leave it open.
        The rule reaching a bracket is drawn from the line of the parent whose
        marriage it is, so where someone has two spouses, each spouse&rsquo;s line
        either carries that rule or does not. {_p(86)} and {_p(87)} are the case that
        would need an editorial reading on another plate &mdash; one woman, two
        husbands, one bracket &mdash; and here it needs none: {_p(86)}&rsquo;s line
        carries the rule and {_p(87)}&rsquo;s does not. <strong>Nothing on this table
        is attributed by this edition</strong>; the plate does it.</li>
    <li id="note-overdrawn">One bracket on the plate is <strong>drawn further than it
        should be</strong>. {_p(22)}&rsquo;s children are {_p(80)} and {_p(82)}, and
        {_p(83)} is {_p(25)}&rsquo;s son &mdash; {_p(25)}&rsquo;s own rule reaches him
        on his row. But the vertical that gathers {_p(80)} and {_p(82)} is carried on
        down past {_p(82)} to touch {_p(83)} as well, so on the plate the three appear
        to share one bracket. <strong>The chart draws two</strong>, which is the one
        place on this table where it departs from the scan. All three children are of
        the Corn clan, as both mothers are, so matrilineal descent cannot settle it;
        the two rules entering the bracket, on {_p(22)}&rsquo;s row and
        {_p(25)}&rsquo;s, are the evidence. Elsewhere this plate keeps adjacent
        brackets plainly apart, offsetting their verticals where they meet.</li>
    <li>Four couples are <strong>printed twice</strong>, the second setting carrying
        the plate&rsquo;s own &ldquo;For her descendants, see above&rdquo; in place of
        the bracket: {_p(7)}+{_p(8)}, {_p(91)}+{_p(92)}, {_p(124)}+{_p(125)} and
        {_p(152)}+{_p(153)}. Each is drawn once here, where its issue is drawn.
        {_p(166)} and {_p(167)} are printed twice as well, with no descendants line.
        Person {_p(8)} is what <strong>joins the plate&rsquo;s two blocks</strong>:
        husband of {_p(7)} in the first and a son of {_p(236)}+{_p(237)} in the
        second.</li>
    <li id="note-duplicate-numbers">The numbering on this plate is
        <strong>not a unique key</strong>. The numbers <strong>258</strong> and
        <strong>259</strong> are each printed on two different people, and
        <strong>256</strong> and <strong>257</strong> appear nowhere. Both pairs were
        read at high magnification to rule out a misreading. This edition states the
        fact and does not guess at the cause: all four rows print the number the plate
        prints, and each is a separate entry in the register with its own card.</li>
    <li id="note-misprint">Three entries are <strong>printed with a value this
        edition reads as a misprint</strong>, and each is set on the chart as the
        plate sets it, ringed in red with this note beneath it. The chart is not
        corrected; the correction is here.
        <ul>
          <li>{_p(37)} is printed <strong>M.</strong> She is the mother of
              {_p(109, "109-112")}, whose clan is Chaparral Cock, where her husband
              {_p(36)} is Lizard. <strong>Clan descent is matrilineal</strong>, so
              the children&rsquo;s clan is hers and she is a woman; her bracket and
              her name agree. Read at high magnification to rule out a broken
              <em>F</em>. Everything computed from this table &mdash; her bracket,
              her children&rsquo;s clan, the register &mdash; follows the reading,
              not the letter.</li>
          <li>{_p(50)}&rsquo;s clan is printed <strong>Chapparral Cock</strong>, with
              a doubled <em>p</em>. Her own four children {_p(135, "135-138")}, and
              every other line on this plate, print <em>Chaparral Cock</em>.</li>
          <li>{_p(255)}&rsquo;s clan is printed <strong>Bager</strong> for
              <em>Badger</em>.</li>
        </ul></li>
    <li id="note-crossref">This plate&rsquo;s cross-references into
        Genealogies&nbsp;I and&nbsp;II have been checked against this edition&rsquo;s
        transcriptions of those plates, by name, sex and clan &mdash; fifty-one
        references on the numbered lines and five more in the two prose notes under
        {_p(155)}. They are <strong>reproduced exactly as printed</strong>, and the
        numbers are deliberately not links: another plate&rsquo;s numbering is not an
        anchor on this page, so following one means opening that table and looking
        the number up there.
        <strong>This plate does not share Genealogy&nbsp;II&rsquo;s
        displacement.</strong> Table&nbsp;2&rsquo;s references into
        Genealogy&nbsp;I run <a href="../genealogy-ii/#note-crossref">one too high
        from its person&nbsp;66 onward</a>; this plate&rsquo;s are exact right across
        that same range, its citations of Genealogy&nbsp;I&rsquo;s 78, 79, 97, 98,
        99, 100, 101, 103 and 104 each reaching the person finally printed under that
        number. Genealogy&nbsp;III was numbered against the Genealogy&nbsp;I that went
        to press; Genealogy&nbsp;II was not. Neither plate&rsquo;s numbering may be
        used to correct the other.
        <strong>Four references do not resolve as printed</strong>, and no two are
        the same kind of slip:
        <ul>
          <li>{_p(170, "170-174")} cite Genealogy&nbsp;II&rsquo;s 191 to 195 on five
              consecutive lines. The people named there are its <strong>201 to
              205</strong> &mdash; five for five, in that order, all of the Sun clan.
              Genealogy&nbsp;II&rsquo;s actual 191 to 195 are a group of Oak people,
              and this plate cites <em>them</em> correctly elsewhere, from {_p(194)},
              {_p(195)} and {_p(198)}.</li>
          <li>{_p(218)} cites Genealogy&nbsp;I&rsquo;s 101, who is the father already
              cited from {_p(257)}. The person meant is its <strong>102</strong>. Her
              sisters {_p(219)} and {_p(261)} cite 103 and 104 exactly, so this is an
              isolated slip and not a run.</li>
          <li>{_p(173)} cites &ldquo;Gen.&nbsp;I, 149&rdquo;, and
              <strong>Genealogy&nbsp;I numbers 104 people</strong>. The reference
              resolves to nobody. It was read again at high magnification: 149 is
              what the plate prints. The person is Genealogy&nbsp;I&rsquo;s
              <strong>49</strong>, which is how Genealogy&nbsp;II cites her from its
              own 204.</li>
          <li>The prose note under {_p(155)} sends a reader to
              &ldquo;Gen.&nbsp;I, 8, 90&rdquo;. The husband, 8, is exact; the
              descendant is <strong>89</strong>, the one child of
              Genealogy&nbsp;I&rsquo;s 73+8. This is the only place on this plate
              where Genealogy&nbsp;II&rsquo;s displacement appears, and it attests it
              from a third plate. {_p(155)}&rsquo;s other prose note, into
              Genealogy&nbsp;II, resolves exactly.</li>
        </ul></li>
    <li>Many entries carry <strong>no sex letter</strong> &mdash; most of the unnamed
        children in the last two generations, and {_p(155)}, {_p(193)}, {_p(195)} and
        {_p(260)}, who are named or numbered but printed without one. Where the plate
        records no sex, none is supplied.</li>
""",
    },
    "iv": {
        "numeral": "IV",
        "plate": "Table 4",
        "module": "transcription_iv",
        "roots": [1, 59],          # the founding women, in plate order
        "slug": "genealogy-iv",
        "couples": f"{_p(1)}+{_p(2)} and {_p(59)}+{_p(60)}",
        "notes": f"""
    <li>Persons {_p(3)} and {_p(4)} appear twice on the plate. Person {_p(4)} links the
        two families: husband of {_p(3)} in the upper half, son of {_p(59)}+{_p(60)} in
        the lower. They are drawn once, with &ldquo;For descendants, see above&rdquo;
        standing in for the repeat, as the plate has it.</li>
    <li>Two sibling groups are printed collectively &mdash;
        &ldquo;{_p(36, "36-43")}. 8 children deceased&rdquo; and
        &ldquo;{_p(50, "50-53")}. 4 children deceased&rdquo;. The plate assigns
        each of them a number, so each is drawn, with the collective setting recorded
        in the data.</li>
    <li>The English names in parentheses &mdash; Hugh, Frank, Paul and Joe Johnson,
        and Mana &mdash; are printed on the plate. They are part of the transcription,
        not additions to it.</li>
    <li>Persons {_p(19)} and {_p(20)} are printed with no sex, and person {_p(73)} with
        no father: the plate records neither, so neither is supplied.</li>
""",
    },
}

# Plates referenced by a transcribed table but not yet transcribed themselves.
# They appear on the landing page as inert cards so the edition states its own
# scope rather than implying Genealogy I is the whole work.
PENDING = []

NUMBER_WORDS = {1: "one", 2: "two", 3: "three", 4: "four", 5: "five",
                6: "six", 7: "seven"}

# The chart's geometry, in rem at the fixed 16px root. Single source of truth:
# the CSS custom properties are emitted from here, and the print reduction is
# computed from here, so the two can never disagree. The values themselves are
# the long-verified ones -- --col must exceed the widest line at generations
# 1 to n-1 or a line will spill into the next column (see the CSS comment).
GEOM = {"lh_rem": 1.55, "stub_rem": 2.6, "col_rem": 24, "sheet_pad_rem": 2.2}


def geom_css():
    """The geometry tokens, emitted from GEOM rather than typed into the CSS."""
    return (":root{--lh:%srem;--stub:%srem;--col:%srem;--sheet-pad:%srem}\n"
            % (GEOM["lh_rem"], GEOM["stub_rem"], GEOM["col_rem"],
               GEOM["sheet_pad_rem"]))


def print_zoom(n_gens):
    """
    The zoom that fits one sheet on a landscape page, from the same GEOM the
    layout uses. 1030px is the printable width of landscape A4/letter at 96dpi
    inside the 12mm @page margins; the browser's shrink-to-fit is the backstop.
    """
    rem = (2 * GEOM["sheet_pad_rem"]
           + (n_gens - 1) * (GEOM["col_rem"] + GEOM["stub_rem"])
           + GEOM["col_rem"])
    sheet_px = rem * 16 + 2   # + the sheet's own border
    return min(1.0, round(1030 / sheet_px, 3))

SITE = "https://pueblogenealogy.github.io"
REPO = "https://github.com/PuebloGenealogy/pueblogenealogy.github.io"

# There is deliberately no DOI constant here. The edition carried the Zenodo
# concept doi from launch until 2026-08-08, when the user withdrew the archive
# from the site: no deposit is advertised, the webhook is off, and no release
# is cut. The v1.0.0 deposit was deleted by its owner the same day (Zenodo
# allows this within 30 days of publishing); both dois now return 410 Gone at a
# tombstone that keeps the metadata. Re-adding a doi to this file is how the
# archive would come back. See CLAUDE.md -> Exposure posture first.

# The social preview card. A 1200x630 band of the real Table 1 plate, derived
# once from sources/parsons-1923-table-1.jpg and committed at assets/og-cover.jpg
# rather than regenerated per build -- the source scan is 33 MB and `sips` is
# macOS-only, so a build step would be both slow and unportable. write_site()
# copies it into docs/. Regenerate with:
#   sips -s format jpeg -Z 1200 sources/parsons-1923-table-1.jpg --out /tmp/w.jpg
#   sips -c 630 1200 /tmp/w.jpg --out assets/og-cover.jpg
OG_IMAGE = f"{SITE}/og-cover.jpg"
OG_ALT = ("A detail of the 1923 foldout plate: numbered individuals with Keresan "
          "names, clans and marriage lines, in five generation columns.")

# The subject terms, in one place: they feed the Dataset structured data and the
# <meta name="keywords"> on the landing page. These describe the edition -- they
# are not a place to add search bait the pages do not actually deliver.
KEYWORDS = [
    "Laguna Pueblo genealogy",
    "Elsie Clews Parsons",
    "Laguna Genealogies",
    "Pueblo genealogy",
    "Keresan",
    "Kawaika",
    "Southwest anthropology",
    "American Museum of Natural History",
    "Anthropological Papers",
    "matrilineal clan descent",
    "digital edition",
    "New Mexico",
]

CITATION_TEXT = ('Elsie Clews Parsons, "Laguna Genealogies", Anthropological Papers '
                 "of the American Museum of Natural History, vol. 19, pt. 5 (1923), "
                 "pp. 133-292.")

# The landing page's one-sentence summary, used for the meta description, the
# Open Graph card and the CollectionPage structured data, so all three agree.
SITE_DESCRIPTION = (
    "A digital edition of the Laguna Pueblo genealogical plates published by "
    "Elsie Clews Parsons in \"Laguna Genealogies\" (1923). All four plates "
    "transcribed character by character from the originals and redrawn as text "
    "you can search, copy and check against the plate.")

# Questions people actually type, answered on the page. The answers are the page
# text verbatim, which is what FAQPage requires -- markup that answers something
# the page does not say is a guideline violation, not a shortcut.
FAQ = [
    ("What are the Laguna Genealogies?",
     "They are a set of foldout genealogical plates published by the anthropologist "
     "Elsie Clews Parsons in \"Laguna Genealogies\", Anthropological Papers of the "
     "American Museum of Natural History, volume 19, part 5 (1923), pages 133-292. "
     "Each plate charts several generations of related families at Laguna Pueblo, "
     "New Mexico, recording each person's number, sex, Keresan name and clan."),
    ("Which plates are transcribed here?",
     "All four. Table 1 (Genealogy I), Table 2 (Genealogy II), Table 3 "
     "(Genealogy III) and Table 4 (Genealogy IV) are transcribed in full and "
     "published on this site, which is the whole of the genealogical material "
     "published with Parsons's 1923 paper."),
    ("How accurate is the transcription?",
     "The plates are transcribed character by character, including the Americanist "
     "phonetic diacritics, and nothing is corrected, normalised or filled in. Where "
     "the plate contains a misprint it is reproduced and annotated rather than "
     "silently fixed, and where Parsons recorded no name the entry is left blank. "
     "Because clan descent at Laguna is matrilineal, every child's clan must match "
     "its mother's, which checks the bracket readings wherever the possible mothers "
     "belong to different clans; where they do not, the evidence is the geometry of "
     "the plate's own brackets, read at full resolution."),
    ("Can I use this for family history research?",
     "Yes. The edition is a finding aid for the printed 1923 record and is released "
     "under CC BY 4.0. Note that it publishes the 1923 transcription only: no modern "
     "names, census matches or identifications of living people are included."),
    ("Why are the names hard to match to census records?",
     "Parsons recorded names in an Americanist phonetic transcription, which differs "
     "sharply from the spellings used by census takers. Each name is therefore also "
     "stored as a diacritic-free key -- Kiwaʼd˙yuwi becomes kiwadyuwi -- so it can be "
     "joined against records that spell the same name differently."),
]


def faq_html():
    """The FAQ, as ordinary page text. The structured data quotes these answers."""
    items = "".join(
        f"    <details class=\"faq\">\n"
        f"      <summary>{esc(q)}</summary>\n"
        f"      <p>{esc(a)}</p>\n"
        f"    </details>\n"
        for q, a in FAQ)
    return f"  <h2>Common questions</h2>\n{items}"


def jsonld_faq():
    data = {
        "@context": "https://schema.org",
        "@type": "FAQPage",
        "mainEntity": [
            {"@type": "Question", "name": q,
             "acceptedAnswer": {"@type": "Answer", "text": a}}
            for q, a in FAQ
        ],
    }
    return ('<script type="application/ld+json">'
            + json.dumps(data, ensure_ascii=False) + "</script>")


def social_meta(title, description, url, kind="article"):
    """
    The Open Graph and Twitter block. One definition for every page, so a card
    can never be right on one page and missing on another.
    """
    return f"""<meta property="og:type" content="{kind}">
<meta property="og:title" content="{title}">
<meta property="og:description" content="{esc(description)}">
<meta property="og:url" content="{url}">
<meta property="og:site_name" content="Laguna Genealogies: A Digital Edition">
<meta property="og:locale" content="en_US">
<meta property="og:image" content="{OG_IMAGE}">
<meta property="og:image:width" content="1200">
<meta property="og:image:height" content="630">
<meta property="og:image:alt" content="{esc(OG_ALT)}">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:title" content="{title}">
<meta name="twitter:description" content="{esc(description)}">
<meta name="twitter:image" content="{OG_IMAGE}">
<meta name="twitter:image:alt" content="{esc(OG_ALT)}">"""
AUTHOR = "Elizabeth Heger-Vlahovic"

# Google Search Console ownership token, from Settings -> Ownership verification
# -> HTML tag (the content="..." value only, not the whole tag). Left empty until
# a property is created; while empty no tag is emitted and the output is
# unchanged. It lives here rather than as a hand-placed file in docs/ because
# docs/ is generated and a stray file there would be easy to lose.
GOOGLE_SITE_VERIFICATION = "7SJ_xFuG2D2skZIWOxUhmKwkBZUXD3HOkisTxSxSIlQ"


def describe(spec, n_persons, n_gens):
    """The meta/OG description for one table. Generated so counts cannot go stale."""
    gens = NUMBER_WORDS.get(n_gens, str(n_gens))
    return (
        f"A verified digital transcription of {spec['plate']}, "
        f"“Genealogy {spec['numeral']}”, from Elsie Clews Parsons’s Laguna "
        f"Genealogies (1923): {n_persons} individuals across {gens} generations, "
        "with names, ages, clans and marriages as printed."
    )

# Where the plate replaces a sibling bracket with a cross-reference because the
# same children are already drawn elsewhere on the sheet. Union U04 (7 x 8) is
# drawn under 7 in the upper half; under 8 in the lower half the plate prints
# this note instead. See PLATE_NOTES in the workbook.
SECOND_VISIT_NOTE = {
    "U04": "For descendants, see above, 13-15, 28-30",   # Table 1, under 8
    "V02": "For descendants, see above",                 # Table 4, under 4
}

# Where the plate does not repeat the marriage AT ALL on the second visit --
# no '+' line, no bracket, no child-column note, only a prose cross-reference
# printed in the block. This is NOT SECOND_VISIT_NOTE: there the plate does
# print the '+' line and replaces only the sibling bracket, which is why the
# note sits in the child column on that spouse's row.
#
# The distinction is structural, not cosmetic. Table 1's person 8 has two
# DIFFERENT wives, 7 and 73, so his two groups hang off two different rows and
# nothing collides. Genealogy II's 169 has two husbands and is the mother of
# both groups, so `u["wife"] == pid` gives both `mother_row = 0` -- and two
# brackets cannot begin on one line. The push logic then moves 169's own line
# down to meet the second group and strands the first, one --lh out. Parsons
# has no such problem because she prints 169 TWICE, one marriage each: under
# 156+157 as 168's wife with the bracket to 196-200, and under 164+165 as her
# parents' daughter with the bracket to 225, 226. Verified on the scan at
# x 3650 y 7500 and x 3650 y 9700, 1500 px wide, 2026-07-30.
#
# So the collision was self-inflicted -- the renderer printed a marriage in a
# block where the plate prints none -- and suppressing it is what makes the
# page agree with the scan. Keyed by union id; the value is the plate's own
# line, verbatim.
SECOND_VISIT_OMITTED = {
    "U43": "For second husband and descendants, see above",  # Table 2, under 164+165
}

RESEARCH_KEYS = ("english_name", "census_name", "census_year", "match_confidence", "notes")
BASE_KEYS = ("sex", "name_as_printed", "alt_name", "age", "clan", "vital_note",
             "origin", "cross_ref", "plate_note")


def load_baseline(spec):
    """
    The same four structures load() returns, built from a table's 1923 baseline.

    Every research key is set to "" -- not omitted -- so Chart and person_line
    work unchanged and the English-name and census chips cannot render. The
    absence of research data here is a property of the module, not of this
    function: a transcription module has no such columns to read.
    """
    sys.path.insert(0, str(ROOT / "scripts"))
    T = importlib.import_module(spec["module"])

    problems = T.self_check()
    if problems:
        raise SystemExit(f"{spec['module']}.py failed its structural checks:\n  "
                         + "\n  ".join(problems))

    # A table may declare DUPLICATE_PLATE_NUMBERS: {synthetic id: the number the
    # plate prints}. Parsons numbers two different people 101 on Genealogy II, so
    # the plate's numbering is not a unique key and one of the two needs an id of
    # its own to be addressable -- anchors, register entries and relation chips
    # all key on it. But the id is PLUMBING, and every place a number is SHOWN
    # must show the plate's. Without this both rows were addressable and the
    # second printed "1010.", a number that appears nowhere on the plate, in
    # three places: its chart line, its register entry and every relation chip
    # pointing at it.
    #
    # Note what this is NOT: a misprint. PLATE_NUMBER_MISPRINTS also separates a
    # shown number from an id, but there the plate is WRONG and the number is
    # ringed in --sic and carries an annotation row. Here the plate is right and
    # simply reuses a number, so it is set exactly like every other number, with
    # no ring and no annotation. A footer note explains the reuse once.
    dupes = getattr(T, "DUPLICATE_PLATE_NUMBERS", {})

    # PLATE_MISPRINTS: {field: {id: what the plate prints}}, for the fields that
    # are not the number -- Genealogy III prints "M." on 37, who is plainly a
    # woman, and misspells two clans. Same principle as PLATE_NUMBER_MISPRINTS
    # and the same treatment: the PLATE's value is what is shown, ringed in
    # --sic with an annotation row, while the DATA keeps the reading the plate's
    # own bracket and clan descent establish, because that is what the structure
    # is computed from. Declared per table; read with getattr.
    field_misprints = getattr(T, "PLATE_MISPRINTS", {})

    persons = {}
    for pid, _gen, sex, name, alt, age, clan, vital, origin, xref, note in T.PERSONS:
        p = dict(zip(BASE_KEYS, (sex, name, alt, age, clan, vital, origin, xref, note)))
        p = {k: ("" if v is None else str(v).strip()) for k, v in p.items()}
        p.update({k: "" for k in RESEARCH_KEYS})
        p["id"] = pid
        p["plate_number"] = dupes.get(pid, pid)
        p["generation"] = _gen
        # "M." and "M" both read as the sex letter: person_line sets the point.
        p["printed_sex"] = str(field_misprints.get("sex", {})
                               .get(pid, "")).strip().rstrip(".")
        p["printed_clan"] = str(field_misprints.get("clan", {})
                                .get(pid, "")).strip()
        persons[pid] = p

    # UNIONS rows are 6-tuples, or 7 with a trailing drawn_under -- the id of
    # the block the plate prints this marriage inside, when that is not either
    # partner's own block. Table 1 uses none, so it is read as absent there.
    # A table may declare PLATE_NUMBER_MISPRINTS: the number the plate prints on
    # a '+' line where it is not the number of the person that line names. Read
    # with getattr so a table without one needs no entry -- Table 4 has none.
    misprints = getattr(T, "PLATE_NUMBER_MISPRINTS", {})
    # LEADER_ON_SPOUSE_ROW: unions whose sibling bracket the plate hangs off the
    # '+' SPOUSE's line rather than off the mother's. The default -- the bracket
    # is the mother's row -- is right on every plate for a first marriage, and on
    # Genealogy III it is wrong for a second one: that plate draws the leader
    # from the line of the parent whose marriage the group belongs to, so a
    # woman's second husband carries his own leader (see transcription_iii's
    # docstring, point 1). Without it BOTH of person 43's unions claim her line,
    # the second group cannot start there, and the push logic below moves her
    # line down to meet it -- stranding the first group, which is the identical
    # failure CLAUDE.md records for Genealogy II's 169. Read with getattr:
    # Tables 1, 2 and 4 declare none.
    leader_on_spouse = set(getattr(T, "LEADER_ON_SPOUSE_ROW", ()))
    unions = []
    for row in T.UNIONS:
        uid, wife, husband, _wo, _ho, note = row[:6]
        unions.append({
            "union_id": uid,
            "wife": wife or 0,
            "husband": husband or 0,
            "note": note or "",
            "drawn_under": (row[6] if len(row) > 6 else 0) or 0,
            "printed_number": misprints.get(uid, 0),
            "leader_on_spouse": uid in leader_on_spouse,
        })

    kids_by_union, kids_by_mother = {}, {}
    for uid, mother, _father, child, _note in T.CHILDREN:
        if uid:
            kids_by_union.setdefault(uid, []).append(child)
        else:
            # Paternity not assignable on the plate (83-85), so the group hangs
            # off the mother's line alone.
            kids_by_mother.setdefault(mother, []).append(child)

    # UNATTACHED_BLOCKS: a couple the plate prints inside another couple's
    # child column with no leader stub -- placed on the page, but no descent
    # asserted. Keyed by the column it is printed in, since that is where
    # Chart.render needs it. Read with getattr: Tables 1 and 4 have none.
    unattached = {}
    for uid, primary, parent_uid, after, _note in getattr(T, "UNATTACHED_BLOCKS", []):
        unattached.setdefault(parent_uid, []).append((after, primary))

    return persons, unions, kids_by_union, kids_by_mother, unattached


def load():
    from openpyxl import load_workbook

    wb = load_workbook(XLSX, data_only=False)

    def rows(sheet):
        ws = wb[sheet]
        hdr = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip only fully blank rows. A blank first cell is meaningful:
            # CHILDREN rows whose paternity cannot be assigned carry no union_id.
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            yield dict(zip(hdr, row))

    persons = {}
    for r in rows("PERSONS"):
        persons[int(r["id"])] = {
            k: ("" if r.get(k) is None else str(r[k]).strip())
            for k in ("sex", "name_as_printed", "alt_name", "age", "clan", "vital_note",
                      "origin", "cross_ref", "plate_note", "english_name", "census_name",
                      "census_year", "match_confidence", "notes")
        }
        persons[int(r["id"])]["id"] = int(r["id"])
        # The workbook has no duplicate-number column; see load_baseline.
        persons[int(r["id"])]["plate_number"] = int(r["id"])
        persons[int(r["id"])]["generation"] = r.get("generation") or 0

    unions = [
        {
            "union_id": r["union_id"],
            "wife": int(r["wife_id"] or 0),
            "husband": int(r["husband_id"] or 0),
            "note": r.get("note") or "",
            "drawn_under": 0,
            "printed_number": 0,
        }
        for r in rows("UNIONS")
    ]

    kids_by_union, kids_by_mother = {}, {}
    for r in rows("CHILDREN"):
        uid = r.get("union_id") or ""
        child, mother = int(r["child_id"]), int(r["mother_id"])
        if uid:
            kids_by_union.setdefault(uid, []).append(child)
        else:
            kids_by_mother.setdefault(mother, []).append(child)

    # The workbook has no unattached-block sheet; it is a plate-layout fact and
    # lives in the transcription modules. See load_baseline.
    return persons, unions, kids_by_union, kids_by_mother, {}


def esc(s):
    return html.escape(str(s))


def dotted(s):
    """Add the plate's trailing period, unless the value already ends in one ('d.')."""
    s = esc(s)
    return s if s.endswith(".") else s + "."


# 'see above' / 'see below' / 'see elsewhere on this table' -- the directional
# phrase a plate uses instead of repeating a sibling bracket. Matched only to
# be wrapped in an anchor; the words themselves are never rewritten.
SEE_PHRASE = re.compile(r"\bsee (?:above|below|elsewhere on this table)\b", re.I)


def linkify_xref(text, persons, target=None, cross_plate=False):
    """
    Wrap person-number tokens in same-table anchors: '76' -> <a href="#p76">76</a>,
    '90-3' -> <a href="#p90">90-3</a>. `text` is already HTML-escaped.

    The directional phrase is linked too, so 'see above' is itself a jump and
    not just a printed instruction to go hunting. `target` is the id it points
    at, supplied by the caller that knows it structurally; where it is None the
    phrase takes the first person-number in its own text, which is what the
    phrase refers to ('see below, 76, 90-3' -> 76). A phrase with neither is
    left as plain text rather than pointed somewhere plausible.

    A cross-reference into another genealogy ('see Gen. II, 21, 74') is left
    entirely untouched: those numbers are another plate's numbering, and every
    href here is an anchor on THIS page. The visible text is identical either
    way -- only the markup around it changes.

    `cross_plate` is that same judgement made by the caller, and it exists
    because the guard below reads one ROW while the reference is a whole
    sentence. A long reference is split at the plate's own line break with '|'
    (see 160 and 169 on Genealogy II, and 155 on Genealogy III), and the row
    that carries the numbers need not be the row that carries the words 'Gen.'
    -- Parsons breaks 155 as 'see Gen. | II, 126, 158, 160'. Judged row by row
    those three numbers looked local and were linked to Genealogy III's own
    126, 158 and 160: three real people, none of them the ones referred to, and
    nothing on the page to reveal it. Callers that split a reference pass the
    verdict for the whole of it.
    """
    if cross_plate or "Gen." in text:
        return text

    found = []

    def repl(m):
        if int(m.group(1)) in persons:
            found.append(int(m.group(1)))
            return f'<a href="#p{m.group(1)}">{m.group(0)}</a>'
        return m.group(0)

    out = re.sub(r"(\d+)(?:-\d+)?", repl, text)

    dest = target if target in persons else (found[0] if found else None)
    if dest is not None:
        out = SEE_PHRASE.sub(
            lambda m: f'<a href="#p{dest}">{m.group(0)}</a>', out, count=1)
    return out


def wrap_line(inner, lead=False, anchor=None):
    """
    Wrap a line's spans in its row div.

    `lead` draws the plate's leader rule running from the end of the text to the
    child column -- 'F. Lupi. Sun (Navaho)————————'. Only the mother's line gets
    one, since that is the line the sibling bracket hangs off. The text is kept
    in its own .txt span so the surrounding flex context cannot collapse the
    spaces between fields.

    `anchor` is the citable id ("p13") a person's first printed occurrence
    carries, so #p13 is a stable URL for that line.
    """
    aid = f' id="{anchor}"' if anchor else ""
    if lead:
        return (f'<div class="line lead-line"{aid}><span class="txt">{inner}</span>'
                '<span class="lead"></span></div>')
    return f'<div class="line"{aid}>{inner}</div>'


def person_line(p, is_spouse, english_seen, printed_number=0):
    """
    The spans of one printed line: '7. F. Dziwaiʼᶦdyitsʼa. d. 1908. Sun'

    `printed_number` is the number the PLATE prints, on the one kind of line
    where that is not the person's own number: a misprint. The plate's number
    is what is shown -- this edition reproduces the plate and annotates it, it
    does not silently correct it -- while the link still resolves to the person
    the line actually names, and a marker points at the note that explains the
    discrepancy. Getting this backwards prints a corrected chart that disagrees
    with the scan beside it.
    """
    bits = []
    if is_spouse:
        bits.append('<span class="plus">+</span>')
    # The number is the plate's own citation apparatus, so it is a link: #p13
    # is the stable address of person 13's first printed line.
    if printed_number and printed_number != p["id"]:
        # The plate's number, ringed. data-printed lets the person card show the
        # same number when it is opened from THIS line -- the register keeps the
        # true one, so the card must be told, not left to guess.
        plus = bits.pop() if is_spouse else ""
        bits.append(f'<span class="sic-ring">{plus} '
                    f'<a class="num num-sic" data-printed="{printed_number}" '
                    f'href="#p{p["id"]}">{printed_number}.</a></span>')
    else:
        # href on the id (unique, addressable), text from plate_number (what the
        # plate prints). They differ only where the plate reuses a number.
        bits.append(f'<a class="num" href="#p{p["id"]}">{p["plate_number"]}.</a>')
    # Sex and clan take the same treatment the number does where the plate is
    # wrong: its value, ringed. The ring is a CLASS ON THE EXISTING SPAN, never
    # a wrapper -- the person card drops the header's .sex/.clan by reading a
    # direct child's className, and moves the .clan node itself into its badge,
    # so a wrapper would either leak the letter into the card title or strip the
    # ring off the badge. .sic-ring is an outline, so no row moves.
    printed_sex = str(p.get("printed_sex", "") or "")
    sex_cls = "sex sic-ring" if printed_sex else "sex"
    bits.append(f'<span class="{sex_cls}">{esc(printed_sex or p["sex"])}.</span>')

    name, alt = p["name_as_printed"], p["alt_name"]
    if name and alt:
        if "braced" in p["plate_note"]:
            nm = f'<span class="brace">{{</span>{esc(name)} / {esc(alt)}<span class="brace">}}</span>'
        else:
            nm = f'{esc(name)} <span class="alt">({esc(alt)})</span>'
    elif name:
        nm = esc(name)
    else:
        nm = '<span class="blank">———</span>'
    # kjq = Western Keres. Without this the names sit inside lang="en" and a
    # screen reader applies English phonics to Americanist transcription.
    lang = ' lang="kjq"' if name else ""
    bits.append(f'<span class="name"{lang}>{nm}.</span>')

    eng = p["english_name"]
    if eng and eng != alt:
        english_seen.add(p["id"])
        bits.append(f'<span class="eng">{esc(eng)}</span>')

    if p["age"]:
        bits.append(f'<span class="age">{dotted(p["age"])}</span>')
    if p["vital_note"]:
        bits.append(f'<span class="vital">{dotted(p["vital_note"])}</span>')

    clan_cls = "clan"
    printed_clan = str(p.get("printed_clan", "") or "")
    if p["clan"]:
        if printed_clan:
            clan_cls = "clan sic-ring"
        clan = (esc(printed_clan or p["clan"])
                + (f' ({esc(p["origin"])})' if p["origin"] else ""))
    elif p["origin"]:
        clan = f'of {esc(p["origin"])}'
        # NOT a clan: the plate names only where this person came from ("of
        # Zuñi", person 101 on Table 1 -- the single case in either table). The
        # card labels its badge "Clan: X", and that label must not be printed
        # over a value that is not one. Marked here, where the distinction is
        # still in the data, rather than sniffed out of the rendered string.
        clan_cls = "clan clan-origin"
    else:
        clan = ""
    if clan:
        bits.append(f'<span class="{clan_cls}">{clan}</span>')

    census = " ".join(x for x in (p["census_name"], p["census_year"]) if x)
    if census:
        bits.append(f'<span class="census" title="census match">{esc(census)}</span>')

    return " ".join(bits)


class Chart:
    def __init__(self, persons, unions, kids_by_union, kids_by_mother, unattached=None):
        self.P = persons
        self.U = unions
        self.KU = kids_by_union
        self.KM = kids_by_mother
        # {parent union id: [(after this child, primary of the block to splice)]}
        self.UB = unattached or {}
        self.rendered_unions = set()
        self.xref_printed = set()
        self.english_seen = set()
        self.placed = set()  # drawn as a descendant line; drives sibling-group dedupe
        self.seen = set()    # drawn anywhere, including as a '+' spouse line

    def unions_of(self, pid):
        # Every marriage this person is in, including ones already drawn from the
        # other spouse's side: the plate prints the '+' line in both places and
        # replaces the repeated sibling bracket with a cross-reference.
        return [u for u in self.U
                if pid in (u["wife"], u["husband"]) or u["drawn_under"] == pid]

    def render(self, pid, depth=0, unattached=False):
        """
        Draw one block (a person plus their '+' spouse lines) and its descendants.

        `unattached` marks a block the plate prints inside this child column
        without a leader stub -- see UNATTACHED_BLOCKS. It changes nothing but
        the node's class; the block itself is drawn exactly as any other.

        Vertical alignment follows the plate: a sibling bracket hangs off the
        MOTHER's line, not off the top of the block. Where the mother is the
        block's primary -- 66 with children 80-82 -- that is row 0. Where the
        primary is male and the mothers are his wives -- 8 with wives 7 and 73 --
        each group sits on its own wife's row. Verified against the plate at
        sources/parsons-1923-table-1.jpg.
        """
        p = self.P[pid]
        self.placed.add(pid)
        # The FIRST printed occurrence of a person -- primary line or '+' spouse
        # line, whichever the document reaches first -- carries the citable
        # id="p{n}". seen-order equals document order here: a block's own lines
        # are all marked before its child column is rendered.
        first = pid not in self.seen
        self.seen.add(pid)

        # Entries are (kind, content, anchor); the list index is the row, which
        # is what the leader rules and the child-column offsets are keyed on.
        block = [("line", person_line(p, False, self.english_seen),
                  f"p{pid}" if first else None)]
        shown = {pid}           # ids already printed in this block
        row = 0                # index of the last line written into the block
        groups = []            # (mother_row, kind, payload) -> the child column

        # A misprinted sex letter or clan on the primary's own line. Counted
        # with row += 1 exactly as the number misprint's annotation is, so every
        # row below stays a whole --lh and no bracket leaves its mother's line.
        # Printed once per person, on the line that carries the id.
        if first and (p.get("printed_sex") or p.get("printed_clan")):
            block.append(("sic", SIC_ROW, None))
            row += 1

        if p["cross_ref"] and pid not in self.xref_printed:
            self.xref_printed.add(pid)
            # The verdict is taken on the WHOLE reference, before the split:
            # the '|' is the plate's line break, not a change of subject.
            xp = "Gen." in p["cross_ref"]
            for part in p["cross_ref"].split("|"):
                block.append(("xref", linkify_xref(esc(part.strip()), self.P,
                                                   cross_plate=xp), None))
                row += 1

        # Pass 1: lay out this block's own lines and decide where each sibling
        # group belongs. No recursion here, so union bookkeeping resolves in the
        # same order it did before descendants are drawn.
        deferred_xref = []     # plate lines that stand in for an omitted union
        for u in self.unions_of(pid):
            drawn_before = u["union_id"] in self.rendered_unions
            # The plate prints no second occurrence of this marriage -- see
            # SECOND_VISIT_OMITTED. Print nothing for it here except the
            # cross-reference the plate puts in its place, held back until the
            # block's other union lines are down, because that is where the
            # plate sets it: under 183, not between 169 and him.
            if drawn_before and u["union_id"] in SECOND_VISIT_OMITTED:
                kids = self.KU.get(u["union_id"], [])
                deferred_xref.append((SECOND_VISIT_OMITTED[u["union_id"]],
                                      kids[0] if kids else None))
                continue
            self.rendered_unions.add(u["union_id"])
            if pid in (u["wife"], u["husband"]):
                other = u["husband"] if u["wife"] == pid else u["wife"]
            else:
                # A drawn_under union: the primary is not a partner, so the '+'
                # line to print is whichever partner this block has not shown.
                # Table 4 sets 6's second husband 7 under 5, because 6 is
                # already on the line above.
                other = u["husband"] if u["wife"] in shown else u["wife"]
            mother_row = row
            if other:
                sp = self.P[other]
                first_sp = other not in self.seen
                self.seen.add(other)
                shown.add(other)
                block.append(("line", person_line(sp, True, self.english_seen,
                                                  u["printed_number"]),
                              f"p{other}" if first_sp else None))
                row += 1
                mother_row = row
                # The misprint annotation is its own row, directly under the
                # line it annotates -- the same slot a cross-reference takes,
                # and counted the same way, so the rows below stay on the grid.
                if (u["printed_number"] and u["printed_number"] != other) or (
                        first_sp and (sp.get("printed_sex") or sp.get("printed_clan"))):
                    block.append(("sic", SIC_ROW, None))
                    row += 1
                if sp["cross_ref"] and other not in self.xref_printed:
                    self.xref_printed.add(other)
                    xp = "Gen." in sp["cross_ref"]
                    for part in sp["cross_ref"].split("|"):
                        block.append(("xref", linkify_xref(esc(part.strip()), self.P,
                                                           cross_plate=xp), None))
                        row += 1
            # The primary is the mother, so the bracket is her own row -- unless
            # the plate draws this marriage's leader from the '+' spouse's line,
            # which is what LEADER_ON_SPOUSE_ROW declares and what Genealogy III
            # does for a second husband. There mother_row is already the spouse's
            # row, set when his line went down.
            if u["wife"] == pid and not u.get("leader_on_spouse"):
                mother_row = 0

            kids = self.KU.get(u["union_id"], [])
            if kids and (drawn_before or all(k in self.placed for k in kids)):
                note = SECOND_VISIT_NOTE.get(u["union_id"], "see elsewhere on this table")
                # The note stands in for THESE children, already drawn on an
                # earlier row -- so the first of them is exactly where "see
                # above" points. Taken from the union, never from the English.
                groups.append((mother_row, "note", (note, kids[0]), u["union_id"]))
            else:
                new = [k for k in kids if k not in self.placed]
                if new:
                    groups.append((mother_row, "kids", new, u["union_id"]))

            # A fatherless group -- one the plate brackets under a mother
            # without saying which marriage it belongs to -- hangs off HER
            # line, and she is not always this block's primary. Genealogy II
            # brackets 116-118 under 48, who has two husbands and is printed
            # only as a '+' line under the first of them, 47. Looked up on the
            # primary alone (below) those three children were never drawn at
            # all: silently, because an undrawn person is not an error anywhere
            # in this file. Tables 1 and 4 never showed it -- Table 1's one
            # such group is under 68, who is a child and therefore a primary.
            # mother_row is already her row here, which is the row the bracket
            # must hang on; and `not in self.placed` is what keeps a mother
            # printed as a '+' in two blocks from drawing them twice.
            if other and u["wife"] == other:
                spouse_kids = [k for k in self.KM.get(other, [])
                               if k not in self.placed]
                if spouse_kids:
                    # No union id: this group is bracketed under the mother
                    # alone, so there is no column for an unattached block to
                    # be spliced into.
                    groups.append((mother_row, "kids", spouse_kids, ""))

        # The omitted marriages' cross-references, now that every '+' line this
        # block does print is down. Counted like any other xref row so the rows
        # below stay on the --lh grid; no group hangs off one.
        for text, tgt in deferred_xref:
            block.append(("xref", linkify_xref(esc(text), self.P, tgt), None))
            row += 1

        orphans = [k for k in self.KM.get(pid, []) if k not in self.placed]
        if orphans:
            groups.append((0, "kids", orphans, ""))

        # Pass 2: draw the child column, each group on its own mother's row.
        #
        # A group is as tall as the subtree it contains, so two multi-child
        # groups in one block cannot both start on their mother's line unless
        # the block's own lines are spread apart -- which is exactly what the
        # plate does. Table 4 sets 11 and 12 under 10 with three children each,
        # and prints 12 three lines below 11, not one. So when a group cannot
        # begin at its mother's row, the mother's line is pushed down to meet
        # it, and every line after it follows in normal flow.
        groups.sort(key=lambda g: g[0])
        col = []
        child_cursor = 0     # next free row in the child column
        pushed = 0           # rows inserted into the block so far
        line_pad = {}        # block line index -> extra rows before that line
        for mother_row, kind, payload, uid in groups:
            target = mother_row + pushed
            if child_cursor > target:
                delta = child_cursor - target
                line_pad[mother_row] = line_pad.get(mother_row, 0) + delta
                pushed += delta
                target = child_cursor
            gap = target - child_cursor
            if kind == "note":
                note_text, note_target = payload
                inner = ('<div class="node"><div class="block">'
                         '<div class="xref xref-cell">'
                         f'{linkify_xref(esc(note_text), self.P, note_target)}'
                         '</div></div></div>')
                rows = 1
            else:
                # An unattached block is spliced in after the child the plate
                # prints it below, and counts its rows like any sibling -- it
                # occupies the column, so everything under it must move down.
                # Only the leader stub is withheld, by the node's class.
                splice = dict(self.UB.get(uid, []))
                parts = []
                for k in payload:
                    parts.append(self.render(k, depth + 1))
                    if k in splice:
                        parts.append(self.render(splice[k], depth + 1, unattached=True))
                inner = "".join(h for h, _ in parts)
                rows = sum(r for _, r in parts)
            style = f' style="margin-top:calc(var(--lh) * {gap})"' if gap else ""
            col.append(f'<div class="kids"{style}>{inner}</div>')
            child_cursor = target + rows

        # A leader rule goes on every row that a sibling group hangs off, i.e.
        # each mother's line -- whether she is the primary (66) or a '+' spouse
        # (12, 17, 57). Same row index the child column is offset by.
        lead_rows = {g[0] for g in groups}
        lines = []
        for i, (kind, content, anchor) in enumerate(block):
            if kind == "line":
                html_line = wrap_line(content, lead=(i in lead_rows), anchor=anchor)
            elif kind == "sic":
                html_line = f'<div class="sic-row">{content}</div>'
            else:
                html_line = f'<div class="xref">{content}</div>'
            if line_pad.get(i):
                html_line = html_line.replace(
                    "<div ", f'<div style="margin-top:calc(var(--lh) * {line_pad[i]})" ', 1)
            lines.append(html_line)

        # Leaf blocks need no fixed width: nothing hangs off them, so a long
        # generation-5 entry can run past the column instead of widening the sheet.
        node_class = "node" if col else "node leaf"
        if unattached:
            node_class += " unattached"
        out = [f'<div class="{node_class}">', '<div class="block">', *lines, "</div>"]
        if col:
            out.append('<div class="kidcol">' + "".join(col) + "</div>")
        out.append("</div>")  # .node
        # Height in rows, so the caller can stack sibling groups without
        # assuming every group is one line tall.
        return "".join(out), max(len(block) + pushed, child_cursor)


def font_css():
    """
    @font-face rules with the subset woff2 base64-inlined, or "" if absent.

    Inlining keeps the page a single self-contained file -- no external request,
    nothing to break if the fonts folder is ever moved, and archivable as one
    document. ~35 kB for both faces.

    The face is named "Laguna Serif" rather than Gentium: the OFL reserves the
    latter for unmodified releases, and a subset is a modification. See
    scripts/subset_font.py.
    """
    faces = []
    for style, weight in (("regular", "normal"), ("italic", "italic")):
        path = FONT_DIR / f"laguna-serif-{style}.woff2"
        if not path.exists():
            print(f"  note: {path.name} missing -- falling back to system fonts. "
                  "Run scripts/subset_font.py to embed it.")
            return ""
        b64 = base64.b64encode(path.read_bytes()).decode("ascii")
        faces.append(
            "@font-face{font-family:'Laguna Serif';font-style:%s;font-weight:400;"
            "font-display:swap;src:url(data:font/woff2;base64,%s) format('woff2')}"
            % (weight, b64)
        )
    return "\n".join(faces) + "\n"


CSS = """
/* ---- tokens ------------------------------------------------------------ */
/* One palette, both themes. Custom-property declarations are never dropped at
   parse time (any token stream is valid), so a static-then-light-dark() pair
   would NOT fall back -- the second declaration always wins and turns
   invalid-at-computed-value-time where light-dark() is unsupported. Hence the
   structure below: static light tokens unconditionally; the light-dark()
   pairs only inside @supports; a static dark set (via the media query and
   [data-theme]) for engines without light-dark().
   The [data-theme] color-scheme flips are all the JS theme control needs.

   THE DEFAULT IS LIGHT, and it is set HERE rather than in the script -- set by
   the user 2026-08-10. `color-scheme:light` at :root is what makes light-dark()
   resolve to the light half of every pair, so a reader on a dark OS gets the
   light palette with the script dead, blocked or still parsing. Dark is now
   reachable only through [data-theme="dark"], i.e. only by pressing the
   control. Do NOT reintroduce a prefers-color-scheme palette block: it would
   restore exactly the OS-follows-you behaviour this replaced, and it would do
   it only in the no-JS case, which is the one nobody looks at. */
:root{color-scheme:light}
:root[data-theme="light"]{color-scheme:light}
:root[data-theme="dark"]{color-scheme:dark}
:root{
  --paper:#FAF8F4; --panel:#FFFDF9; --ink:#1C1A17; --muted:#635D53;
  /* plate rules: >=3:1 against --panel in BOTH themes (non-text minimum) */
  --rule:#7D766B;
  /* chrome hairlines only -- never the plate's own rules */
  --rule-faint:#C4BFB4;
  --accent:#7A5C1E; --accent-strong:#5C450F; --ink-lineage:#3A342A;
  /* The misprint red. Marks the plate's error where it occurs and nowhere
     else, so it must clear 4.5:1 on BOTH papers on its own -- it is text,
     and it is the only colour on a table page that is not --ink. Measured
     6.16:1 on #FAF8F4 and 7.84:1 on #191713. */
  --sic:#B3261E;
  /* The clan. One colour for the whole field, NOT one per clan -- a 13-clan
     palette was built and reverted (it collapsed to about one just-noticeable
     difference under deuteranopia) and this is not that decision re-opened.
     Two colours have to be told apart here, not thirteen, and they differ in
     lightness as well as hue, so the distinction survives any colour vision.
     It is text, so it clears 4.5:1 on its own against every background it can
     sit on: paper 5.86:1 light / 9.53:1 dark, panel 6.12 / 8.74, selected row
     6.22 / 10.40. The values coincide with --accent -- this is the gold the
     clan was set in before body.chart flattened the page to --ink -- but the
     token is separate on purpose: --accent means "interactive" everywhere
     else, and recolouring the chrome must not recolour the genealogy. */
  --clan:#7A5C1E;
  --wash:#F4E6CA; --shadow:rgba(0,0,0,.07);
  /* Selected row. Moves AWAY from --paper rather than toward it, so the row
     that is picked out is the one where the text reads BEST, not worst: the
     clan gold goes 5.86:1 -> 6.22:1 on light and 9.53:1 -> 10.4:1 on dark,
     where the old --wash tint dropped it to 5.04:1. The highlight itself is
     therefore carried by the leading rule and ring, which are non-text and
     clear 8.5:1, not by a background wash competing with the type. */
  --sel-bg:#FFFFFF;
  /* research chips -- private build only; class names are the leak check's markers */
  --eng-bg:#E8DFC8; --eng-fg:#4A3A12; --cen-bg:#DCE6EF; --cen-fg:#22384C;
  /* two type voices: the plate speaks 1923, the chrome speaks 2026 */
  --font-plate:"Laguna Serif","Iowan Old Style","Palatino Linotype",Palatino,Georgia,serif;
  --font-ui:system-ui,-apple-system,"Segoe UI",Roboto,"Helvetica Neue",Arial,sans-serif;
  /* --tap floors every hit area in the site chrome. 2rem clears the 24px
     WCAG 2.5.8 minimum with room to spare on a mouse; a coarse pointer gets
     the full 44px below. --bar-h is derived so anchor scroll-margin, which
     is keyed on it, tracks the bar instead of being restated. */
  /* The real --muted, captured at :root. body.chart redefines --muted to
     --ink for the table pages, and a var() is substituted with the value the
     element it is declared on computes -- so anything that must keep the
     dimmer grey THROUGH that flatten reads --muted-fixed instead. Declared
     once here: every theme block below only ever restates --muted, and this
     picks up whichever of those wins. */
  --muted-fixed:var(--muted);
  --tap:2rem; --bar-h:calc(var(--tap) + var(--s2) * 2);
  /* Type ramp for the CHROME AND PROSE ONLY. Nothing inside the sheet is
     sized from these: the plate's metrics are rem against the fixed 16px
     root (see GEOM), so a size change there would move --col and --stub and
     break the zero-drift invariant. --t-xs is a floor, not a default -- 13px
     is as small as the uppercase, letterspaced labels are allowed to get.
     --t-prose is the only fluid step; it grows on wide screens where the
     measure can afford it. */
  --t-xs:.8125rem; --t-sm:.875rem; --t-base:1rem; --t-md:1.0625rem; --t-lg:1.125rem; --t-xl:1.375rem;
  --t-prose:clamp(1rem,.95rem + .28vw,1.125rem);
  --measure:40rem; --measure-wide:72rem;
  --s1:.25rem; --s2:.5rem; --s3:.75rem; --s4:1rem; --s5:1.5rem; --s6:2.5rem; --s7:4rem;
}
@media (pointer:coarse){:root{--tap:2.75rem}}
@supports (color: light-dark(#000,#fff)){
  :root{
    --paper:light-dark(#FAF8F4,#191713);
    --panel:light-dark(#FFFDF9,#221F1B);
    --ink:light-dark(#1C1A17,#E9E6DF);
    --muted:light-dark(#635D53,#A49E93);
    --rule:light-dark(#7D766B,#756F66);
    --rule-faint:color-mix(in oklab,var(--rule) 45%,var(--paper));
    --accent:light-dark(#7A5C1E,#DBB970);
    --accent-strong:light-dark(#5C450F,#DBB970);
    --sic:light-dark(#B3261E,#FF8A80);
    --clan:light-dark(#7A5C1E,#DBB970);
    --ink-lineage:light-dark(#3A342A,#C9C0AF);
    --wash:light-dark(#F4E6CA,#2E250D);
    --shadow:light-dark(rgba(0,0,0,.07),rgba(0,0,0,.45));
    --eng-bg:light-dark(#E8DFC8,#3A3016);
    --eng-fg:light-dark(#4A3A12,#F0DFAE);
    --cen-bg:light-dark(#DCE6EF,#1D2E3D);
    --cen-fg:light-dark(#22384C,#BCD7EE);
    --sel-bg:light-dark(#FFFFFF,#0E0C09);
  }
}
/* Engines without light-dark() (Firefox ESR 115, Safari <=17.4): the same
   dark set, statically. Kept in step with the pairs above -- change one,
   change both. Keyed on [data-theme="dark"] ALONE: there is no
   prefers-color-scheme branch here any more, because the OS no longer selects
   the palette on either kind of engine. The static light tokens at :root are
   the default, and this is the only thing that overrides them. */
@supports not (color: light-dark(#000,#fff)){
  :root[data-theme="dark"]{
    --paper:#191713; --panel:#221F1B; --ink:#E9E6DF; --muted:#A49E93;
    --rule:#756F66; --rule-faint:#423F38; --accent:#DBB970;
    --accent-strong:#DBB970; --ink-lineage:#C9C0AF; --wash:#2E250D;
    --sic:#FF8A80; --clan:#DBB970;
    --shadow:rgba(0,0,0,.45); --eng-bg:#3A3016; --eng-fg:#F0DFAE;
    --cen-bg:#1D2E3D; --cen-fg:#BCD7EE;
    --sel-bg:#0E0C09;
  }
}
@media (prefers-contrast:more){:root{--muted:var(--ink);--rule:var(--muted)}}

/* ---- base -------------------------------------------------------------- */
*{box-sizing:border-box}
/* 'Laguna Serif' is the embedded Gentium subset (see font_css). It leads the
   stack because it is the only face guaranteed to carry the phonetic modifier
   letters ᶦ ᵘ ᵃ; the rest are fallbacks for a build made without it.
   Root size is fixed at 16px: the chart's metrics were measured against it. */
body{margin:0;background:var(--paper);color:var(--ink);
  font:16px/1.55 var(--font-plate)}
:focus-visible{outline:2px solid var(--accent-strong);outline-offset:2px}
.visually-hidden{position:absolute;width:1px;height:1px;margin:-1px;padding:0;
  overflow:hidden;clip-path:inset(50%);white-space:nowrap;border:0}
.skip{position:fixed;inset-inline-start:-100vw;inset-block-start:var(--s3);
  z-index:100;background:var(--paper);color:var(--ink);
  border:1px solid var(--rule);padding:.5rem 1rem;
  font:600 .8125rem/1 var(--font-ui);text-decoration:none}
.skip:focus{inset-inline-start:var(--s3)}

/* ---- masthead (sticky site chrome) ------------------------------------- */
/* The bar's own inline padding is pulled in by --s2, the same amount every
   control adds back as hit area, so the enlarged targets stay optically
   flush with the page edge instead of the chrome appearing to widen. */
.masthead{position:sticky;inset-block-start:0;z-index:30;min-height:var(--bar-h);
  display:flex;align-items:center;flex-wrap:wrap;gap:var(--s1) var(--s2);
  padding:var(--s2) calc(var(--s4) - var(--s2));background:var(--paper);
  border-block-end:1px solid var(--rule-faint);
  font:400 var(--t-xs)/1.2 var(--font-ui);letter-spacing:.08em}
/* One hit-area rule for every control in the bar: links included, since a
   roman numeral is only a few px of glyph and was the worst target here. */
.wordmark,.masthead nav a,.mast-btn{
  display:inline-flex;align-items:center;min-height:var(--tap);
  padding-inline:var(--s2);vertical-align:middle}
.wordmark{text-transform:uppercase;letter-spacing:.14em;font-weight:600;
  color:var(--ink);text-decoration:none}
a.wordmark:hover{color:var(--accent)}
a.wordmark:focus-visible{color:var(--accent)}
.masthead nav{display:flex;flex-wrap:wrap;gap:var(--s1);
  color:var(--muted);text-transform:uppercase}
/* Table links are buttons, sharing the 2px radius of the Theme and Scale
   controls rather than inventing a second button shape. The current page is
   a filled inversion, not a colour shift, so "which table am I on" survives
   both themes, a monochrome screen and colour-blind vision. aria-current is
   the only source of truth for that state -- there is no parallel class to
   fall out of step with it. */
/* --s2 rather than --s3 inline: at 375px the two pills plus the Theme button
   overran the row by 2.9px, and the tighter padding buys ~17px of headroom
   so the bar settles at two rows instead of three. Targets stay far past the
   24px minimum -- the labels alone are ~100px wide. */
/* The pills carry the NUMERAL ALONE at every width. "Genealogy " is still in
   the markup and still in the accessible name -- hidden here, not removed, so
   the link is still called "Genealogy I" by ear. --tap squares the pill, since
   a bare numeral is a few px of glyph and was the worst target in this bar.
   No gap is needed: the word is out of flow, so it is not a flex item. */
.masthead nav a{justify-content:center;padding-inline:var(--s2);
  min-width:var(--tap);
  border:1px solid var(--rule-faint);border-radius:2px;
  color:var(--muted);text-decoration:none;white-space:nowrap}
.masthead nav .nav-word{position:absolute;width:1px;height:1px;margin:-1px;
  padding:0;overflow:hidden;clip-path:inset(50%);white-space:nowrap;border:0}
.masthead nav a:hover{color:var(--ink);border-color:var(--rule)}
.masthead nav a:focus-visible{color:var(--ink);border-color:var(--rule)}
.masthead nav a[aria-current="page"]{background:var(--ink);color:var(--paper);
  border-color:var(--ink);font-weight:600}
/* Below ~26rem the row is still too tight for Search's word beside four
   pills, so the SAME hide-the-word mechanism runs there -- scoped to
   .mast-right, because the nav's copy of it is now unconditional above and
   the two must be able to move independently. Hidden VISUALLY, not removed:
   the accessible name stays "Search", and the glyph rule below gives it a
   visible mark at that width. */
@media (max-width:26rem){
  .mast-right .nav-word{position:absolute;width:1px;height:1px;margin:-1px;
    padding:0;overflow:hidden;clip-path:inset(50%);white-space:nowrap;border:0}
}
.mast-right{margin-inline-start:auto;display:flex;align-items:center;gap:var(--s2)}
.mast-btn{font:inherit;letter-spacing:inherit;text-transform:uppercase;
  color:var(--muted);background:none;border:1px solid var(--rule-faint);
  border-radius:2px;cursor:pointer}
/* Search is an <a> wearing the button shape -- the rule above was written for
   <button>, which has no underline to clear. */
a.mast-btn{text-decoration:none}
/* The glyph is the narrow-screen half of the Search label; above the
   breakpoint the word carries it and the glyph would be a second mark saying
   the same thing. Pills fall back to a --tap square there, and so does this. */
.mast-glyph{display:none}
@media (max-width:26rem){
  .mast-glyph{display:block}
  a.mast-btn{min-width:var(--tap);justify-content:center}
}
.mast-btn:hover{color:var(--ink);border-color:var(--rule)}
.mast-btn:focus-visible{color:var(--ink);border-color:var(--rule)}

/* ---- the theme control, at the foot ------------------------------------- */
/* Moved out of the masthead 2026-08-10. It keeps .mast-btn's shape and the
   --tap floor -- it is site chrome wherever it stands -- and only needs the
   bar's own font, which it no longer inherits down here. Right-aligned on the
   same rail as the text above it, so it reads as the end of the page rather
   than as a stray control; the block-start rule closes the page off under the
   folded apparatus. */
.theme-foot{display:flex;justify-content:flex-end;
  max-width:var(--measure-wide);margin:0 auto;
  padding:var(--s4) var(--s5) var(--s6);
  border-block-start:1px solid var(--rule-faint);
  font:400 var(--t-xs)/1.2 var(--font-ui);letter-spacing:.08em}
.theme-foot .mast-btn{display:inline-flex;align-items:center;
  min-height:var(--tap);padding-inline:var(--s3)}

/* ---- title page -------------------------------------------------------- */
.titlepage{max-width:var(--measure);margin:0 auto;
  padding:var(--s7) var(--s5) var(--s5);text-align:center}
.plate-label{font-variant:small-caps;letter-spacing:.22em;font-size:var(--t-xs);
  color:var(--muted)}
h1{font-size:clamp(1.6rem,1.15rem + 1.9vw,2.5rem);font-weight:400;
  letter-spacing:.09em;line-height:1.15;margin:.35rem 0 0;text-wrap:balance}
.rule-double{width:8rem;height:4px;margin:var(--s4) auto;
  border-block-start:2px solid var(--ink);border-block-end:1px solid var(--ink)}
.cite{font-size:var(--t-base);color:var(--muted);line-height:1.65;
  text-wrap:pretty}
/* The table page's statistics line. --muted-fixed, not --muted: this is the
   one thing on a table page that keeps the landing page's grey through
   body.chart's flatten to --ink, so the two pages read as one edition. A step
   larger than the landing page's .c-stats, which is --t-sm. */
.imprint{margin-block-start:var(--s3);font-variant:small-caps;
  letter-spacing:.14em;font-size:var(--t-base);color:var(--muted-fixed);
  line-height:1.6}
/* A table page's title block no longer holds the citation, so it no longer
   needs the 40rem prose measure -- and at --t-base the statistics line does
   not fit inside one. The landing page keeps --measure, because its citation
   still has to read as prose. */
body.chart .titlepage{max-width:var(--measure-wide)}

/* ---- plate bar --------------------------------------------------------- */
/* Find sits hard left, scale hard right, spanning the bar. The push comes from
   an auto margin on #scale-mount rather than space-between, because #find is
   [hidden] until the script unhides it: with space-between a no-JS reader would
   get the scale buttons stranded on the left. An auto start-margin puts them
   right whether or not find is in the row. */
/* The bar rides the PLATE's rail, not the prose column: no max-width, and the
   same var(--s5) inline padding .scroll carries, so Find lands exactly on the
   sheet's left edge and Scale on its right. It was centred at --measure-wide
   before, which matched the title block's BOX but nothing the reader can see --
   the statistics line inside that box is centred text and sits ~270px further
   in, so the bar's ends lined up with thin air. Controls belong to the thing
   they control; if this ever moves back to a measure, it has to move with
   .scroll's padding or the two rails part again. */
.plate-bar{padding:0 var(--s5) var(--s3);display:flex;
  align-items:baseline;gap:var(--s3) var(--s4);flex-wrap:wrap}
.plate-tools{display:flex;align-items:center;gap:var(--s3);
  font-family:var(--font-ui);flex-wrap:wrap;inline-size:100%}
/* :not([hidden]) -- an unconditional display would defeat the hidden
   attribute (author display beats the UA's [hidden]{display:none}) and ship
   a dead search form to no-JS readers. */
#find:not([hidden]){display:flex;align-items:center;gap:var(--s2)}
/* 16px on the find field, not --t-xs: anything smaller and iOS Safari zooms
   the viewport on focus, which on a horizontally scrolling plate is a nasty
   way to lose your place. */
#find input{font:var(--t-base) var(--font-ui);color:var(--ink);
  background:var(--panel);border:1px solid var(--rule);border-radius:2px;
  padding:.35rem .6rem;width:17rem;max-width:60vw;min-height:var(--tap)}
#find input:focus-visible{outline:2px solid var(--accent-strong);
  outline-offset:1px}
.find-note{font:var(--t-xs) var(--font-ui);color:var(--muted)}
#scale-mount{display:flex;align-items:center;gap:var(--s1);
  margin-inline-start:auto}
.scale-l{font:var(--t-xs) var(--font-ui);text-transform:uppercase;
  letter-spacing:.08em;color:var(--muted);margin-inline-end:var(--s1)}
.scale-btn{font:var(--t-xs) var(--font-ui);color:var(--muted);background:none;
  border:1px solid var(--rule-faint);border-radius:2px;padding:.3rem .55rem;
  cursor:pointer;min-height:var(--tap);min-width:var(--tap)}
.scale-btn[aria-pressed="true"]{color:var(--ink);border-color:var(--rule)}
.scale-btn:hover{color:var(--ink)}
.scale-btn:focus-visible{color:var(--ink)}

/* ---- the plate: shell, scroller, ruler, sheet -------------------------- */
.plate{margin:0}
/* The shell owns the edge fades as overlays, so nothing about the scroller's
   own layout -- the container the chart invariants live in -- changes. */
.scroll-shell{position:relative;timeline-scope:--plate-x}
.scroll-shell::before,.scroll-shell::after{content:"";position:absolute;
  inset-block:0 var(--s6);width:2rem;pointer-events:none;z-index:5}
.scroll-shell::before{inset-inline-start:0;visibility:hidden;
  background:linear-gradient(to right,var(--paper),transparent)}
.scroll-shell::after{inset-inline-end:0;opacity:.5;
  background:linear-gradient(to left,var(--paper),transparent)}
/* The region is focusable so a keyboard user can scroll the wide chart with the
   arrow keys; without that the later generations are unreachable. Give the
   focus ring room so it is not clipped by the scroll container. */
.scroll{overflow-x:auto;padding:var(--s3) var(--s5) var(--s6);
  -webkit-overflow-scrolling:touch;overscroll-behavior-x:contain;
  scrollbar-color:var(--rule) transparent;scrollbar-gutter:stable;
  scroll-timeline:--plate-x inline}
.scroll:focus-visible{outline:2px solid var(--accent);outline-offset:-4px}
/* Reduction wrapper: the scale control and the print fit both zoom HERE, so
   the ruler and the sheet always scale together and the column grid survives
   scaling by construction. */
.plate-zoom{display:inline-block;min-width:100%;zoom:var(--plate-zoom,1)}
/* Generation ruler: editorial apparatus OUTSIDE the sheet's frame. Its labels
   are sized from the same --col/--stub/--sheet-pad tokens as the grid, so
   alignment is exact by construction and the 0px-drift check covers it.
   It pans with the plate; viewport-pinned identity is the masthead's job
   (position:sticky cannot pin vertically inside a horizontal scroller). */
/* Two bands, not one: the identity chip rides at the block-start, the labels
   sit at flex-end. The height is what separates them, and it is the fix for a
   real collision -- the chip is pinned to the inline start, so whichever label
   the reader has panned to that edge sat directly underneath it, and the chip's
   opaque fill ate the first half of the word ("GENERA|TION 2"). Nothing here
   touches .gen widths, so the ruler still measures the same column grid. */
.ruler{display:flex;align-items:flex-end;width:max-content;min-width:100%;
  height:3.4rem;margin-block-end:var(--s2);
  padding-inline-start:calc(var(--sheet-pad) + 1px);
  font:600 var(--t-xs)/1.9 var(--font-ui);text-transform:uppercase;
  letter-spacing:.08em;color:var(--muted);
  border-block-end:1px solid var(--rule-faint);position:relative}
.ruler .gen{flex:none;width:var(--col);margin-inline-end:var(--stub);
  border-inline-start:1px solid var(--rule-faint);padding-inline-start:.5rem}
.ruler .gen:last-child{width:auto;margin-inline-end:0}
/* Identity chip: a zero-width sticky slot on the INLINE axis (which genuinely
   scrolls), so it can never displace a label. Invisible at rest; fades in over
   the first 8% of pan where scroll-driven animations exist. */
/* min-width:0 defeats the flex item's min-width:auto -- without it the slot
   takes the chip's content width and displaces every label by that much. */
.ruler-chipslot{flex:0 0 0px;min-width:0;overflow:visible;position:sticky;
  inset-inline-start:var(--s3);z-index:1;align-self:flex-start}
/* 1.5 rather than the labels' 1.9: the chip is a lozenge, so its line-height
   is its height, and every 0.1 here is 1.3px the ruler has to reserve. */
.ruler-chip{display:inline-block;white-space:nowrap;padding:.05rem .55rem;
  border:1px solid var(--rule-faint);border-radius:999px;background:var(--paper);
  color:var(--ink);font:400 var(--t-xs)/1.5 var(--font-plate);text-transform:none;
  letter-spacing:.02em;opacity:0}
/* Progress hairline under the ruler: inks left-to-right as the plate pans. */
.ruler::after{content:"";position:absolute;inset-inline:0;inset-block-end:-1px;
  height:2px;background:var(--accent);transform-origin:0 50%;transform:scaleX(0)}
@media (prefers-reduced-motion:no-preference){
  @supports (animation-timeline:scroll()){
    .ruler-chip{animation:chip-in linear both;animation-timeline:--plate-x;
      animation-range:0% 8%}
    .ruler::after{animation:plate-progress linear both;animation-timeline:--plate-x}
    .scroll-shell::before{visibility:visible;opacity:0;
      animation:fade-in linear both;animation-timeline:--plate-x;animation-range:0% 6%}
    /* Base 0, first keyframe 1: when the plate does not overflow the timeline
       is INACTIVE, the animation applies nothing, and the base 0 hides the
       fade -- a raised base would leave it stuck over a non-scrolling plate. */
    .scroll-shell::after{opacity:0;
      animation:fade-right linear both;animation-timeline:--plate-x;animation-range:94% 100%}
  }
}
@keyframes chip-in{to{opacity:1}}
@keyframes plate-progress{to{transform:scaleX(1)}}
@keyframes fade-in{to{opacity:1}}
@keyframes fade-right{from{opacity:1}to{opacity:0}}
/* The sheet: a flat matted plate. Layout rules below are byte-identical to the
   verified originals -- skin only may change. */
.sheet{display:inline-block;min-width:100%;background:var(--panel);
  border:1px solid var(--rule);padding:2rem var(--sheet-pad)}
.tree + .tree{margin-top:3.2rem;padding-top:2.4rem;border-top:1px dashed var(--rule)}
.node{display:flex;align-items:flex-start}
/* THE COLUMN GRID. Every .node is [.block][.kidcol] and every nested node adds
   exactly one --stub of padding plus one block, so a fixed block width puts
   generation d at d x (--col + --stub) on every path -- which is what makes the
   two families share one set of columns, as they do on the plate. --col is the
   single tunable: it must exceed the widest line at generations 1-4 (measured
   339px) or that line will spill into the next column.
   Vertical padding must stay 0: every line is exactly one --lh tall, which is
   what lets a sibling group be offset onto its mother's row by whole rows. */
.block{white-space:nowrap;padding:0;width:var(--col)}
.node.leaf > .block{width:auto}
/* A ROW'S HEIGHT IS STATED, NOT INFERRED FROM ITS LINE BOX. This is the whole
   of the sibling-bracket alignment, and line-height alone does not carry it.
   `--lh` is 1.55rem = 24.8px, and every vertical offset in the chart is a
   multiple of it: .kids gets margin-top:calc(var(--lh) * N) to sit on its
   mother's row, and line_pad puts the same calc on a pushed line. Those margins
   are lengths, so an engine keeps them to LayoutUnit precision -- 24.796875px.
   A LINE BOX is not a length, and WebKit quantises it to a whole pixel:
   measured in Safari 26.3 on 2026-08-10, every .line/.xref/.sic-row box came
   back 24.000px against a declared line-height of 24.799999px, while the same
   page's .kids margins were 24.796875px. So each row of offset lost 0.796875px,
   and it accumulates down the tree -- 69 of 141 brackets off, the worst
   -20.016px (25 x 0.797) at III's 21 -> 74, with the sign set by whether the
   mismatched margin sat on the group or on a line_pad push inside the block.
   Chromium reported 24.797px for both and was clean at 0.003px, which is why
   this was invisible here for the whole life of the edition; it is what a
   reader in Safari sees as a break in the leader rule at III 113 -> 204, where
   an only child leaves no bracket vertical to bridge the step.
   Stating the height makes the box exactly --lh in both engines whatever the
   line box inside it does. Do not replace it with line-height alone again --
   .block is white-space:nowrap, so a .line can never need to grow. */
.line{line-height:var(--lh);height:var(--lh);
  scroll-margin-block:calc(var(--bar-h) + 1.5rem) 1rem;
  scroll-margin-inline:var(--stub)}
/* Selected row. There is exactly ONE selected row, and which mechanism draws
   it depends on whether the card script is running:
     - card script live (html[data-card], set in the popoverOK block):
       .is-selected only. markSelected() is then the single owner of the
       highlight, so it can be cleared. :target CANNOT be cleared -- the hash
       survives every click -- so leaving it live left a second row lit after
       the reader followed a relation link out of a card and clicked away,
       until they left the page and came back.
     - no card script: :target, exactly as before. A #p anchor is then the only
       way a row is ever selected, and it must still light up.
   Background, box-shadow and outline only: a padding or border change here
   would move the row and break the bracket's mother_row alignment.
   The signal is carried by the leading rule and the ring, both
   --accent-strong at ~8.5:1 against the page, because the BACKGROUND has
   almost no room to move -- see --sel-bg.
   Every part of it is drawn OUTSIDE the border box: an inset shadow put the
   bar on top of the first glyphs, and a hugging outline sat on the text. The
   leading rule is a shadow offset -.3rem with no spread, the halo a .3rem
   spread behind it, the ring an outline at a matching offset. Nothing here
   changes layout, so the row still cannot move off its mother_row. */
.line.is-selected,html:not([data-card]) .line:target{background:var(--sel-bg);
  box-shadow:-.3rem 0 0 0 var(--accent-strong),0 0 0 .3rem var(--sel-bg);
  outline:2px solid var(--accent-strong);outline-offset:.3rem}
/* The whole row is the click target for the card (see rowClick). Hover is
   BACKGROUND ONLY: any padding, border or height change here would move the
   row and break the bracket's mother_row alignment. :has() is the feature
   test -- where it is unsupported the row simply keeps the default cursor,
   and the number link still works. */
.line:has(a.num){cursor:pointer}
.line:has(a.num):hover{background:color-mix(in oklab,var(--ink) 7%,transparent)}
/* The plate's leader rule, filling the gap from the end of the mother's text to
   the child column: '1. F. Lupi. Sun (Navaho)————————3. F. Nayowʼ˙ăitsa. Sun'.
   It ends exactly at --col, where the child's bracket bar begins, so the two
   read as one continuous rule. */
.lead-line{display:flex;align-items:baseline}
.lead-line > .txt{white-space:nowrap}
.lead{flex:1 1 auto;min-width:1.2rem;align-self:stretch;position:relative}
.lead::before{content:"";position:absolute;left:.5rem;right:0;top:50%;
  border-top:1px solid var(--rule)}
/* One .kids group per sibling bracket, stacked so each can be offset onto the
   row of its own mother -- see the docstring on Chart.render. */
.kidcol{display:flex;flex-direction:column}
.kids{display:flex;flex-direction:column}
.kids > .node{position:relative;padding-left:var(--stub)}
.kids > .node::before{content:"";position:absolute;left:0;width:var(--stub);
  top:calc(var(--lh)/2);border-top:1px solid var(--rule)}
.kids > .node::after{content:"";position:absolute;left:0;border-left:1px solid var(--rule)}
.kids > .node:first-child::after{top:calc(var(--lh)/2);bottom:0}
.kids > .node:last-child::after{top:0;height:calc(var(--lh)/2)}
.kids > .node:not(:first-child):not(:last-child)::after{top:0;bottom:0}
.kids > .node:only-child::after{display:none}
/* A block the plate prints INSIDE this column with no leader stub joining it
   to the bracket -- see UNATTACHED_BLOCKS. Genealogy II sets 31+32 between
   9+10's children 29 and 33 at exactly the children's indent, and draws no
   stub: the vertical passes the row. So ::before goes and ::after stays, and
   the indent is untouched -- withholding the padding too would move the block
   out of the column the plate puts it in, which is the error this replaced.
   It can never be first or last in a column (self_check() forbids it), so the
   :first-child / :last-child terminus rules above are unaffected. */
.kids > .node.unattached::before{display:none}
/* Lineage inking: hovering or focusing within a block darkens the rules a
   reader would trace with a finger -- its own bracket and the brackets it
   hangs. Color-only, on the existing 1px borders; the plate is never dimmed.
   Engines without :has() simply read the chart as before. */
.kids > .node:has(.block:hover,.block:focus-within)::before,
.kids > .node:has(.block:hover,.block:focus-within)::after{
  border-color:var(--ink-lineage)}
.node:has(> .block:hover,> .block:focus-within) > .block .lead::before,
.node:has(> .block:hover,> .block:focus-within) > .kidcol > .kids > .node::before,
.node:has(> .block:hover,> .block:focus-within) > .kidcol > .kids > .node::after{
  border-color:var(--ink-lineage)}
@media (prefers-reduced-motion:no-preference){
  .kids > .node::before,.kids > .node::after,.lead::before{
    transition:border-color .12s ease}
}
.num{color:var(--muted);font-variant-numeric:tabular-nums}
/* A little air after the number's point, so "65." reads as the entry's label
   rather than as the first word of the name. A margin, not a wider space
   character: it is independent of the font and cannot be swallowed by white-
   space collapsing. Layout-safe -- the widest line on Table 1 ("+ 61. M.
   Ishŭrneai (Garsia). Chaparral Cock") leaves 108px of slack inside its 384px
   --col block, so nothing is pushed toward the stub. Printed person lines
   only; the card's relation chips space themselves. */
.line .num,.reg-line .num{margin-inline-end:.2em}
a.num{text-decoration:none}
a.num:hover{color:var(--accent);text-decoration:underline}
a.num:focus-visible{color:var(--accent);text-decoration:underline}
/* A number the plate misprints. The digits stay exactly as every other number
   is set -- this is what the plate says -- and the ring around them is drawn
   with outline, never border or padding: a border would widen the row and
   throw the sibling bracket off its mother_row. The annotation itself is a
   separate row below, so nothing is inserted into the printed line at all. */
.sic-ring{outline:1.5px solid var(--sic);outline-offset:3px;border-radius:1rem}
/* height, not just line-height -- see the note on .line. nowrap, so it can
   never need to grow. */
.sic-row{line-height:var(--lh);height:var(--lh);
  padding-inline-start:1.4rem;white-space:nowrap}
.sic{font:italic .8rem var(--font-ui);color:var(--sic);
  text-decoration:underline dotted;text-underline-offset:.15em}
.sic:hover,.sic:focus-visible{text-decoration:underline solid}
.sex{color:var(--muted)}
/* + and ——— are text content (spouse marker, unrecorded name), so they hold
   the 4.5:1 text minimum via --muted; --rule is reserved for drawn rules. */
.plus{color:var(--muted);font-weight:600}
.name{font-weight:500}
.alt,.brace{color:var(--muted)}
.age{font-style:italic;font-variant-numeric:tabular-nums}
.vital{font-style:italic;color:var(--muted)}
/* The clan is the one field on a printed line that is not the person -- it is
   what the matrilineal check is run on -- so it carries its own colour. Reads
   through body.chart's flatten because --clan is its own token, exactly as
   --sic and --muted-fixed do; see the token for the measured contrasts. */
.clan{color:var(--clan)}
.blank{color:var(--muted);letter-spacing:-.05em}
.eng{background:var(--eng-bg);color:var(--eng-fg);border-radius:2px;
  padding:.02em .38em;font-size:.86em;font-weight:600}
.census{background:var(--cen-bg);color:var(--cen-fg);border-radius:2px;
  padding:.02em .38em;font-size:.8em}
/* Wraps at the column edge, as the plate sets it:
   'For second husband and offspring see Gen.' / 'II, 21, 74'

   EACH VISUAL LINE MUST BE EXACTLY --lh TALL, which is why line-height is the
   token and the block padding is zero. Chart.render counts a cross-reference
   row as `row += 1`, so a sibling group whose mother's line sits BELOW one is
   offset by mother_row * --lh -- and this rule used to render 21.09px against a
   24.8px budget, putting seven of Genealogy II's brackets 3.7px off their
   mother's line (measured 2026-07-29). Table 1 never showed it: no group there
   has a mother's line below an xref row, so the whole defect sat latent in
   shared CSS until a plate with six generations and 30 cross-references
   arrived. Exactly the reasoning .sic-row already carries -- keep the two
   together, and see the wrapping note below. */
/* min-height, not height -- see the note on .line for why the height has to be
   stated at all. This is the one row type that is white-space:normal, so a
   reference CAN wrap; min-height lifts a quantised 24px line box back to --lh
   without capping a wrap at one row, which would overlap the row below instead
   of merely mis-budgeting it. Wrapping is still the unsolved case the note
   above describes: split at the plate's own break with `|`. */
.xref{font-size:.8rem;color:var(--muted);font-style:italic;
  white-space:normal;max-width:var(--col);line-height:var(--lh);
  min-height:var(--lh);padding:0 0 0 1.4rem}
.xref a{color:inherit;text-decoration:underline dotted;text-underline-offset:.15em}
.xref a:hover{color:var(--accent)}
.xref a:focus-visible{color:var(--accent)}
/* A cross-reference standing in the child column, where the plate prints it in
   place of a sibling bracket: sits on its mother's baseline, no indent. */
/* padding must be 0 so the cell is exactly one line tall; otherwise it pushes
   the next sibling group off its mother's row. */
.xref-cell{padding:0;line-height:var(--lh);white-space:nowrap;max-width:none}
/* The caption now carries the pan hint and nothing else, so it is the caption
   that hides above 1400px -- hiding only the span would leave an empty
   figcaption holding its own bottom padding open. */
.plate-caption{display:none;max-width:var(--measure-wide);margin:0 auto;
  padding:0 var(--s5) var(--s5);font-size:var(--t-sm);line-height:1.6;
  color:var(--muted);text-wrap:pretty}
@media (max-width:1400px){.plate-caption{display:block}}

/* ---- register of persons ----------------------------------------------- */
.apparatus-register{max-width:var(--measure-wide);margin:0 auto;
  padding:var(--s6) var(--s5) 0}
/* A section head must not be smaller than the prose it introduces, and at
   13px small-caps it was. It tracks --t-prose rather than --t-base so it
   stays level as the body grows on wide screens -- a fixed --t-base head
   would fall behind again at the top of the clamp. Same size as the body is
   deliberate: the small caps, the weight and the tracking do the work, which
   is how the rest of this apparatus already distinguishes rank. */
.apparatus-register h2,footer h2,.prose h2{font-family:var(--font-plate);
  font-size:var(--t-prose);font-variant:small-caps;letter-spacing:.14em;
  font-weight:600;color:var(--ink);margin:0 0 var(--s3)}
.reg-note{font:var(--t-xs)/1.55 var(--font-ui);color:var(--muted);margin:0 0 var(--s3)}
.register-d{background:var(--panel);border:1px solid var(--rule-faint);
  border-radius:2px}
.register-d summary{cursor:pointer;padding:var(--s3) var(--s4);
  font:.8125rem var(--font-ui);color:var(--muted)}
.register-d summary:hover{color:var(--ink)}
.register-d[open] summary{border-block-end:1px solid var(--rule-faint)}
.reg-list{list-style:none;margin:0;padding:var(--s4);columns:1;
  column-gap:var(--s6)}
@media (min-width:900px){.reg-list{columns:2}}
.reg{break-inside:avoid;padding:var(--s2);
  scroll-margin-block:calc(var(--bar-h) + 1.5rem) 1rem}
.reg:target{background:var(--sel-bg);
  box-shadow:-.3rem 0 0 0 var(--accent-strong),0 0 0 .3rem var(--sel-bg);
  outline:2px solid var(--accent-strong);outline-offset:.3rem}
.reg + .reg{border-block-start:1px solid var(--rule-faint)}
.reg-line{font-size:var(--t-base);line-height:1.5}
.reg-rel{font-size:var(--t-sm);line-height:1.65;color:var(--muted);
  padding-inline-start:1.4rem}
.rel-l{font-family:var(--font-ui);font-size:var(--t-xs);text-transform:uppercase;
  letter-spacing:.08em;color:var(--muted)}
.reg-rel a{color:inherit;text-decoration:underline dotted;
  text-underline-offset:.15em}
.reg-rel a:hover{color:var(--accent)}
.reg-rel a:focus-visible{color:var(--accent)}

/* ---- footer apparatus --------------------------------------------------- */
/* The apparatus is long-form scholarly prose and was set at 14px, below a
   comfortable reading size. --t-prose also shortens the measure in
   characters at a fixed --measure, which is the other half of the fix. */
/* Two columns, so the apparatus stops being one long scroll and lands on the
   same left edge as the register above it. --measure-wide, not --measure:
   each column then measures about 34rem, still inside a comfortable line
   length at --t-prose, and the font size does not change. Grid of whole
   sections rather than CSS columns, because multicolumn would happily break a
   heading away from the list it introduces. */
footer{max-width:var(--measure-wide);margin:0 auto;
  padding:var(--s6) var(--s5) var(--s7);font-size:var(--t-prose);
  color:var(--muted);line-height:1.7;text-wrap:pretty}
.app-cols{display:grid;grid-template-columns:1fr;gap:var(--s5) var(--s6);
  align-items:start}
@media (min-width:56rem){.app-cols{grid-template-columns:repeat(2,minmax(0,1fr))}}
.app-sec{break-inside:avoid}
footer h2{margin:var(--s6) 0 var(--s3)}
footer h2:first-child,.app-sec h2:first-child{margin-block-start:0}
/* Editorial notes, Provenance and Citation fold away; The record and
   Navigating this chart do not. The split is what each section is FOR: the
   first two orient a reader who has just arrived at the plate, the other three
   are reference material consulted once and then in the way.

   Same disclosure the landing page's FAQ and the register above already use --
   the marker, the sizes and the hover are deliberately identical, because a
   third idiom for "this opens" is a cost with no reader on the other end of
   it. The <h2> lives INSIDE the summary so the heading outline survives the
   fold; a screen reader still finds five headings in the apparatus.

   Two things this must not break, both already solved elsewhere and reused
   rather than rebuilt: a deep link into a folded section (#note-misprint,
   #note-paternity, #note-crossref) is opened by openDetailsFor(), the same
   fragment insurance the register's disclosure relies on; and the offprint
   carries every section whatever the reader left folded -- see @media print. */
.app-d > summary{cursor:pointer;list-style:none;padding:var(--s3) 0;
  display:flex;gap:var(--s3);align-items:baseline}
.app-d > summary::-webkit-details-marker{display:none}
.app-d > summary::before{content:"+";color:var(--muted);
  font-variant-numeric:tabular-nums}
.app-d[open] > summary::before{content:"\\2013"}
.app-d > summary h2{margin:0}
.app-d > summary:hover h2{color:var(--accent)}
.app-d > summary:focus-visible{outline:2px solid var(--accent-strong);
  outline-offset:.2rem}
/* The fold is the section's own top edge, so the heading's block margin would
   double the gap the grid already sets. */
.app-sec > .app-d{margin:0}
footer ul{margin:.3rem 0;padding-left:1.2rem}
footer li{margin:.25rem 0}
/* An apparatus note that something on the chart links to. Same scroll-margin
   as .line and .reg so the sticky bar does not sit on top of it, and the same
   target treatment, so a reader who followed a marker can see which of eight
   notes they were sent to. */
footer li[id]{scroll-margin-block:calc(var(--bar-h) + 1.5rem) 1rem}
footer li:target{background:var(--sel-bg);
  box-shadow:-.3rem 0 0 0 var(--accent-strong),0 0 0 .3rem var(--sel-bg);
  outline:2px solid var(--accent-strong);outline-offset:.3rem}
footer a{color:var(--accent)}
footer code{font-family:var(--font-ui);font-size:.9em}
.cite-block{margin:var(--s3) 0;padding:var(--s3) var(--s4);
  background:var(--panel);border-inline-start:2px solid var(--rule)}
.cite-block p{margin:.4rem 0}
.copy-btn{font:var(--t-xs) var(--font-ui);color:var(--muted);background:none;
  border:1px solid var(--rule-faint);border-radius:2px;padding:.3rem .6rem;
  cursor:pointer;margin-inline-start:var(--s2);min-height:var(--tap)}
.copy-btn:hover{color:var(--ink);border-color:var(--rule)}
.copy-btn:focus-visible{color:var(--ink);border-color:var(--rule)}
.updated{margin-top:var(--s6);padding-top:var(--s4);
  border-top:1px solid var(--rule-faint);font-size:var(--t-xs)}
.print-url{display:none;font-size:var(--t-xs)}

/* ---- person card (popover; JS-filled from the register) ----------------- */
.pcard{position:fixed;margin:var(--s2);padding:0;
  width:min(38rem,92vw);background:var(--panel);color:var(--ink);
  border:1px solid var(--rule);border-radius:8px;
  box-shadow:0 6px 24px var(--shadow);font-size:var(--t-base);line-height:1.55}
/* ---- card header: the printed line, set as a title -------------------- */
/* The clan is lifted out of the line into a chip beside the name. It can be:
   person_line emits it as its own trailing <span class="clan">, with the
   sentence point living inside .name, so removing it leaves no orphan
   punctuation behind. Nothing else about the line is rearranged -- the number,
   sex mark and name keep the plate's order and the plate's wording. */
.pc-head{display:flex;align-items:center;gap:var(--s3);flex-wrap:wrap;
  padding:var(--s3) var(--s4);background:var(--wash);
  border-block-end:1px solid var(--rule-faint);border-radius:7px 7px 0 0;
  position:relative;z-index:1;box-shadow:0 3px 6px -2px var(--shadow)}
/* The badge carries the plate number, which is why the header text no longer
   does. Tabular figures so 8, 68 and 104 all sit centred in the same circle. */
.pc-mark{flex:none;display:grid;place-items:center;width:2.25rem;height:2.25rem;
  border-radius:50%;background:var(--accent);color:var(--paper);
  font:600 var(--t-sm)/1 var(--font-ui);font-variant-numeric:tabular-nums}
/* Name, age and vital note are separate fields, so they are separate layout
   elements and the space between them is a gap -- not a typed space, which
   would tie their separation to the font's word width. Each field keeps its
   own trailing point (person_line puts it inside the span), so punctuation
   travels with the field it belongs to. The whitespace text nodes between
   them survive only so the header still copies out as readable text; a
   whitespace-only node is not rendered as a flex item, so it contributes
   nothing to layout. */
.pc-title{display:flex;flex-wrap:wrap;align-items:baseline;
  column-gap:.5rem;row-gap:.1rem;
  margin:0;font:var(--t-xl)/1.25 var(--font-plate);font-weight:400;
  color:var(--ink);letter-spacing:0}
/* The clan keeps --clan inside the chip; the chip's border is --rule-faint so
   the colour is doing the work, not a second border weight. */
.pc-clan{margin-inline-start:auto;margin-inline-end:0;
  display:inline-flex;align-items:baseline;
  gap:.3rem;padding:.25rem .7rem;border:1px solid var(--rule-faint);
  border-radius:4px;background:var(--panel);font-size:var(--t-sm)}
.pc-clan-l{font:var(--t-xs) var(--font-ui);color:var(--muted)}
/* The vital note is metadata, not part of the name, so it steps back from it.
   --muted-fixed rather than --muted because body.chart flattens --muted to
   --ink; this is the second deliberate user of it after .imprint. Measured on
   the header band, both themes, below. */
.pcard .pc-title .vital{color:var(--muted-fixed);font-style:italic}
/* ---- card body: a column per spouse ------------------------------------ */
.pc-main{padding:var(--s3) var(--s4) var(--s4)}
.pc-cols{display:grid;grid-template-columns:repeat(auto-fit,minmax(13rem,1fr));
  column-gap:var(--s4);row-gap:var(--s5)}
/* The parents block owns the space beneath it. Without this the next
   section's heading box began exactly at the last parent button's bottom edge
   -- 0px clearance -- and its decorative rule, which paints centred in that
   box, landed 8px above the button. Margin here rather than a top margin on
   the heading, because .pc-h's top margin is zeroed so the two family columns
   start level with each other. */
.pc-sec{margin-block-end:var(--s5)}
/* A single spouse is centred at a readable width rather than stretched. */
.pc-cols--single{grid-template-columns:minmax(0,22rem);justify-content:center}
.pc-parents{display:grid;grid-template-columns:repeat(auto-fit,minmax(11rem,1fr));
  gap:var(--s2)}
.pc-parents > .pc-row{margin-block-start:0}
/* The divider is drawn only when there are exactly two columns. Columns wrap
   when there are more -- person 68 has two spouses plus a group whose father
   the plate does not record -- and a wrapped column is first on its own row
   while still matching `.pc-col + .pc-col`, which would hang a rule off its
   left edge with nothing to its left. Beyond two, the gap separates them. */
.pc-cols--pair > .pc-col + .pc-col{border-inline-start:1px solid var(--rule-faint);
  padding-inline-start:var(--s4)}
/* A section rule that reads as a rule, not as an underline on the words. */
.pc-h{display:flex;align-items:center;gap:var(--s2);margin:var(--s3) 0 var(--s2);
  font:var(--t-xs) var(--font-ui);text-transform:uppercase;letter-spacing:.08em;
  color:var(--muted)}
.pc-h::before,.pc-h::after{content:"";flex:1;height:1px;
  background:var(--rule-faint)}
.pc-col > .pc-h:first-child,.pc-sec > .pc-h:first-child{margin-block-start:0}
/* A relative's row is set at --t-base, the size the register entry and the
   plate line give a person -- not --t-sm. The row IS a person line: number,
   name, clan. Setting it smaller than the name in the header made the card's
   whole point -- who this person's family are -- the smallest text on it. The
   clan stays proportional at .92em, so one declaration moves both. */
.pc-row{display:flex;align-items:center;gap:var(--s2);
  padding:.35rem .55rem;margin-block-start:.3rem;
  border:1px solid var(--rule-faint);border-radius:5px;background:var(--paper);
  color:var(--ink);text-decoration:none;font-size:var(--t-base);
  min-height:var(--tap)}
.pc-row:first-of-type{margin-block-start:0}
a.pc-row:hover,a.pc-row:focus-visible{border-color:var(--rule);
  background:var(--sel-bg);text-decoration:none}
/* The parent's clan sits after the name and before the chevron, which keeps
   its auto start-margin and so stays hard right. Slightly smaller than the
   name: it is the same secondary weight the chart line gives it. */
.pc-row-clan{font-size:.92em}
/* The editorial-attribution mark. Deliberately quiet: it flags a reading, it
   does not warn about an error, so it is not --sic. It sits in the heading and
   in the register label, and always links to #note-paternity. */
.edmark{margin-inline-start:.35em;text-decoration:none;color:var(--accent);
  font-size:1.1em;line-height:1}
.edmark:hover,.edmark:focus-visible{text-decoration:underline}
.pc-h .edmark{flex:none}
.pc-chev{margin-inline-start:auto;color:var(--accent);font-size:1.1em;
  line-height:1}
a.pc-row:hover .pc-chev,a.pc-row:focus-visible .pc-chev{color:var(--ink)}
/* ---- card actions ------------------------------------------------------ */
.pcard-actions{display:flex;gap:var(--s3);align-items:center;
  justify-content:space-between;margin:0;
  padding:var(--s3) var(--s4);border-block-start:1px solid var(--rule-faint);
  font-family:var(--font-ui);font-size:.8125rem}
.pcard-actions button{font:var(--t-xs) var(--font-ui);color:var(--muted);
  background:none;border:0;padding:.3rem 0;cursor:pointer;
  min-height:var(--tap);text-decoration:underline;text-underline-offset:.15em}
.pcard-actions button:hover{color:var(--ink)}
.pcard-actions button:focus-visible{color:var(--ink)}
.pcard-actions a{color:var(--accent)}
/* The one filled control on the card. Its background is the affordance, which
   is why it is exempt from body.chart's link flatten below -- it needs no
   underline and --paper on --accent is the same measured pair as the clan
   gold, read the other way round: 5.86:1 light, 9.53:1 dark. */
.pcard-actions a.pc-btn{display:inline-flex;align-items:center;
  padding:.4rem .85rem;border-radius:5px;background:var(--accent);
  color:var(--paper);text-decoration:none;min-height:var(--tap)}
.pcard-actions a.pc-btn:hover,.pcard-actions a.pc-btn:focus-visible{
  background:var(--accent-strong);text-decoration:none}
@supports (anchor-name:--pc){
  .pcard{inset:auto;position-anchor:--pc;
    position-area:block-end span-inline-end;
    position-try-fallbacks:flip-block,flip-inline}
}
@media (prefers-reduced-motion:no-preference){
  .pcard{opacity:1;transition:opacity .12s ease;
    transition-behavior:allow-discrete}
  @starting-style{.pcard:popover-open{opacity:0}}
}
@media (max-width:640px){
  .pcard{inset:auto 0 0 0;width:100%;max-width:none;margin:0;
    max-height:75vh;max-height:75dvh;overflow-y:auto;
    border-radius:8px 8px 0 0;position-area:none}
  .pc-head{border-radius:8px 8px 0 0;position:sticky;top:0;z-index:2}
  /* Columns are stacked here, so the divider has nothing to its left. It must
     out-specify .pc-cols--pair > .pc-col + .pc-col, which is (0,3,0) -- a bare
     .pc-col + .pc-col is (0,2,0) and lost, leaving the second column indented
     16px with a rule hanging down its left edge on a phone. */
  .pc-cols > .pc-col + .pc-col{border-inline-start:0;padding-inline-start:0}
  .pc-cols,.pc-cols--single,.pc-parents{grid-template-columns:1fr}
}

/* ---- flat text colour, chart pages -------------------------------------- */
/* Every piece of text on a table page renders in --ink, the colour of the h1.
   --muted is REDEFINED rather than each rule rewritten -- the same technique
   the prefers-contrast:more block above already uses -- which sweeps the
   numbers, sex marks, ages, vital notes, cross-references, ruler, register
   labels, citation and footer in one declaration.
   --rule is deliberately NOT touched: the plate's brackets and leader rules
   are drawn lines, not text, and they carry the genealogy's structure. */
body.chart{--muted:var(--ink)}
/* Colour was the only thing marking these as links, so they take an explicit
   underline in its place -- dropping both would leave no affordance at all
   (WCAG 1.4.1). The xref and register links already carry a dotted underline. */
body.chart footer a,body.chart .pcard-actions a:not(.pc-btn){
  color:var(--ink);text-decoration:underline;text-underline-offset:.15em}
/* Hover and focus still change, but by weight of underline rather than by
   introducing a second colour. */
body.chart a.wordmark:hover,body.chart a.wordmark:focus-visible,
body.chart a.num:hover,body.chart a.num:focus-visible,
body.chart .xref a:hover,body.chart .xref a:focus-visible,
body.chart .reg-rel a:hover,body.chart .reg-rel a:focus-visible,
body.chart .register-d summary:hover,
body.chart .masthead nav a:hover,body.chart .masthead nav a:focus-visible{
  color:var(--ink);text-decoration:underline}
/* The one exception: the current-table pill is set ON --ink, so its text has
   to stay --paper or it disappears into its own background. */
body.chart .masthead nav a[aria-current="page"]{color:var(--paper)}

/* ---- print: the offprint ------------------------------------------------ */
@media print{
  :root,:root[data-theme]{color-scheme:light}
  body{background:#fff;color:#000}
  /* .theme-foot joins the list because Theme left the masthead on
     2026-08-10: hiding .masthead used to cover it, and no longer does. */
  .masthead,.theme-foot,.skip,.plate-tools,.plate-caption,.apparatus-register,
  [popover],.ruler-chipslot{display:none}
  .scroll{overflow:visible;padding:0}
  .scroll-shell::before,.scroll-shell::after,.ruler::after{display:none}
  .sheet{border:0;box-shadow:none}
  .plate-zoom{zoom:var(--print-zoom,.7)}
  /* The chip is hidden in print, so the band reserved above the labels for it
     is dead space on the sheet. Back to one band. */
  .ruler{border-block-end-color:#000;height:2rem}
  .tree{break-inside:avoid-page}
  footer{break-before:page}
  footer a[href^="http"]::after,.prose a[href^="http"]::after{
    content:" (" attr(href) ")";font-size:.9em}
  .line:target,.line.is-selected,.reg:target{
    background:none;box-shadow:none;outline:none}
  /* The offprint carries the WHOLE apparatus, whatever the reader left folded
     -- a printed edition with its citation collapsed away is not an edition.
     ::details-content does it with no script on a current engine; the
     beforeprint handler in UI_JS is the same guarantee for the rest, and is
     what a reader printing an older browser gets. Belt and braces on purpose:
     this is the one place where the fold silently losing content would leave
     no trace on screen to notice it by. */
  .app-d::details-content{content-visibility:visible!important;
    block-size:auto!important;opacity:1!important}
  .app-d > summary{padding:0}
  .app-d > summary::before{display:none}
  /* The offprint is black on white. The clan gold prints as a weak grey, and
     the colour was never carrying information a reader could not get from the
     word itself, so it flattens here rather than degrading. */
  a.num,.xref a,.clan{color:inherit;text-decoration:none}
  .print-url{display:block}
  @page{size:landscape;margin:12mm}
}
"""

# A bracket mark: two generations joined, which is what the plate is made of.
FAVICON = (
    "data:image/svg+xml,"
    "%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 32 32'%3E"
    "%3Crect width='32' height='32' rx='5' fill='%231c1a17'/%3E"
    "%3Cg stroke='%23d9b872' stroke-width='2' fill='none'%3E"
    "%3Cpath d='M7 16h6M13 8v16M13 8h6M13 16h6M13 24h6'/%3E%3C/g%3E%3C/svg%3E"
)


CITE = """Elsie Clews Parsons, &ldquo;Laguna Genealogies,&rdquo;<br>
    <em>Anthropological Papers of the American Museum of Natural History</em>,
    vol.&nbsp;19, pt.&nbsp;5 (1923), pp.&nbsp;133&ndash;292."""

# Notes true of every plate in the edition. Plate-specific notes live on the
# TABLES entry, because a note about Table 1's misprint printed under Table 4
# would be a false statement about that plate.
READING_COMMON = """
    <li>Clan membership is matrilineal: every sibling group carries its mother's clan.
        That rule was used to verify the bracket readings against the plate.</li>
    <li>A dash in place of a name means the plate printed no name.</li>
    {blocks}
"""


def blocks_note(spec):
    """
    The note on the dashed rules between descent blocks, counted from
    spec["roots"] rather than typed.

    This existed as fixed copy reading "the dashed rule separating the two
    halves ... the two founding couples" while both published plates happened
    to have exactly two. Genealogy II has THREE, and `.tree + .tree` draws a
    rule before every block after the first, so the fixed copy stated two
    wrong numbers at once. Anything here that counts must come from the data,
    for the same reason the statistics line does.
    """
    n = len(spec["roots"])
    couples = spec["couples"]
    if n < 2:
        return f"<li>The chart draws one descent block, from {couples}.</li>"
    if n == 2:
        # The published wording, kept verbatim DOWN TO THE SOURCE LINE BREAK,
        # so this refactor leaves the two cited pages byte-identical. Without
        # the break the rendered text is the same and the file still differs,
        # which is a diff to explain on a page nobody meant to touch.
        return ("<li>The dashed rule separating the two halves is an editorial addition,\n"
                "        not on the plate. It marks the two founding couples, "
                f"{couples}.</li>")
    blocks = NUMBER_WORDS.get(n, str(n))
    return (f"<li>The dashed rules separating the {blocks} descent blocks are an "
            "editorial addition, not on the plate &mdash; one above each block "
            f"after the first. They mark the {blocks} founding couples, "
            f"{couples}.</li>")

# The redesign's one disclosure: the reading aids are 2026 apparatus, the sheet
# is 1923. Printed with the editorial notes on every chart page.
APPARATUS_NOTE = """
    <li>The generation ruler above the chart, the linked person numbers, and the
        register below the chart are editorial apparatus of this edition, not
        part of the plate.</li>
"""

def navigating_html(spec):
    """The footer's how-to list, and since the on-page key was removed the only
    place `+`, `F.`/`M.` and the leader rule are explained. Do not thin those
    three out: nothing else on the page decodes them. The example anchor uses
    this table's first founding person, so it is correct on every table
    regardless of size."""
    r = spec["roots"][0]
    return f"""
    <li>A <em>+</em> at the start of a line marks a spouse, printed on the line
        below the person they married.</li>
    <li><em>F.</em> and <em>M.</em> give the sex of each person as printed.</li>
    <li>The horizontal leader rule runs from a person to the bracket holding
        their children; the bracket itself hangs on the mother&rsquo;s line.</li>
    <li>A number after a name is the person&rsquo;s age when the data was collected.
        <em>d.</em> means they had already died when Parsons recorded the genealogy,
        during her fieldwork of 1918&ndash;19; where she knew the year she gives it,
        as in <em>d.&nbsp;1913</em>.</li>
    <li>Every person number is a stable link &mdash; <code>#p{r}</code> addresses
        person {r}&rsquo;s first printed line, and the address bar keeps it while
        you read.</li>
    <li>The chart region can be focused and panned with the arrow keys; on a
        narrow screen, scroll sideways to follow the later generations.</li>
    <li>With JavaScript: with the chart focused, press <code>/</code> to find a
        person by number or name; selecting a person number opens its register
        card, and Esc closes it. The Theme and Scale controls remember your
        choice.</li>
"""

# Applies a stored manual theme before first paint so the page cannot flash the
# wrong palette. Inline in <head> on every page; theme state is otherwise CSS.
THEME_SNIPPET = ('<script>try{var t=localStorage.getItem("lg-theme");'
                 'if(t==="light"||t==="dark")document.documentElement.dataset.theme=t}'
                 'catch(e){}</script>')

# The shared theme module: toggle Light <-> Dark on #theme, persist, and keep
# the theme-color meta in step. String-formatted into both scripts below.
#
# There is no Auto state in the control, and since 2026-08-10 there is no
# system-preference state either: an untouched control resolves to LIGHT, not
# to the OS. The CSS says the same thing on its own (see the color-scheme note
# in CSS), so this function only has to agree with it -- it is not the thing
# that decides. Both halves have to change together or a dark-OS reader gets a
# light page whose button says "Theme: Dark".
_THEME_JS = r"""
var THEMES=["light","dark"];
function applyTheme(t){
  if(t!=="light"&&t!=="dark"){t="light"}
  root.dataset.theme=t;
  doc.querySelectorAll('meta[name="theme-color"]').forEach(function(m){
    m.content=t==="light"?"#FAF8F4":"#191713"});
  var b=$("#theme");if(b)b.textContent="Theme: "+t.charAt(0).toUpperCase()+t.slice(1);
}
function cycleTheme(){
  var cur=root.dataset.theme==="dark"?"dark":"light";
  var next=THEMES[(THEMES.indexOf(cur)+1)%THEMES.length];
  store("lg-theme",next);applyTheme(next);
}
var themeBtn=$("#theme");
/* No stored choice resolves to light and is NOT written back: nothing is
   persisted until the reader presses the control, so an untouched control
   leaves no trace in storage. */
if(themeBtn){themeBtn.hidden=false;applyTheme(read("lg-theme"))}
"""

_JS_PRELUDE = r"""
"use strict";
var doc=document,root=doc.documentElement;
function $(s,c){return (c||doc).querySelector(s)}
function store(k,v){try{v===null?localStorage.removeItem(k):localStorage.setItem(k,v)}catch(e){}}
function read(k){try{return localStorage.getItem(k)}catch(e){return null}}
"""

# Landing page: the theme module only.
LANDING_JS = ("(function(){" + _JS_PRELUDE + _THEME_JS + r"""
doc.addEventListener("click",function(e){
  var t=e.target&&e.target.closest?e.target.closest("[data-action=theme]"):null;
  if(t)cycleTheme();
});
})();""")

# Chart pages: theme, scale, finder, register insurance, copy actions and the
# person card. Every module is feature-checked so one failure never cascades;
# everything JS reveals is authored hidden, so a no-JS page has no dead
# controls. Five delegated listeners, no scroll or resize handlers.
UI_JS = ("(function(){" + _JS_PRELUDE + _THEME_JS + r"""
function flip(b,txt){var o=b.textContent;b.textContent=txt;b.disabled=true;
  setTimeout(function(){b.textContent=o;b.disabled=false},1500)}

/* plate scale -- the reduction. Zooms ruler and sheet together. */
var pz=$(".plate-zoom"),mount=$("#scale-mount");
function applyScale(v){
  pz.style.setProperty("--plate-zoom",v);
  mount.querySelectorAll("button").forEach(function(b){
    b.setAttribute("aria-pressed",b.dataset.v===v?"true":"false")});
}
/* Gated on CSS zoom support (Firefox gained it in 126): an engine that would
   drop the zoom declaration must get the documented no-control fallback,
   not three buttons that do nothing. */
if(pz&&mount&&typeof CSS!=="undefined"&&CSS.supports&&CSS.supports("zoom","0.7")){
  mount.setAttribute("role","group");
  mount.setAttribute("aria-label","Plate scale");
  mount.insertAdjacentHTML("beforeend",
    '<span class="scale-l">Scale</span>'+
    ["1","0.85","0.7"].map(function(v){
      return '<button class="scale-btn" data-action="scale" data-v="'+v+'">'+
        Math.round(parseFloat(v)*100)+'%</button>'}).join(""));
  var sv=read("lg-scale");
  applyScale(sv==="0.85"||sv==="0.7"?sv:"1");
}

/* finder: leading integer wins; otherwise diacritic-folded substring match. */
var form=$("#find"),input=$("#find-q"),note=$("#find-note"),
    datalist=$("#persons-list");
function fold(s){return s.normalize("NFD").replace(/\p{M}+/gu,"")
  .replace(/[\p{Lm}˙'’]/gu,"").toLowerCase()}
if(form&&input&&datalist){
  form.hidden=false;
  var opts=[].map.call(datalist.options,function(o){
    /* dataset.n is the number the plate PRINTS, carried only where it is
       shared; everywhere else the id already is it. */
    return {id:o.value,n:o.dataset.n||o.value,nm:o.dataset.nm||"",
            folded:fold(o.textContent||"")}});
  form.addEventListener("submit",function(e){
    e.preventDefault();
    var q=input.value.trim();if(!q)return;
    var id=null,also=null,m=q.match(/^\d+/);
    if(m){
      /* The PRINTED number first -- it is what the reader is holding, and the
         plate prints one number on two people three times in the edition. The
         id is only the fallback, which is what reaches the three people whose
         printed number belongs to somebody else. Getting this order backwards
         is the whole of the defect: an id lookup silently answers with one of
         the two and offers no route to the other. */
      var hits=[];
      for(var j=0;j<opts.length;j++)if(opts[j].n===m[0])hits.push(opts[j]);
      if(hits.length){id=hits[0].id;if(hits.length>1)also=hits[1]}
      else id=m[0];
    }
    if(!id){var f=fold(q);
      for(var i=0;i<opts.length;i++){
        if(opts[i].folded.indexOf(f)>=0){id=opts[i].id;break}}}
    var el=id&&doc.getElementById("p"+id);
    if(!el){note.textContent="No person “"+q+"” in this table.";return}
    note.textContent="";
    /* The plate's own reuse of a number is worth surfacing, not hiding. Built
       as nodes rather than innerHTML: the name is plate text, not ours. */
    if(also){
      note.appendChild(doc.createTextNode(
        "The plate prints "+m[0]+" on two people — also "));
      var a2=doc.createElement("a");
      a2.href="#p"+also.id;a2.textContent=also.nm;
      note.appendChild(a2);
      note.appendChild(doc.createTextNode("."));
    }
    if(location.hash!=="#p"+id)location.hash="#p"+id;
    /* Unconditional: finding the person already named by the hash fires no
       hashchange, and the row may have been deselected by a click since. */
    syncSelection();
    el.scrollIntoView({inline:"start",block:"center",
      behavior:matchMedia("(prefers-reduced-motion:no-preference)").matches
        ?"smooth":"auto"});
  });
  /* WCAG 2.1.4: the bare "/" shortcut is active only while focus is inside
     the plate figure -- the component it serves -- never page-wide. */
  var plateEl=$(".plate");
  if(plateEl)plateEl.addEventListener("keydown",function(e){
    if(e.key==="/"&&!e.altKey&&!e.ctrlKey&&!e.metaKey){
      e.preventDefault();input.focus();input.select();
    }
  });
  input.addEventListener("keydown",function(e){
    if(e.key==="Escape")input.blur();
  });
}

/* fragment insurance: a #r{n} link must open the register's disclosure, and
   hash navigation dismisses the person card. */
function openDetailsFor(hash){
  if(!hash||hash.length<2)return;
  var el=doc.getElementById(hash.slice(1));
  if(el&&el.closest){var d=el.closest("details");
    /* If we had to open the disclosure ourselves, the browser's own fragment
       scroll already missed (the target was display:none) -- finish the jump. */
    if(d&&!d.open){d.open=true;el.scrollIntoView({block:"center"})}}
}
addEventListener("hashchange",function(){
  openDetailsFor(location.hash);
  if(popoverOK){try{card.hidePopover()}catch(e){}syncSelection()}
  if(navFromCard){
    navFromCard=false;
    var el=doc.getElementById(location.hash.slice(1));
    var na=el&&el.querySelector?el.querySelector("a.num"):null;
    if(na)na.focus({preventScroll:true});
  }
});
openDetailsFor(location.hash);

/* print insurance: the offprint carries every apparatus section, whatever the
   reader left folded. The print stylesheet does this on its own where
   ::details-content is supported; this covers the engines where it is not.
   It restores exactly what it opened, so a reader's own folds survive the
   print dialog -- reopening all five would be a change they never made. */
var printOpened=[];
addEventListener("beforeprint",function(){
  printOpened=[];
  doc.querySelectorAll(".app-d").forEach(function(d){
    if(!d.open){d.open=true;printOpened.push(d)}
  });
});
addEventListener("afterprint",function(){
  printOpened.forEach(function(d){d.open=false});
  printOpened=[];
});

/* copy citation */
var citeText=$("#cite-text"),copyMount=$("#copy-mount");
if(citeText&&copyMount&&navigator.clipboard){
  copyMount.insertAdjacentHTML("beforeend",
    '<button class="copy-btn" data-action="copycite" aria-live="polite">Copy citation</button>');
}

/* person card: filled by cloning the register entry -- single source of
   truth. Without popover support the number stays a plain working anchor. */
var card=$("#pcard"),CANON=root.getAttribute("data-canonical")||"",
    lastInvoker=null,anchoredEl=null,navFromCard=false,
    popoverOK=!!card&&("showPopover" in HTMLElement.prototype),
    anchorsOK=typeof CSS!=="undefined"&&CSS.supports&&
      CSS.supports("anchor-name","--pc");
/* The selected row. Opening by click set no visible state at all before --
   only #p anchors lit a row up -- so the reader could lose track of which
   person the card belonged to. Once the card script is live this is the ONLY
   thing that lights a row: the stylesheet drops .line:target under
   html[data-card]. See the CSS for why.
   cardRow is the row the open card belongs to, kept so closing the card can
   clear the highlight it set WITHOUT clearing one that hash navigation has
   already moved elsewhere -- the popover's toggle event and hashchange both
   arrive as tasks and their order is not guaranteed. */
var selRow=null,cardRow=null;
function markSelected(el){
  if(selRow)selRow.classList.remove("is-selected");
  selRow=el||null;
  if(selRow)selRow.classList.add("is-selected");
}
/* The row a fragment names, or null -- for #r{n}, #note-{x} and every other
   hash, which select no row. Deliberately not "the element's closest .line":
   a hash that lands inside the register must leave the plate unselected. */
function lineFor(hash){
  var el=hash&&hash.charAt(0)==="#"&&hash.length>1
    ? doc.getElementById(hash.slice(1)) : null;
  return el&&el.classList&&el.classList.contains("line") ? el : null;
}
function syncSelection(){markSelected(lineFor(location.hash))}
if(popoverOK){
  card.hidden=false;
  /* The stylesheet's switch. Set here rather than in <head> because it must
     mean "the card script is running", not merely "JavaScript is enabled":
     where the popover is unsupported the numbers stay plain anchors and
     :target is still the only thing that can select a row. */
  root.dataset.card="1";
  card.addEventListener("toggle",function(e){
    /* preventScroll: closing via a relation link must not yank the plate
       back to the invoker -- the reader is on their way somewhere else. */
    if(e.newState==="closed"){
      if(selRow===cardRow)markSelected(null);
      cardRow=null;
      if(lastInvoker){lastInvoker.focus({preventScroll:true});lastInvoker=null}}});
  /* A shared #p link still arrives selected, now by class rather than :target. */
  syncSelection();
}
function openCard(a){
  var id=(a.getAttribute("href")||"").slice(2);
  var reg=doc.getElementById("r"+id);
  if(!reg)return false;
  card.innerHTML="";
  /* The register entry is still the single source of truth and still the no-JS
     person card; the popover REGROUPS a detached copy of it rather than
     rendering it. src is never inserted into the document. */
  var src=doc.createElement("div");
  src.innerHTML=reg.innerHTML;
  var body=doc.createElement("div");
  body.className="pcard-body";

  /* ---- header: badge, name, clan ---------------------------------------- */
  var line=src.querySelector(".reg-line");
  var clan=line?line.querySelector(".clan"):null;
  var numEl=line?line.querySelector(".num"):null;
  /* Opened from a line the plate misnumbers: the badge shows the number the
     reader is looking at, not the register's. The register keeps the true one,
     so the printed one has to be carried on the link. The number only, never
     the annotation -- the chart row the reader opened it from already sits
     under its own '(misprint)' row. */
  var printed=a.dataset?a.dataset.printed:null;
  var plateNo=printed||(numEl?numEl.textContent.replace(/\.$/,""):id);

  var head=doc.createElement("header");head.className="pc-head";
  var mark=doc.createElement("span");
  mark.className="pc-mark";mark.setAttribute("aria-hidden","true");
  mark.textContent=plateNo+".";
  head.appendChild(mark);

  var title=doc.createElement("h2");
  title.className="pc-title";title.id="pcard-t";
  /* The number and sex mark leave the header TEXT -- the badge carries the
     number now -- but they stay in the dialog's accessible NAME, which is this
     element. Removing them outright would silently rename every card.
     Rebuilt from the element children rather than moved wholesale, so exactly
     one space lands between the name and the vital note: "Tsaʼtsiʼ. d. in
     1905." can never come out closed up, whatever the source spacing was. */
  var vh=doc.createElement("span");
  vh.className="visually-hidden";
  vh.textContent=plateNo+". "+(line&&line.querySelector(".sex")
    ? line.querySelector(".sex").textContent+" " : "");
  title.appendChild(vh);
  var DROP=/(^| )(num|sex|clan|plus|sic-ring)( |$)/;
  var kept=[];
  if(line)[].slice.call(line.children).forEach(function(el){
    if(!DROP.test(" "+el.className+" "))kept.push(el);
  });
  kept.forEach(function(el,i){
    if(i)title.appendChild(doc.createTextNode(" "));
    title.appendChild(el);
  });
  head.appendChild(title);

  if(clan){
    var chip=doc.createElement("span");chip.className="pc-clan";
    /* "Clan: Water". Suppressed for the one value that is an origin rather
       than a clan -- person 101's "of Zuñi" -- which person_line marks with
       .clan-origin so this never has to read the string to find out. */
    if(!/(^| )clan-origin( |$)/.test(" "+clan.className+" ")){
      var lbl=doc.createElement("span");
      lbl.className="pc-clan-l";lbl.textContent="Clan:";
      chip.appendChild(lbl);chip.appendChild(doc.createTextNode(" "));
    }
    chip.appendChild(clan);head.appendChild(chip);
  }
  body.appendChild(head);

  /* ---- body: one column per spouse ------------------------------------ */
  /* Read off data-rel / data-with, never off the label: "Children (with 66)"
     is prose, and digging a person number out of prose is the mistake _p()
     exists to prevent. A row with no data-rel is a cross-reference, which the
     card does not carry -- the chart prints it under the line already. */
  function heading(text,editorial){
    var h=doc.createElement("h3");h.className="pc-h";h.textContent=text;
    /* Same dagger, same target as the register's: the grouping under this
       heading is the edition's reading, not the plate's bracket. */
    if(editorial){
      var m=doc.createElement("a");
      m.className="edmark";m.href="#note-paternity";
      m.title="editorial attribution";m.textContent="†";
      h.appendChild(m);
    }
    return h;
  }
  /* Every related person's clan, read from that person's OWN register entry
     rather than added to rel_link -- which would also rewrite the register.
     Omitted entirely when the plate records neither clan nor origin (person 89
     is the only such case in either table); no placeholder stands in for it.
     The span keeps its .clan class, so it keeps --clan. */
  function withClan(el){
    var m=/^#p(\d+)$/.exec(el.getAttribute("href")||"");
    if(!m)return el;
    var e=doc.getElementById("r"+m[1]);
    var c=e?e.querySelector(".reg-line .clan"):null;
    if(!c)return el;
    var s=c.cloneNode(true);
    s.className+=" pc-row-clan";
    var chev=el.querySelector(".pc-chev");
    /* A real space, not just the flex gap: a whitespace-only text node is not
       rendered as a flex item, so this costs nothing on screen, but without it
       the row copies out as "NayowʼăitsaSun". */
    el.insertBefore(doc.createTextNode(" "),chev);
    el.insertBefore(s,chev);
    return el;
  }
  function row(el){
    el.className="pc-row"+(el.tagName==="A"?"":" pc-row-flat");
    var n=el.querySelector(".num");
    if(n&&n.textContent.slice(-1)!==".")n.textContent+=".";
    if(el.tagName==="A"){
      var c=doc.createElement("span");
      c.className="pc-chev";c.setAttribute("aria-hidden","true");
      c.textContent="›";el.appendChild(c);
    }
    return withClan(el);
  }
  var parents=null,parentsEd=false,spouses=[],kids={},kidsEd={},order=[];
  src.querySelectorAll(".reg-rel").forEach(function(r){
    var kind=r.getAttribute("data-rel");
    if(!kind)return;
    var items=[].slice.call(r.children).filter(function(c){
      return c.className&&(" "+c.className+" ").indexOf(" rel-x ")>=0;
    });
    if(!items.length)return;
    if(kind==="parents"){
      parents=items;
      /* Same flag the Children rows use, so an attributed father is marked on
         the card exactly as he is in the register. */
      parentsEd=r.getAttribute("data-editorial")==="1";
    }
    else if(kind==="spouses")spouses=items;
    else if(kind==="children"){
      var w=r.getAttribute("data-with")||"0";
      kids[w]=items;kidsEd[w]=r.getAttribute("data-editorial")==="1";
      if(order.indexOf(w)<0)order.push(w);
    }
  });

  var main=doc.createElement("div");main.className="pc-main";
  if(parents){
    var ps=doc.createElement("section");ps.className="pc-sec";
    ps.appendChild(heading("Parents",parentsEd));
    /* Mother and father side by side in equal columns, stacking themselves
       when the card is too narrow -- auto-fit does both, so there is no
       breakpoint here to keep in step with the one on .pc-cols. */
    var pg=doc.createElement("div");pg.className="pc-parents";
    parents.forEach(function(x){pg.appendChild(row(x));});
    ps.appendChild(pg);
    main.appendChild(ps);
  }
  var cols=doc.createElement("div");cols.className="pc-cols";
  var used={};
  spouses.forEach(function(sp){
    var sid=(sp.getAttribute("href")||"").slice(2);
    var col=doc.createElement("section");col.className="pc-col";
    col.appendChild(heading("Spouse"));
    col.appendChild(row(sp));
    if(sid&&kids[sid]){
      col.appendChild(heading("Children",kidsEd[sid]));
      kids[sid].forEach(function(k){col.appendChild(row(k));});
      used[sid]=1;
    }
    cols.appendChild(col);
  });
  /* A group whose other parent the plate does not give -- data-with="0" -- and
     any group whose partner is not among the spouses. Its own column, headed
     only "Children", so nothing is silently dropped. */
  order.forEach(function(w){
    if(used[w])return;
    var col=doc.createElement("section");col.className="pc-col";
    col.appendChild(heading("Children",kidsEd[w]));
    kids[w].forEach(function(k){col.appendChild(row(k));});
    cols.appendChild(col);
  });
  /* One column -- a single spouse, or a lone group whose other parent the
     plate does not give -- is centred at a readable width instead of being
     stretched across a card sized for two. Decided from what was actually
     built, so "one spouse" and "no spouse, one group" behave alike. */
  if(cols.children.length===1)cols.className+=" pc-cols--single";
  else if(cols.children.length===2)cols.className+=" pc-cols--pair";
  if(cols.children.length)main.appendChild(cols);
  if(main.children.length)body.appendChild(main);
  card.appendChild(body);

  var act=doc.createElement("p");act.className="pcard-actions";
  var inner='<a class="pc-btn" href="#r'+id+'">Open register entry</a>';
  if(navigator.clipboard&&CANON)
    inner='<button data-action="copylink" data-id="'+id+'" aria-live="polite">Copy link</button> '+inner;
  act.innerHTML=inner;card.appendChild(act);
  if(anchoredEl)anchoredEl.style.removeProperty("anchor-name");
  anchoredEl=a;lastInvoker=a;
  cardRow=a.closest(".line");markSelected(cardRow);
  if(anchorsOK)a.style.setProperty("anchor-name","--pc");
  card.style.cssText="";
  card.showPopover();
  if(!anchorsOK&&!matchMedia("(max-width:640px)").matches){
    var r=a.getBoundingClientRect(),t=r.bottom+4,l=r.left;
    if(t+card.offsetHeight+8>innerHeight)t=Math.max(8,r.top-card.offsetHeight-4);
    if(l+card.offsetWidth+8>innerWidth)l=Math.max(8,innerWidth-card.offsetWidth-8);
    card.style.inset="auto";card.style.top=t+"px";card.style.left=l+"px";
  }
  card.focus();
  return true;
}

/* The whole printed line opens the person's card, not only the number. It
   delegates to openCard with the row's own a.num, so the card still anchors
   to the number and focus still returns to a real focusable element on close.
   Deliberately NOT a tabindex on the row: the a.num is already the tab stop,
   and a second one per line would double the tab stops on a 104-person plate.
   Guards, in order:
     - no popover support: leave the browser's #p anchor behaviour alone;
     - a modified click is the browser's (open in a new tab);
     - a real anchor inside the row keeps its own behaviour -- the plate's
       cross-references are links;
     - a click that ends a text selection is a copy gesture, not a tap. The
       names here exist to be selected and cited, so this must not eat that. */
function rowClick(e){
  if(!popoverOK)return;
  if(e.metaKey||e.ctrlKey||e.shiftKey||e.altKey)return;
  var el=e.target;
  if(!el||!el.closest)return;
  if(el.closest("a,button,input,summary,[popover]"))return;
  var sel=doc.getSelection&&doc.getSelection();
  if(sel&&!sel.isCollapsed&&String(sel).trim())return;
  var line=el.closest(".line");
  var a=line&&line.closest(".sheet")?line.querySelector("a.num"):null;
  if(a){if(openCard(a))e.preventDefault();return}
  /* A click on bare plate deselects. Without this the row a relation link
     selected stayed lit with nothing to clear it: its card closed on the way
     out, so no close event is coming, and the hash does not change again. */
  markSelected(null);
}

/* one delegated click listener for every control */
doc.addEventListener("click",function(e){
  var t=e.target&&e.target.closest?e.target.closest("[data-action],a.num,#pcard a"):null;
  if(!t){rowClick(e);return}
  var act=t.dataset?t.dataset.action:"";
  if(act==="theme"){cycleTheme();return}
  if(act==="scale"){store("lg-scale",t.dataset.v);applyScale(t.dataset.v);return}
  if(act==="copycite"&&citeText){
    navigator.clipboard.writeText(citeText.textContent.replace(/\s+/g," ").trim())
      .then(function(){flip(t,"Copied")})
      .catch(function(){flip(t,"Copy failed")});return}
  if(act==="copylink"){
    navigator.clipboard.writeText(CANON+"#p"+t.dataset.id)
      .then(function(){flip(t,"Copied")})
      .catch(function(){flip(t,"Copy failed")});return}
  /* a hash link inside the card: close toward the target, never back to the
     invoker; when the hash will not change, no hashchange fires, so finish
     the jump here. */
  if(popoverOK&&t.tagName==="A"&&t.closest("#pcard")){
    var h=t.getAttribute("href")||"";
    if(h.charAt(0)==="#"){
      lastInvoker=null;navFromCard=true;
      try{card.hidePopover()}catch(err){}
      if(h===location.hash){
        navFromCard=false;
        var same=doc.getElementById(h.slice(1));
        if(same){same.scrollIntoView({inline:"start",block:"center"});
          var na=same.querySelector("a.num");if(na)na.focus({preventScroll:true})}
        /* No hashchange will fire, so the selection has to be moved here --
           and after cardRow is dropped, or the close handler reads them as
           equal and clears the row we just selected. */
        cardRow=null;syncSelection();
        e.preventDefault();
      }
    }
    return;
  }
  if(popoverOK&&t.matches&&t.matches("a.num")&&t.closest(".sheet")){
    /* a modified click asks for the browser's own behavior (new tab etc.) */
    if(e.metaKey||e.ctrlKey||e.shiftKey||e.altKey)return;
    if(openCard(t))e.preventDefault();
  }
});
})();""")


# Drawn, not typed. A magnifier character (U+2315, U+1F50D) is either missing
# from the UI stack or an emoji, and this bar is set in system-ui -- there is no
# embedded face to guarantee either. currentColor keeps it on the same hover and
# focus transitions as the label it replaces.
SEARCH_GLYPH = (
    '<svg class="mast-glyph" viewBox="0 0 16 16" width="14" height="14" '
    'aria-hidden="true" focusable="false">'
    '<circle cx="6.75" cy="6.75" r="4.75" fill="none" stroke="currentColor" '
    'stroke-width="1.6"/>'
    '<path d="M10.4 10.4 14.4 14.4" stroke="currentColor" stroke-width="1.6" '
    'stroke-linecap="round"/></svg>'
)


def masthead_html(tables, current_slug, prefix, home):
    """
    The sticky site bar -- the guaranteed carrier of identity and the way home
    2800px deep in the page. `tables` is [(numeral, slug), ...]; `prefix`
    starts every table link ("../" on a published chart page, "" on the
    landing page, the absolute site URL in the private build); `home` is the
    wordmark's href, or None on the landing page, which names itself.
    """
    # "Home", not the edition's name -- set by the user 2026-08-10. The bar is
    # now a way back rather than a nameplate; the edition names itself in the
    # <title>, the title block and the citation, which is where a reader who
    # wants to cite it looks anyway.
    if home:
        mark = f'<a class="wordmark" href="{home}">Home</a>'
    else:
        mark = '<span class="wordmark" aria-current="page">Home</span>'
    # The pills show the NUMERAL ALONE at every width (user, 2026-08-10). The
    # word is still in the markup and still in the accessible name -- it is
    # hidden visually by .masthead nav .nav-word, not removed -- because "I"
    # on its own is not a link name anyone can act on by ear, and the nav's
    # aria-label="Tables" groups them but does not name them individually.
    # This is the SAME .nav-word mechanism the Search label uses, now applied
    # at every width here and still only below 26rem there; the two are
    # deliberately separate selectors, so widening one does not silently
    # unhide the other.
    links = "".join(
        f'<a href="{prefix}{slug}/"'
        + (' aria-current="page"' if slug == current_slug else "")
        + f'><span class="nav-word">Genealogy </span>{numeral}</a>'
        for numeral, slug in tables)
    # Search is NOT in the Tables nav, which is a list of plates -- the search
    # page is not a fifth plate, and putting it there would say it was to a
    # screen reader reading the group. It is a link rather than a button so it
    # works with the script dead, unlike Theme.
    #
    # It sits in mast-right, which held Theme as well until 2026-08-10, when
    # Theme moved to the page foot. The measurement that governed this corner
    # is worth keeping even though its pressure is off: at 375px the pills and
    # Theme filled their row to 360px against 359px of usable width, so adding
    # Search wrapped the bar to a THIRD sticky row (109px -> 157px, a fifth of
    # an 812px viewport). Theme leaving and the pills dropping to bare numerals
    # both give width back -- re-measure before spending it.
    #
    # Do NOT try to buy a row back by shaving gaps: 44px is --tap, the floor,
    # and this file's own history (the 2.9px overrun comment above) is the
    # argument against living on a 3px margin. Its label wears .nav-word so the
    # ≤26rem rule hides it, reused rather than reinvented; below that width the
    # glyph alone stands in and the accessible name stays "Search".
    search = (f'<a class="mast-btn" href="{prefix}search/">'
              f'<span class="nav-word">Search</span>{SEARCH_GLYPH}</a>')
    return f"""<header class="masthead">
  {mark}
  <nav aria-label="Tables">{links}</nav>
  <span class="mast-right">
    {search}
  </span>
</header>"""


# The theme control, at the FOOT of every page (user, 2026-08-10; it rode in
# the masthead until then). Bare "Theme" in the markup: the server cannot know
# which palette the reader will resolve to, and applyTheme() names it on the
# first tick. It must not say Auto -- there is no Auto state, and this label is
# what ships in the HTML and shows in the moment before the script runs.
#
# One string for all three page types, so the control cannot drift between
# them. Authored hidden and unhidden by the script, exactly as it was in the
# bar: with the script dead a reader gets the light palette and no dead
# control, which is the whole reason Theme is a <button> and Search is an <a>.
THEME_FOOT = ('<div class="theme-foot">'
              '<button id="theme" class="mast-btn" data-action="theme" hidden>'
              'Theme</button></div>')


def ruler_html(spec, n_gens):
    """
    The generation ruler: a column-header row that pans with the plate. Sized
    from the same --col/--stub/--sheet-pad tokens as the grid, so alignment is
    exact by construction. Decorative to AT (the chart states its generation
    count in the region label), hence aria-hidden.
    """
    chip = f"{spec['plate']} &middot; Genealogy {spec['numeral']}"
    # "Generation 1", spelled out. Note this is the ONLY "Gen." in the project
    # that means a generation column: every other one lives in transcription*.py
    # and is 1923 plate text meaning Genealogy II/III -- a cross-reference, and
    # not ours to reword.
    gens = "".join(f'<span class="gen">Generation {i}</span>'
                   for i in range(1, n_gens + 1))
    return ('<div class="ruler" aria-hidden="true">'
            f'<span class="ruler-chipslot"><span class="ruler-chip">{chip}</span></span>'
            f"{gens}</div>")


def datalist_html(persons, drawn):
    """Finder suggestions: number, printed name, clan. Baseline fields only.
    Persons not drawn on the chart (unreachable from the founding couples --
    a reported data condition) are omitted: the finder cannot jump to them."""
    # The numbers the plate prints on more than one person. Three times in the
    # whole edition (II 101, III 258, III 259), and the finder has to be told
    # about them or a reader holding the plate can reach only one of the two --
    # see the `data-n` note below.
    seen_n, shared = set(), set()
    for pid in persons:
        if pid in drawn:
            n = str(persons[pid]["plate_number"])
            (shared if n in seen_n else seen_n).add(n)

    opts = []
    for pid in sorted(persons):
        if pid not in drawn:
            continue
        p = persons[pid]
        nm = p["name_as_printed"] or "———"
        if p["alt_name"]:
            nm += f" ({p['alt_name']})"
        # The LABEL carries the plate's number, the VALUE the id. The script
        # jumps to "#p" + value, so the value has to stay the id or a name match
        # lands on the wrong person; the label is what the reader reads, so it
        # has to be what the plate prints. They differ only where the plate
        # reuses a number, and the suggestion list is the last place the
        # synthetic id was still visible.
        label = f"{p['plate_number']} · {nm}" + (f" · {p['clan']}" if p["clan"] else "")
        # `data-n` is the number the plate PRINTS, and it is emitted only on the
        # handful of options that need it: the script reads `dataset.n ||
        # value`, so for the other ~710 the id already is the printed number and
        # the markup is unchanged. Both members of a colliding pair carry it,
        # including the one whose number and id agree, so the pair is symmetric
        # and the script never has to reason about which kind it is holding.
        # `data-nm` rides with it because the note names the OTHER person, and
        # digging that name back out of the label would be parsing our own
        # prose -- the mistake `_p()` exists to prevent. It is the label minus
        # the number, clan included: both of III 259's are unnamed, so the name
        # alone would offer the reader a choice between "———" and "———".
        extra = ""
        if str(p["plate_number"]) in shared:
            who = nm + (f" · {p['clan']}" if p["clan"] else "")
            extra = f' data-n="{esc(str(p["plate_number"]))}" data-nm="{esc(who)}"'
        opts.append(f'<option value="{pid}"{extra}>{esc(label)}</option>')
    return '<datalist id="persons-list">' + "".join(opts) + "</datalist>"


def register_html(persons, unions, ku, km, drawn, paternity=None):
    """
    The Register of persons: a generated index, one entry per person in plate
    order, with parents, spouses and children as #p{n} links. It is the no-JS
    search, the no-JS person card, and the single source the person-card
    popover clones -- never a second truth.

    Research keys are blanked on a copy before person_line runs, so the
    English-name and census chips can never render here, in either build.
    A person the chart did not draw (`drawn` is Chart.seen) still gets an
    entry -- the register records everyone -- but no #p link, because its
    target would not exist.
    """
    parents = {}
    for u in unions:
        for k in ku.get(u["union_id"], []):
            parents[k] = (u["wife"], u["husband"])
    for m, kids in km.items():
        for k in kids:
            parents[k] = (m, 0)

    # The apparatus's own view of the children. `paternity` moves a child out
    # of its mother-only group into the union of that mother with the named
    # father. ku/km themselves are untouched, so the CHART is unaffected: it
    # keeps drawing the one bracket the plate draws. Only this copy is regrouped.
    pat = paternity or {}
    union_of = {(u["wife"], u["husband"]): u["union_id"]
                for u in unions if u["wife"] and u["husband"]}
    ku_a = {k: list(v) for k, v in ku.items()}
    km_a, editorial_unions = {}, set()
    for m, kids in km.items():
        rest = []
        for k in kids:
            uid = union_of.get((m, pat.get(k, 0)))
            if uid:
                ku_a.setdefault(uid, []).append(k)
                editorial_unions.add(uid)
            else:
                rest.append(k)
        if rest:
            km_a[m] = rest

    # The attribution read from the CHILD's end as well. Without this the
    # apparatus stated it in one direction only: 48 and 49 both listed 116-118,
    # while 116's own entry named just her mother, so a reader who came to Julia
    # first never learned the edition had supplied her a father.
    #
    # The dagger goes on the ROW, not on 49's chip, and the two are not the same
    # claim. What is editorial here is the PAIRING -- that this child belongs to
    # 48's marriage with 49 rather than with 47 -- exactly as it is on the
    # Children rows that already carry a row-level mark. The mother is never in
    # doubt; she is the plate's own bracket, and the footnote the dagger links
    # to says so in its first sentence. Marking 49 alone would also strand the
    # person card, whose rows are anchors: an <a> dagger cannot nest inside the
    # chip's <a>, and a bare span would give the card an unclickable mark.
    parents_ed = set()
    for k, f in pat.items():
        m = parents.get(k, (0, 0))[0]
        if m and f and union_of.get((m, f)):
            parents[k] = (m, f)
            parents_ed.add(k)

    spouses, children = {}, {}
    for u in unions:
        w, h = u["wife"], u["husband"]
        if w and h:
            spouses.setdefault(w, []).append(h)
            spouses.setdefault(h, []).append(w)
        kids = sorted(ku_a.get(u["union_id"], []))
        if kids:
            ed = u["union_id"] in editorial_unions
            if w:
                children.setdefault(w, []).append((h, kids, ed))
            if h:
                children.setdefault(h, []).append((w, kids, ed))
    for m, kids in km_a.items():
        children.setdefault(m, []).append((0, kids, False))

    def rel_link(pid):
        p = persons[pid]
        if p["name_as_printed"]:
            nm = f'<span class="name" lang="kjq">{esc(p["name_as_printed"])}</span>'
        else:
            nm = '<span class="blank">———</span>'
        # One element per person either way. A person the chart did not draw has
        # no #p target, so it cannot be a link -- but the person card enumerates
        # these as rows, and loose sibling spans cannot be enumerated. Both
        # forms therefore carry .rel-x, and the card checks the tag, not the
        # class, to decide whether a row is navigable.
        # Shown number is the plate's; the href and the key stay the id.
        shown = p["plate_number"]
        if pid not in drawn:
            return f'<span class="rel-x"><span class="num">{shown}</span> {nm}</span>'
        return (f'<a class="rel-x" href="#p{pid}">'
                f'<span class="num">{shown}</span> {nm}</a>')

    def rel_row(label, links, kind, with_id=None, editorial=False):
        """
        One labelled relation row.

        `kind` and `with_id` are written for the person card, which regroups
        these rows into a column per spouse: they let it pair a group of
        children with the other parent WITHOUT reading the label. The label is
        prose -- "Children (with 66)" -- and digging a person number out of
        prose is exactly the mistake `_p()` exists to prevent. Both attributes
        are invisible; the register renders as it always has.
        """
        w = "" if with_id is None else f' data-with="{with_id}"'
        # The dagger says "this grouping is ours, not the plate's" and links to
        # the note that says so. data-editorial carries the same fact to the
        # person card, which builds its own headings.
        ed_a = ('<a class="edmark" href="#note-paternity"'
                ' title="editorial attribution">&dagger;</a>') if editorial else ""
        ed_d = ' data-editorial="1"' if editorial else ""
        return (f'<div class="reg-rel" data-rel="{kind}"{w}{ed_d}>'
                f'<span class="rel-l">{label}</span>{ed_a} '
                + " &middot; ".join(links) + "</div>")

    items = []
    for pid in sorted(persons):
        p = dict(persons[pid])
        p.update({k: "" for k in RESEARCH_KEYS})
        rows = []
        pm, pf = parents.get(pid, (0, 0))
        par = [rel_link(x) for x in (pm, pf) if x]
        if par:
            rows.append(rel_row("Parents", par, "parents",
                                editorial=pid in parents_ed))
        sp = [rel_link(x) for x in spouses.get(pid, [])]
        if sp:
            rows.append(rel_row("Spouse" if len(sp) == 1 else "Spouses",
                                sp, "spouses"))
        groups = children.get(pid, [])
        for other, kids, ed in groups:
            if not other:
                label = "Children (father not recorded)"
            elif len(groups) > 1 or ed:
                # `or ed`: an editorial group NAMES the parent even when it is
                # the person's only one. Genealogy II's 48 has both her
                # husbands' children in a single attributed group, so without
                # this the row would read a bare "Children" with a dagger and
                # the reader would have to open the note to learn which
                # marriage is being asserted. It changes exactly one row on
                # Table 1 as well -- person 69, whose only group is the
                # editorial one; 68 and 70 each have two groups and were
                # already naming the parent.
                #
                # plate_number, not `other`: this is a number SHOWN, and the id
                # is a key. No duplicate-numbered person is a spouse on any
                # plate today, so the two agree everywhere -- which is exactly
                # why this path survived the 2026-07-29 sweep that fixed four
                # others. It is written correctly now rather than left to bite.
                label = f"Children (with {persons[other]['plate_number']})"
            else:
                label = "Children"
            rows.append(rel_row(label, [rel_link(k) for k in kids],
                                "children", other, ed))
        xr = persons[pid]["cross_ref"]
        if xr:
            xp = "Gen." in xr
            for part in xr.split("|"):
                rows.append('<div class="reg-rel"><em>'
                            + linkify_xref(esc(part.strip()), persons,
                                           cross_plate=xp)
                            + "</em></div>")
        line = person_line(p, False, set())
        if pid not in drawn:
            # No chart line to link to: demote the number anchor to a span.
            line = line.replace(f'<a class="num" href="#p{pid}">',
                                '<span class="num">', 1).replace("</a>", "</span>", 1)
        items.append(f'<li class="reg" id="r{pid}">'
                     f'<div class="reg-line">{line}</div>'
                     + "".join(rows) + "</li>")

    return f"""<section class="apparatus-register" id="register-sec" aria-labelledby="reg-h">
  <h2 id="reg-h">Register of persons</h2>
  <p class="reg-note">Editorial apparatus, generated from the transcription.
     Numbers link back to the chart; the plate itself is unchanged.</p>
  <details id="register" class="register-d">
    <summary>All {len(persons)} entries</summary>
    <ol class="reg-list">{"".join(items)}</ol>
  </details>
</section>"""


def cite_html(spec, today):
    """
    The recommended two-part citation, generated so it can never go stale.

    The <h2> is the caller's, not this function's: since 2026-08-09 the section
    is a disclosure and the heading lives inside its <summary>.
    """
    canonical = f"{SITE}/{spec['slug']}/"
    root_id = spec["roots"][0]
    return f"""<blockquote class="cite-block" id="cite-text">
    <p>Elsie Clews Parsons, &ldquo;Laguna Genealogies,&rdquo;
       <em>Anthropological Papers of the American Museum of Natural History</em>,
       vol.&nbsp;19, pt.&nbsp;5 (1923), pp.&nbsp;133&ndash;292, {spec['plate']}.</p>
    <p>Digital transcription: {esc(AUTHOR)},
       <em>Laguna Genealogies: A Digital Edition</em>, {today.year},
       {canonical}. CC&nbsp;BY&nbsp;4.0.</p>
  </blockquote>
  <p>To cite one person&rsquo;s line, use its number&rsquo;s link:
     <a href="#p{root_id}">{canonical}#p{root_id}</a> is person {root_id}.<span id="copy-mount"></span></p>"""


def jsonld_chart(spec, description, today):
    """Bibliographic structured data -- public builds only, counts computed."""
    canonical = f"{SITE}/{spec['slug']}/"
    data = {
        "@context": "https://schema.org",
        "@type": "Dataset",
        "name": (f"Genealogy {spec['numeral']} — Parsons 1923, Laguna Pueblo: "
                 "a digital transcription"),
        "description": description,
        "url": canonical,
        "creator": {"@type": "Person", "name": AUTHOR},
        # No "identifier": it held the Zenodo doi until 2026-08-08. The field
        # is optional to Google's Dataset guidance and the build's own
        # check_structured_data() does not require it, so dropping it costs
        # nothing and keeps the archive unadvertised.
        "license": "https://creativecommons.org/licenses/by/4.0/",
        "isBasedOn": {"@type": "CreativeWork", "name": "Laguna Genealogies",
                      "author": {"@type": "Person", "name": "Elsie Clews Parsons"},
                      "datePublished": "1923"},
        # Google's Dataset validator rejects a WebSite here: the "belongs to a
        # larger collection" relation is includedInDataCatalog + DataCatalog,
        # even though Dataset.isPartOf accepts any CreativeWork in schema.org.
        "includedInDataCatalog": {"@type": "DataCatalog",
                                  "name": "Laguna Genealogies: A Digital Edition",
                                  "url": SITE + "/"},
        "dateModified": today.isoformat(),
        "image": OG_IMAGE,
        "inLanguage": ["en", "kjq"],
        "keywords": KEYWORDS,
        "temporalCoverage": "1870/1923",
        "spatialCoverage": {
            "@type": "Place",
            "name": "Laguna Pueblo, Valencia County, New Mexico, United States",
        },
        "citation": CITATION_TEXT,
        "distribution": {
            "@type": "DataDownload",
            "encodingFormat": "text/html",
            "contentUrl": canonical,
        },
    }
    return ('<script type="application/ld+json">'
            + json.dumps(data, ensure_ascii=False) + "</script>")


def check_structured_data(paths):
    """
    Walk every JSON-LD block in the built pages and assert that each Dataset --
    at any nesting depth -- carries the fields Google requires.

    This exists because a nested Dataset is validated as a Dataset in its own
    right, not as a pointer to one: the landing page's hasPart entries were once
    name-and-url stubs, and Search Console reported them as invalid items even
    though the page they pointed at was complete. A malformed block costs rich
    results silently -- nothing renders differently -- so it needs a build check
    rather than an eye.
    """
    required = ("name", "description", "url")

    # Google's Dataset validator is stricter than schema.org about what type may
    # sit in each field. isPartOf: {"@type": "WebSite"} is legal schema.org and
    # was still reported as an invalid object type -- the collection relation it
    # wants is includedInDataCatalog + DataCatalog. Only fields this build emits
    # are listed; add a row when a new one is emitted.
    field_types = {
        "creator": ("Person", "Organization"),
        "includedInDataCatalog": ("DataCatalog",),
        "isBasedOn": ("CreativeWork", "Book", "ScholarlyArticle"),
        "distribution": ("DataDownload",),
        "spatialCoverage": ("Place",),
        "publisher": ("Person", "Organization"),
    }
    problems = []

    def walk(obj):
        if isinstance(obj, dict):
            if obj.get("@type") == "Dataset":
                who = obj.get("name", "(unnamed)")
                missing = [f for f in required if not obj.get(f)]
                if missing:
                    problems.append(f"{who}: missing {missing}")
                for field, allowed in field_types.items():
                    val = obj.get(field)
                    if isinstance(val, dict) and val.get("@type") not in allowed:
                        problems.append(
                            f"{who}: {field} is {val.get('@type')!r}, "
                            f"expected one of {list(allowed)}")
            for v in obj.values():
                walk(v)
        elif isinstance(obj, list):
            for v in obj:
                walk(v)

    n = 0
    for path in paths:
        html = path.read_text(encoding="utf-8")
        for block in re.findall(
                r'<script type="application/ld\+json">(.*?)</script>', html, re.S):
            try:
                data = json.loads(block)
            except json.JSONDecodeError as e:
                problems.append(f"{path.name}: JSON-LD does not parse -- {e}")
                continue
            n += 1
            walk(data)

    if problems:
        print("  STRUCTURED DATA INVALID:")
        for p in problems:
            print(f"    {p}")
        return False
    print(f"  {n} JSON-LD blocks valid")
    return True


def jsonld_breadcrumb(spec):
    """Breadcrumbs, so a result shows 'Laguna Genealogies > Genealogy N'."""
    data = {
        "@context": "https://schema.org",
        "@type": "BreadcrumbList",
        "itemListElement": [
            {"@type": "ListItem", "position": 1,
             "name": "Laguna Genealogies", "item": SITE + "/"},
            {"@type": "ListItem", "position": 2,
             "name": f"{spec['plate']} — Genealogy {spec['numeral']}",
             "item": f"{SITE}/{spec['slug']}/"},
        ],
    }
    return ('<script type="application/ld+json">'
            + json.dumps(data, ensure_ascii=False) + "</script>")


def jsonld_site(built, today):
    data = {
        "@context": "https://schema.org",
        "@type": "CollectionPage",
        "name": "Laguna Genealogies: A Digital Edition",
        "url": SITE + "/",
        "description": SITE_DESCRIPTION,
        "creator": {"@type": "Person", "name": AUTHOR},
        # No "identifier" -- see jsonld_chart. This page carried the doi most
        # precisely of any, since the deposit covered the whole edition; that
        # is exactly why it is the one that had to lose it.
        "license": "https://creativecommons.org/licenses/by/4.0/",
        "dateModified": today.isoformat(),
        "image": OG_IMAGE,
        "inLanguage": ["en", "kjq"],
        "keywords": KEYWORDS,
        "about": {"@type": "Place", "name": "Laguna Pueblo, New Mexico"},
        "citation": CITATION_TEXT,
        # Each part is validated as a Dataset in its own right, not as a
        # pointer, so a name-and-url stub fails on the required "description".
        # These carry the same generated description, creator and licence as the
        # table's own page -- one source, so they cannot disagree.
        "hasPart": [
            {"@type": "Dataset",
             "name": f"Genealogy {spec['numeral']} — Parsons 1923",
             "url": f"{SITE}/{spec['slug']}/",
             "description": describe(spec, st["persons"], st["gens"]),
             "creator": {"@type": "Person", "name": AUTHOR},
             "license": "https://creativecommons.org/licenses/by/4.0/",
             "includedInDataCatalog": {"@type": "DataCatalog",
                                       "name": "Laguna Genealogies: A Digital Edition",
                                       "url": SITE + "/"}}
            for spec, st in built
        ],
    }
    return ('<script type="application/ld+json">'
            + json.dumps(data, ensure_ascii=False) + "</script>")


def build_doc(spec, description, gens, n_gens, trees, status, public, today,
              stats, persons, unions, ku, km, drawn):
    """The full HTML document. Identical in both modes apart from the footer,
    the meta, and the masthead's link base (the private build lives outside
    the site tree, so its links are absolute)."""
    canonical = f"{SITE}/{spec['slug']}/"
    tables = [(TABLES[k]["numeral"], TABLES[k]["slug"]) for k in sorted(TABLES)]
    if public:
        social = social_meta(
            f"Genealogy {spec['numeral']} &mdash; Parsons 1923, Laguna Pueblo",
            description, canonical)
        head_extra = f"""<link rel="canonical" href="{canonical}">
<meta name="description" content="{esc(description)}">
<meta name="author" content="{esc(AUTHOR)}">
{social}
{jsonld_chart(spec, description, today)}
{jsonld_breadcrumb(spec)}"""
        provenance = f"""    <section class="app-sec">
  <details class="app-d">
    <summary><h2>Provenance</h2></summary>
  <ul>
    <li>This is a transcription of a plate in an existing published source. The
        1923 publication is in the public domain in the United States.</li>
    <li>The transcription, encoding and layout are by {esc(AUTHOR)}, released under
        <a href="https://creativecommons.org/licenses/by/4.0/">CC&nbsp;BY&nbsp;4.0</a>.
        Please cite both this edition and Parsons.</li>
    <li>Readings can be checked against the source scan, which is kept in the
        <a href="{REPO}">project repository</a>
        together with the data and the code that drew this page.</li>
    <li>Names are set in a subset of SIL Gentium, embedded in this page under the
        <a href="../fonts/OFL.txt">SIL Open Font License</a>, so the phonetic
        diacritics render the same everywhere.</li>
  </ul>
  </details>
    </section>"""
        mast = masthead_html(tables, spec["slug"], "../", "../")
        canon_attr = f' data-canonical="{canonical}"'
    else:
        head_extra = '<meta name="robots" content="noindex,nofollow">'
        provenance = """    <section class="app-sec">
  <details class="app-d">
    <summary><h2>This is the private build</h2></summary>
  <ul>
    <li>Generated from <code>data/parsons_genealogy_I.xlsx</code>, so it may show
        English names and census matches. It is git-ignored and must not be published.</li>
    <li>Edit the workbook, then re-run <code>python3 scripts/make_chart.py</code>.</li>
    <li>The public page is built separately with <code>--public</code>, from
        <code>scripts/transcription.py</code>, which has no research columns.</li>
  </ul>
  </details>
    </section>"""
        mast = masthead_html(tables, spec["slug"], f"{SITE}/", f"{SITE}/")
        canon_attr = ""

    imprint = (f"{stats['persons']} individuals &middot; {stats['gens']} generations "
               f"&middot; {stats['unions']} marriages &middot; {stats['links']} "
               "parent&ndash;child links")

    return f"""<!doctype html>
<html lang="en"{canon_attr}>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Genealogy {spec['numeral']} &mdash; Parsons 1923, Laguna Pueblo{"" if public else " (private build)"}</title>
<meta name="theme-color" content="#FAF8F4">
<link rel="icon" href="{FAVICON}">
{head_extra}
<style>{font_css()}{geom_css()}{CSS}</style>
{THEME_SNIPPET}
</head>
<body class="chart">
<a class="skip" href="#plate">Skip to chart</a>
<a class="skip" href="#register-sec">Skip past chart to the register</a>
{mast}
<div class="titlepage">
  <div class="plate-label">{spec['plate']}</div>
  <h1>GENEALOGY {spec['numeral']}</h1>
  <div class="rule-double"></div>
  <p class="imprint">{imprint}</p>
</div>
<main>
<div class="plate-bar">
  <span class="plate-tools">
    <search><form id="find" hidden>
      <label class="visually-hidden" for="find-q">Find a person by number or name</label>
      <input id="find-q" list="persons-list" autocomplete="off"
             placeholder="Find: number or name ( / )" enterkeyhint="go">
      <span id="find-note" role="status" class="find-note"></span>
    </form></search>
    <span id="scale-mount"></span>
  </span>
</div>
{datalist_html(persons, drawn)}
<figure class="plate">
  <div class="scroll-shell">
    <div class="scroll" id="plate" tabindex="0" role="region"
         aria-label="Genealogy {spec['numeral']} chart, {gens} generations, scrollable horizontally; use the arrow keys to scroll &mdash; person numbers are links">
      <div class="plate-zoom" style="--print-zoom:{print_zoom(n_gens)}">
        {ruler_html(spec, n_gens)}
        <div class="sheet">{trees}</div>
      </div>
    </div>
  </div>
  <figcaption class="plate-caption"><span class="pan-hint">Scroll sideways to
    follow the later generations &mdash; person numbers are links.</span></figcaption>
</figure>
</main>
{register_html(persons, unions, ku, km, drawn, spec.get("paternity"))}
<footer id="apparatus">
  <div class="app-cols">
    <section class="app-sec">
      <h2>The record</h2>
      <ul>{"".join(status)}</ul>
    </section>
    <section class="app-sec">
      <h2>Navigating this chart</h2>
      <ul>{navigating_html(spec)}</ul>
    </section>
    <section class="app-sec">
      <details class="app-d">
        <summary><h2>Editorial notes</h2></summary>
      <ul>{READING_COMMON.format(blocks=blocks_note(spec))}{spec["notes"]}{APPARATUS_NOTE}</ul>
      </details>
    </section>
{provenance}
    <section class="app-sec">
      <details class="app-d">
        <summary><h2>Citation</h2></summary>
      {cite_html(spec, today)}
      </details>
    </section>
  </div>
  <p class="updated">Last updated
     <time datetime="{today.isoformat()}">{today.strftime("%-d %B %Y")}</time>.
     {f'<span class="print-url">Published at {canonical}</span>' if public else ''}</p>
</footer>
{THEME_FOOT}
<div id="pcard" class="pcard" popover role="dialog" aria-labelledby="pcard-t" tabindex="-1" hidden></div>
<script>{UI_JS}</script>
</body>
</html>
"""


LANDING_CSS_EXTRA = """
.contents{max-width:var(--measure);margin:0 auto;padding:var(--s4) var(--s5)}
.contents ol{list-style:none;margin:0;padding:0}
.contents li{border-block-end:1px solid var(--rule-faint)}
.contents li:first-child{border-block-start:1px solid var(--rule-faint)}
.contents a{display:block;padding:var(--s4) var(--s2);text-decoration:none;
  color:inherit}
.c-title{display:block;font-size:var(--t-md);color:var(--ink)}
.contents a:hover .c-title{color:var(--accent)}
.contents a:focus-visible .c-title{color:var(--accent)}
.c-stats{display:block;font-size:var(--t-sm);color:var(--muted);
  margin-block-start:var(--s1);line-height:1.6}
.contents .pending{padding:var(--s4) var(--s2)}
.contents .pending .c-title{color:var(--muted)}
.pending-tag{font-style:italic;font-size:var(--t-base)}
/* The finding aid is ruled into the contents list without being one of its
   items -- a plate's rule comes from .contents li, so this row carries its
   own. Everything else it needs (block display, padding, hover, focus) it
   already gets from .contents a. */
.c-across{border-block-end:1px solid var(--rule-faint)}
.prose{max-width:var(--measure);margin:0 auto;
  padding:var(--s6) var(--s5) var(--s7);font-size:var(--t-prose);line-height:1.7;
  color:var(--muted);text-wrap:pretty}
.prose h2{margin:var(--s6) 0 var(--s3)}
.prose a{color:var(--accent)}
.prose code{font-size:.9em}
/* The FAQ. Open by default would bury the contents list, so each answer is a
   disclosure -- but the text is in the DOM either way, which is what the
   FAQPage structured data requires and what a crawler reads. */
.faq{border-block-end:1px solid var(--rule-faint)}
.faq:first-of-type{border-block-start:1px solid var(--rule-faint)}
.faq summary{padding:var(--s3) 0;cursor:pointer;color:var(--ink);
  list-style:none;display:flex;gap:var(--s3);align-items:baseline}
.faq summary::-webkit-details-marker{display:none}
.faq summary::before{content:"+";color:var(--muted);font-variant-numeric:tabular-nums}
.faq[open] summary::before{content:"\\2013"}
.faq summary:hover{color:var(--accent)}
.faq p{margin:0 0 var(--s4) var(--s4)}
/* The theme control closes the page on the same rail as the page. A table
   page's apparatus is --measure-wide and the control matches it there by
   default; here the contents list and the prose are both --measure, and a
   closing rule 32rem wider than everything above it reads as a stray element
   rather than as the end of the page. */
.theme-foot{max-width:var(--measure)}
"""


# The two projects store the reader's palette under different keys -- this site
# under "lg-theme", the widget's own default being "laguna-theme" -- and both
# drive the same html[data-theme]. Left alone, a reader who chose Dark on a
# chart page arrives at /search/ and is handed the system preference instead,
# and the same in reverse.
#
# From 2026-08-09 the widget takes the key as configuration, so there is one
# key and nothing to synchronise. This line is the whole of it: it declares the
# key the widget's two halves read -- `THEME_BOOT`, the blocking script that
# applies the palette before first paint, and `themeToggle()`, which writes it
# at mount. Both fall back to "laguna-theme" when it is unset, so the vendored
# page still works on its own.
#
# It must be injected AHEAD of the vendored boot script; everything else this
# function adds goes in at </head>, and putting this there too would leave the
# pre-paint read looking at the wrong key. It is spliced in after the charset
# meta rather than after `<head>` itself, so nothing comes between the document
# and the declaration that decodes it.
#
# This replaced a bridge that mirrored the two keys and carried changes back
# with a MutationObserver. Don't reintroduce one: a second key is the defect,
# not the starting condition.
#
# It carries a SECOND declaration since 2026-08-10: the default palette. Every
# other page here defaults to light in CSS, with no script involved (see the
# color-scheme note in CSS), but the widget's palette is keyed on
# html[data-theme] alone and its own default is the system preference -- so
# without this, choosing nothing gives a dark /search/ beside a light
# everything-else on a dark OS. Setting the attribute here rather than asking
# for another widget option is deliberate: defaulting to light is THIS site's
# decision, and the widget standing alone should keep following the reader's
# OS. It runs ahead of the vendored boot script, which only ever assigns when
# it finds a stored value, so a stored choice still wins.
THEME_KEY_DECL = (
    '<script>window.LAGUNA_THEME_KEY="lg-theme";'
    'document.documentElement.dataset.theme="light"</script>')


def write_search(tables):
    """
    docs/search/ -- the cross-plate finding aid, from vendor/search/.

    That directory is another project's build output (vendor/search/SOURCE.md),
    so this function does the least it can to it: it never rewrites the widget,
    only wraps it. Six things are added, and each is here because the vendored
    file cannot supply it.

    1. THE FONT. `search.css` declares no @font-face at all, and every name it
       shows is Americanist phonetic -- `ʼ` `˙` `ᶦ` `ᵘ` `ᵃ` `ʽ`. The rest of the
       site inlines the subset for exactly this reason, so the same faces are
       injected and prepended to the widget's own `--lg-serif` stack.

       Nothing downstream can catch a regression here. subset_font.py's
       coverage check reads the TEXT of built pages, and the names arrive from
       search-index.json at runtime, so they appear in no HTML file and that
       check sees an empty page. If this injection is ever dropped, the page
       keeps working and silently substitutes.

    2. THE BAR. The widget draws a title block and no navigation -- it was
       built to be mounted in a host page that already had some. Without this a
       reader who lands on /search/ from a link has no route into the edition.
       It is scoped `.lg-host-bar` and takes its COLOURS from the widget's own
       `--lg-*` tokens, so it follows the theme toggle and cannot collide with
       the widget, which is scoped `.laguna-search`.

       Its METRICS are the masthead's, lifted out of `CSS` (user, 2026-08-10:
       make the bar uniform on every page). It used to restate them in round
       numbers and was 69px tall against the masthead's 49px, with a 44px hit
       floor where the site's is 32px on a mouse, a bare `system-ui` stack, a
       16px inset against 8px -- and no Search control at all. Every one of
       those now comes from the site's own tokens, emitted under the site's own
       names. Search is present too, marked `aria-current` rather than linked:
       a bar that drops a control on one page is not uniform.

    3. `--lg-sticky-top`. The widget's filter header is sticky; the token is its
       documented hook for "the host page has a bar this tall". Set it or the
       header comes to rest underneath ours.

    4. THE THEME KEY. One line, `THEME_KEY_DECL`, and it goes in at the top of
       <head> rather than at the bottom with the rest -- see the note there.
       Without it the widget stores the palette under its own key while every
       other page here stores it under "lg-theme", and a reader's choice is
       dropped in whichever direction they crossed the boundary.

    5. THE HEADING'S TYPE. The widget sets its h1 from its own ramp, because it
       was built to stand alone: clamp(2.1rem,5vw,4rem) at .1em, which is 64px
       at 1280px against this edition's 40px, and 40px on a phone against
       25.6px. On this site it is one page of seven, so it takes the site's own
       h1 size, letter-spacing and line-height.

       All three are READ OUT OF `CSS`'s h1 rule rather than restated. That is
       the point of doing it this way: the ramp is one literal, so it can move
       without leaving /search/ behind, and there is no second copy to notice.
       Missing any of the three fails the build rather than shipping a heading
       that is a step off every other page.

    6. THE REST OF THE TITLE BLOCK -- the standfirst and the double rule
       between it and the heading (user, 2026-08-10). Same argument as the h1:
       the widget sizes both for a page where its title block is the whole of
       the page. Its `.lede` is 1.25rem/1.45, 20px, against the table pages'
       statistics line at `--t-base`/1.6, 16px, under the same h1; its `.rule`
       is 452px wide and 7px deep in accent gold against `.rule-double`'s 8rem,
       4px and ink. Both are read whole out of `CSS`, and the rule's ink token
       is substituted for the widget's, exactly as the bar's colours are. Same
       abort if either rule stops stating what is read from it.

    Note what is NOT in this list. The search card's one-line control row and
    the list keeping its columns at every width were both wanted here on
    2026-08-10, and both went UPSTREAM instead, into laguna-search's own
    stylesheet -- they are that widget's layout, not this host's, and an
    override of another project's media queries is a thing to re-read on every
    re-vendor rather than a thing to own. The h1 above is the counter-example
    and the test to apply: it is host-specific by nature, because the widget
    standing alone should keep its own ramp. If a wanted change would look
    wrong on the widget's own site too, it belongs upstream.

    Returns False if the vendored files are missing, which fails the build. That
    is deliberate: METHOD.md's *Identity across plates* describes this page in
    the present tense, so an edition that ships without it is describing
    something that 404s.
    """
    need = ("index.html", "search.js", "search-index.json")
    missing = [n for n in need if not (SEARCH_DIR / n).exists()]
    if missing:
        print(f"ABORTED: vendor/search/ incomplete -- missing {', '.join(missing)}")
        print("  See vendor/search/SOURCE.md; the search page cannot be built.")
        return False

    # The site's own rules, taken apart so /search/ can be given the exact
    # declarations that decide how it reads. No value is typed twice: every
    # number below is lifted out of `CSS`, so the ramp and the spacing scale
    # can move without leaving this page behind.
    def site_decls(pattern):
        """The declarations of one rule in `CSS`, comments stripped."""
        m = re.search(pattern, CSS)
        body = re.sub(r"/\*.*?\*/", "", m.group(1), flags=re.S) if m else ""
        out = {}
        for d in body.replace("\n", " ").split(";"):
            prop, sep, val = d.partition(":")
            if sep:
                out[prop.strip()] = val.strip()
        return out

    def take(decls, props, what, why):
        """`props` out of `decls`, or abort the build naming what went."""
        missing = [p for p in props if p not in decls]
        if missing:
            print(f"ABORTED: the site's {what} no longer states " +
                  ", ".join(missing) + "; " + why)
            return None
        return "".join(f"{p}:{decls[p]};" for p in props)

    h1_type = take(site_decls(r"(?m)^h1\{([^}]*)\}"),
                   ("font-size", "letter-spacing", "line-height"),
                   "h1 rule",
                   "/search/'s heading would drift off every other page's")
    if h1_type is None:
        return False

    # The standfirst. A table page sets its statistics line at --t-base/1.6
    # directly under the same h1; the widget's own lede is 1.25rem/1.45, built
    # for a page where it is the only thing under the title. Same title block,
    # same size (user, 2026-08-10).
    lede_type = take(site_decls(r"(?m)^\.imprint\{([^}]*)\}"),
                     ("font-size", "line-height"),
                     ".imprint rule",
                     "/search/'s standfirst would drift off the table pages'")
    if lede_type is None:
        return False

    # The double rule under the title. The widget draws its own -- 452px wide,
    # 7px deep, in accent gold -- because it was built to stand alone; a table
    # page's is 8rem, 4px and ink, and it is the same mark under the same
    # heading (user, 2026-08-10: match the lines under the header title with
    # the tables' lines). Read whole out of `.rule-double`, colour token
    # substituted, exactly as the bar's rules are.
    rule_double = take(site_decls(r"(?m)^\.rule-double\{([^}]*)\}"),
                       ("width", "height", "margin",
                        "border-block-start", "border-block-end"),
                       ".rule-double rule",
                       "/search/'s rule would drift off the table pages'")
    if rule_double is None:
        return False
    rule_double = rule_double.replace("var(--ink)", "var(--lg-ink)")

    # The bar's metrics, straight off the masthead's own tokens. The names are
    # kept AS THEY ARE rather than namespaced, so every rule below is the
    # masthead's text with only the selectors and the colour tokens changed --
    # a diff between the two is then a real difference and not a rename.
    root = site_decls(r"(?ms)^:root\{(.*?)^\}")
    bar_tokens = ("--font-ui", "--tap", "--bar-h",
                  "--s1", "--s2", "--s4", "--t-xs")
    token_css = take(root, bar_tokens, ":root block",
                     "/search/'s bar could not be built from the site's own "
                     "spacing and type scale")
    if token_css is None:
        return False

    coarse = re.search(r"@media\s*\(pointer:coarse\)\{:root\{([^}]*)\}\}", CSS)
    if not coarse or "--tap" not in coarse.group(1):
        print("ABORTED: the site's (pointer:coarse) --tap override is gone; "
              "/search/'s bar would keep a 32px hit area on a touch screen")
        return False

    html = (SEARCH_DIR / "index.html").read_text(encoding="utf-8")

    # The tokens above are emitted under the SITE's names, which is only safe
    # while the widget keeps every one of its own behind `--lg-`. It does, and
    # this is the check that says so on the day it stops.
    clash = [t for t in bar_tokens if t in html]
    if clash:
        print("ABORTED: vendor/search/ now declares " + ", ".join(clash) +
              "; /search/'s bar would overwrite the widget's own token")
        return False

    # The host bar is only pinned while panning because we widen `body` below,
    # and the widget's own `body` rule sets margin, background and colour but
    # no width. The day it sets one, source order decides which wins and the
    # bar silently goes back to sliding off the left edge -- which reads as a
    # sticky bug and is not one.
    if re.search(r"body\s*\{[^}]*\bwidth\s*:", html):
        print("ABORTED: vendor/search/ now sets a width on `body`; "
              "/search/'s host bar would stop covering the panned document")
        return False

    out = DOCS / "search"
    out.mkdir(parents=True, exist_ok=True)
    for name in ("search.js", "search-index.json"):
        shutil.copyfile(SEARCH_DIR / name, out / name)

    # The bar mirrors the site masthead, and since 2026-08-10 it mirrors it
    # MEASURABLY rather than approximately (user: make the bar uniform across
    # every page). It used to restate the metrics in round numbers -- 13px
    # type, 6px/16px padding, a 44px hit floor, no Search control -- which put
    # it 69px tall against the masthead's 49px, with the wordmark 8px further
    # in. Everything it needs now comes out of `CSS` above, so the two bars are
    # one bar wearing two sets of colour tokens.
    #
    # Search is here for the same reason: the masthead carries it on every
    # other page, and a bar that drops a control on one page is not uniform. It
    # is a <span> with aria-current, not a link -- the wordmark's own idiom for
    # "you are here" on the landing page, and a link to the page you are on is
    # a dead control.
    links = "".join(
        f'<a href="../{slug}/"><span class="lg-hb-word">Genealogy </span>{numeral}</a>'
        for numeral, slug in tables)
    bar = (
        '<header class="lg-host-bar">'
        '<a class="lg-hb-mark" href="../">Home</a>'
        f'<nav aria-label="Tables">{links}</nav>'
        '<span class="lg-hb-right">'
        '<span class="lg-hb-btn" aria-current="page">'
        f'<span class="lg-hb-word">Search</span>{SEARCH_GLYPH}</span>'
        "</span>"
        "</header>")

    # Every rule from `.lg-host-bar` to the print block is the masthead's own,
    # with two substitutions and nothing else: the selectors (.masthead ->
    # .lg-host-bar, .wordmark -> .lg-hb-mark, .mast-right -> .lg-hb-right,
    # .mast-btn -> .lg-hb-btn, .nav-word -> .lg-hb-word) and the colour tokens
    # (--paper/--ink/--muted -> --lg-*, --rule-faint -> --lg-rule, --rule ->
    # --lg-rule-strong). The colours have to be the widget's or the bar stops
    # following the theme; the METRICS are the site's, emitted above under the
    # site's own names, so the two bars cannot drift on size, spacing or the
    # touch floor. Diff this against the masthead section of CSS when either
    # changes -- that diff should show selectors and colours, nothing more.
    #
    # --lg-serif is a stack, not a family: prepending keeps every fallback the
    # widget chose. Georgia stays the second choice, as it is here.
    host_css = f"""{font_css()}:root{{
  --lg-serif:'Laguna Serif',Georgia,"Iowan Old Style","Times New Roman",serif;
  {token_css}
  /* The widget's documented hook for "the host has a bar this tall". +1px is
     the bar's bottom border, which --bar-h does not carry: without it the
     sticky filter header comes to rest a hairline high and the border shows
     through it. */
  --lg-sticky-top:calc(var(--bar-h) + 1px);
}}
@media (pointer:coarse){{:root{{{coarse.group(1)}}}}}
/* The site resets `box-sizing` globally (`*{{box-sizing:border-box}}`); the
   widget scopes its reset to `.laguna-search *`, and this bar is deliberately
   outside that. So without this line every metric below is read as
   content-box and the same declarations build a DIFFERENT bar: measured
   65px tall against the masthead's 49, pills 50x34 against 32x32 -- padding
   and border added on rather than absorbed. It is the one rule here that is
   not the masthead's, and it exists to make the rest of them the masthead's. */
.lg-host-bar,.lg-host-bar *{{box-sizing:border-box}}
/* The document PANS below 651px -- the list is a table at every width (user,
   2026-08-10) -- and `position:sticky` does not stick HORIZONTALLY. The bar's
   containing block is body's content box, which is only the viewport's width,
   so panning right to reach the `#` box slid the bar off to the left and left
   the top of the page bare. A table page has no such problem: there the pan is
   scoped to `.scroll` and the document itself never moves.
   Widening body is the whole fix -- the bar then fills a containing block as
   wide as the panned document, and no rule below changes. `fit-content` is
   min(max-content, max(min-content, available)), so it lands on exactly the
   min-content width the cards already overflow to, and `min-width:100%` keeps
   body the viewport's width once the viewport is the wider of the two.
   `max-content` is the WRONG tool and looks like the obvious one: it would
   size the table to its no-wrap width, blowing the pan out well past the
   cards' own minimum. Do not "simplify" this to it. */
body{{width:fit-content;min-width:100%}}
.lg-host-bar{{position:sticky;inset-block-start:0;z-index:40;
  min-height:var(--bar-h);
  display:flex;align-items:center;flex-wrap:wrap;gap:var(--s1) var(--s2);
  padding:var(--s2) calc(var(--s4) - var(--s2));
  background:var(--lg-paper);border-block-end:1px solid var(--lg-rule);
  font:400 var(--t-xs)/1.2 var(--font-ui);letter-spacing:.08em}}
/* One hit-area rule for every control in the bar, links included. */
.lg-hb-mark,.lg-host-bar nav a,.lg-hb-btn{{
  display:inline-flex;align-items:center;min-height:var(--tap);
  padding-inline:var(--s2);vertical-align:middle}}
.lg-host-bar a{{text-decoration:none}}
.lg-hb-mark{{text-transform:uppercase;letter-spacing:.14em;font-weight:600;
  color:var(--lg-ink)}}
a.lg-hb-mark:hover,a.lg-hb-mark:focus-visible{{color:var(--lg-accent)}}
.lg-host-bar nav{{display:flex;flex-wrap:wrap;gap:var(--s1);
  color:var(--lg-muted);text-transform:uppercase}}
/* Numeral alone at every width, the word hidden but kept in the accessible
   name -- the masthead's rule, restated here in the widget's colour tokens
   because this bar cannot reach the site's stylesheet. */
.lg-host-bar nav a{{justify-content:center;padding-inline:var(--s2);
  min-width:var(--tap);
  border:1px solid var(--lg-rule);border-radius:2px;
  color:var(--lg-muted);white-space:nowrap}}
.lg-host-bar nav a:hover,.lg-host-bar nav a:focus-visible{{color:var(--lg-ink);
  border-color:var(--lg-rule-strong)}}
.lg-host-bar nav .lg-hb-word{{position:absolute;width:1px;height:1px;
  margin:-1px;padding:0;overflow:hidden;clip-path:inset(50%);
  white-space:nowrap;border:0}}
/* Below 26rem the row is too tight for Search's word beside four pills, so
   the same hide-the-word mechanism runs there -- scoped to .lg-hb-right,
   because the nav's copy above is unconditional and the two must be able to
   move independently. Exactly the masthead's pair of selectors. */
@media (max-width:26rem){{
  .lg-hb-right .lg-hb-word{{position:absolute;width:1px;height:1px;margin:-1px;
    padding:0;overflow:hidden;clip-path:inset(50%);white-space:nowrap;border:0}}
}}
.lg-hb-right{{margin-inline-start:auto;display:flex;align-items:center;
  gap:var(--s2)}}
.lg-hb-btn{{text-transform:uppercase;color:var(--lg-muted);
  border:1px solid var(--lg-rule);border-radius:2px}}
/* You are here. The masthead's own current-page treatment -- a filled
   inversion, not a colour shift, so it survives both themes, a monochrome
   screen and colour-blind vision. */
.lg-hb-btn[aria-current="page"]{{background:var(--lg-ink);color:var(--lg-paper);
  border-color:var(--lg-ink);font-weight:600}}
/* The glyph is the narrow-screen half of the Search label, and it is the
   site's own SEARCH_GLYPH string, so it carries .mast-glyph. */
.lg-host-bar .mast-glyph{{display:none}}
@media (max-width:26rem){{
  .lg-host-bar .mast-glyph{{display:block}}
  .lg-hb-btn{{min-width:var(--tap);justify-content:center}}
}}
@media print{{.lg-host-bar{{display:none}}}}
/* The heading, on the site's ramp rather than the widget's -- see (5) above.
   This block follows the widget's whole stylesheet at equal specificity, so it
   also wins over the <=480px rule, which sets a size AND a letter-spacing of
   its own; that is the width where the two disagree most. */
.laguna-search h1{{{h1_type}}}
/* And the standfirst under it, on the table pages' .imprint size -- same
   title block, same measure. Same equal-specificity trick: the <=480px rule
   sets a size of its own. */
.laguna-search .lede{{{lede_type}}}
/* And the double rule between them, which is the table pages' own mark: 8rem
   and 4px deep in ink, not 452px and 7px in accent gold. The widget states it
   with the physical `border-top`/`border-bottom` and this states it with the
   logical pair, which is the same computed property -- later wins, and the
   <=480px rule that re-widens it is earlier still. */
.laguna-search .rule{{{rule_double}}}
"""

    charset = '<meta charset="utf-8">'
    if charset not in html:
        print("ABORTED: vendor/search/index.html has no charset meta to anchor "
              "the theme key to; the palette would not survive /search/")
        return False
    html = html.replace(charset, f"{charset}\n{THEME_KEY_DECL}", 1)
    html = html.replace(
        "</head>",
        f'<link rel="canonical" href="{SITE}/search/">\n'
        f"<style>{host_css}</style>\n</head>", 1)
    html = html.replace("<body>", f"<body>\n{bar}", 1)
    (out / "index.html").write_text(html, encoding="utf-8")

    kb = (out / "search-index.json").stat().st_size // 1024
    print(f"  search page written -- docs/search/, index {kb} KB")
    return True


def write_site(today, built):
    """
    The files around the chart: landing page, robots.txt, sitemap.xml, .nojekyll,
    and the font licence. All generated, so none of docs/ is ever hand-edited.
    """
    DOCS.mkdir(parents=True, exist_ok=True)
    (DOCS / "fonts").mkdir(exist_ok=True)

    # .nojekyll stops Pages running Jekyll, which would otherwise ignore any
    # file or folder whose name begins with an underscore.
    (DOCS / ".nojekyll").write_text("")

    licence = FONT_DIR / "OFL.txt"
    if licence.exists():
        shutil.copyfile(licence, DOCS / "fonts" / "OFL.txt")

    # The social card. Copied rather than generated -- see OG_IMAGE.
    cover = ROOT / "assets" / "og-cover.jpg"
    if cover.exists():
        shutil.copyfile(cover, DOCS / "og-cover.jpg")
    else:
        print("  WARNING: assets/og-cover.jpg missing -- social cards will be blank")

    (DOCS / "robots.txt").write_text(
        f"User-agent: *\nAllow: /\n\nSitemap: {SITE}/sitemap.xml\n", encoding="utf-8"
    )

    stamp = today.isoformat()
    # /search/ is deliberately absent. The vendored page ships
    # <meta name="robots" content="noindex">, and a sitemap entry for a page
    # that asks not to be indexed is a contradictory signal, not a stronger
    # one. If that meta is ever dropped, add the path here in the same commit.
    paths = ["/"] + [f"/{spec['slug']}/" for spec, _ in built]
    urls = "".join(
        f"\n  <url><loc>{SITE}{path}</loc><lastmod>{stamp}</lastmod></url>"
        for path in paths
    )
    (DOCS / "sitemap.xml").write_text(
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">'
        f"{urls}\n</urlset>\n",
        encoding="utf-8",
    )

    verify = (f'<meta name="google-site-verification" content="{GOOGLE_SITE_VERIFICATION}">\n'
              if GOOGLE_SITE_VERIFICATION else "")

    rows = []
    for spec, st in built:
        stats_line = (f'{st["persons"]} individuals &middot; {st["gens"]} '
                      f'generations &middot; {st["unions"]} marriages &middot; '
                      f'{st["links"]} parent&ndash;child links')
        rows.append(
            f'    <li><a href="{spec["slug"]}/">\n'
            f'      <span class="c-title">{spec["plate"]} &mdash; Genealogy {spec["numeral"]}</span>\n'
            f'      <span class="c-stats">{stats_line}</span>\n'
            '    </a></li>\n')
    for i, (plate, name, note) in enumerate(PENDING):
        # The id exists so a cross-reference can point here the day the table
        # ships; nothing links to it yet -- a link must not promise content.
        m = re.search(r"\d+", plate)
        n = m.group() if m else f"x{i + 1}"
        rows.append(
            f'    <li class="pending" id="pending-{n}">\n'
            f'      <span class="c-title">{plate} &mdash; {name} '
            '<span class="pending-tag">&middot; in preparation</span></span>\n'
            f'      <span class="c-stats">{note}</span>\n'
            '    </li>\n')
    rows = "".join(rows)

    # The finding aid, below the plates and outside the <ol>. It is not a fifth
    # plate: the list is the edition's plates in Parsons's order, and numbering
    # /search/ among them would say it is one. Counted from the built tables, so
    # it cannot go stale -- and never given a plate count in words, which is the
    # shape of claim that outlived its truth in SITE_DESCRIPTION. "entries", not
    # "people": the four plates draw 713 entries for rather fewer individuals,
    # and the search page is where that distinction is set out.
    entries = sum(st["persons"] for _, st in built)
    across = (
        '  <a class="c-across" href="search/">\n'
        '    <span class="c-title">Search the whole edition</span>\n'
        f'    <span class="c-stats">All {entries} entries, every plate '
        '&middot; by name, person number or clan</span>\n'
        '  </a>\n')

    tables = [(TABLES[k]["numeral"], TABLES[k]["slug"]) for k in sorted(TABLES)]
    landing = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Laguna Genealogies &mdash; Parsons 1923, a Digital Edition of the Laguna Pueblo Genealogical Plates</title>
<link rel="canonical" href="{SITE}/">
<meta name="description" content="{esc(SITE_DESCRIPTION)}">
<meta name="author" content="{esc(AUTHOR)}">
<meta name="keywords" content="{esc(', '.join(KEYWORDS))}">
{social_meta("Laguna Genealogies: A Digital Edition of Parsons 1923",
             SITE_DESCRIPTION, SITE + "/", kind="website")}
<meta name="theme-color" content="#FAF8F4">
<link rel="icon" href="{FAVICON}">
{verify}{jsonld_site(built, today)}
{jsonld_faq()}
<style>{font_css()}{geom_css()}{CSS}{LANDING_CSS_EXTRA}</style>
{THEME_SNIPPET}
</head>
<body>
{masthead_html(tables, None, "", None)}
<div class="titlepage">
  <div class="plate-label">A Digital Edition</div>
  <h1>LAGUNA GENEALOGIES</h1>
  <div class="rule-double"></div>
  <div class="cite">
    {CITE}
  </div>
</div>
<main>
<nav class="contents" aria-label="Contents">
  <ol>
{rows}  </ol>
{across}</nav>
</main>
<div class="prose">
  <h2>What this is</h2>
  <p>Elsie Clews Parsons published a set of foldout genealogical plates with
     <em>Laguna Genealogies</em> in 1923, in the <em>Anthropological Papers of the
     American Museum of Natural History</em>, vol. 19, pt. 5, pp. 133&ndash;292. They
     chart several generations of related families at <strong>Laguna Pueblo</strong>
     (Kawaika) in New Mexico, giving each person a number, a sex, a Keresan name and a
     clan. They are dense, hand-set, and hard to read from a scan. This edition
     transcribes them character by character &mdash; including the Americanist phonetic
     diacritics &mdash; and redraws them as text you can search, copy, and check against
     the original.</p>
  <p>Every individual on a published plate has a stable address here: select any
     person number to open their card, or link straight to them with a
     <code>#p</code> anchor. Names, clans, marriages and parent&ndash;child links are
     all real text, so they can be searched, copied and cited.</p>
  <p>Nothing has been corrected, normalised or filled in. Where the plate contains a
     misprint it is reproduced and annotated rather than silently fixed; where Parsons
     recorded no name, the entry stays blank.</p>

  <h2>When the genealogies were recorded</h2>
  <p>Parsons recorded Genealogy I at Laguna in <strong>February 1918</strong>. She
     returned in <strong>June 1919</strong> for Genealogies II, III and IV, and used
     that visit to revise Genealogy I as well &mdash; chiefly the spelling of the
     names. The ages and vital notes on the plates are therefore as of that fieldwork,
     not as of publication in 1923.</p>
  <p>This is what <em>d.</em> means on a plate: the person had already died when
     Parsons recorded the genealogy. Where she knew the year she gives it, as in
     <em>d.&nbsp;1913</em>; <em>d.</em> alone means she did not record it. A number
     after a name is that person&rsquo;s age when the data was collected.</p>

  <h2>Provenance and use</h2>
  <p>This is Laguna Pueblo material. Parsons's Laguna fieldwork is itself contested:
     she published information that members of the community regarded as restricted.
     This edition is a transcription of an already published source and adds no new
     information about the community. It is offered as a finding aid for the printed
     record.</p>
  <p>The 1923 publication is in the public domain in the United States. The
     transcription, encoding and layout are by {esc(AUTHOR)} and are released under
     <a href="https://creativecommons.org/licenses/by/4.0/">CC&nbsp;BY&nbsp;4.0</a>.</p>
  <p>Data, source scan and the code that generated these pages are in the
     <a href="{REPO}">project repository</a>.
     Corrections are welcome and are recorded as dated commits, so the edition carries
     its own revision history.</p>
{faq_html()}
  <p class="updated">Last updated
     <time datetime="{stamp}">{today.strftime("%-d %B %Y")}</time>.</p>
</div>
{THEME_FOOT}
<script>{LANDING_JS}</script>
</body>
</html>
"""
    (DOCS / "index.html").write_text(landing, encoding="utf-8")

    # Pages serves docs/404.html for any unmatched path, at any depth, so this
    # page's links must be site-absolute rather than relative.
    not_found = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Page not found &mdash; Laguna Genealogies</title>
<meta name="robots" content="noindex">
<meta name="theme-color" content="#FAF8F4">
<link rel="icon" href="{FAVICON}">
<style>{font_css()}{geom_css()}{CSS}{LANDING_CSS_EXTRA}</style>
{THEME_SNIPPET}
</head>
<body>
{masthead_html(tables, None, "/", "/")}
<div class="titlepage">
  <div class="plate-label">404</div>
  <h1>PAGE NOT FOUND</h1>
  <div class="rule-double"></div>
</div>
<div class="prose">
  <p>That address is not part of this edition. The published plates are:</p>
  <ul>
{"".join(f'    <li><a href="/{spec["slug"]}/">{spec["plate"]} &mdash; Genealogy {spec["numeral"]}</a></li>' + chr(10) for spec, _ in built)}  </ul>
  <p>Person links look like
     <code>/genealogy-i/#p42</code> &mdash; if you followed one and landed here, the
     table slug is probably wrong rather than the person number.</p>
  <p><a href="/">Return to the contents</a>.</p>
</div>
</body>
</html>
"""
    (DOCS / "404.html").write_text(not_found, encoding="utf-8")
    print("  wrote docs/index.html, 404.html, robots.txt, sitemap.xml, "
          ".nojekyll, og-cover.jpg, fonts/OFL.txt")


# ---------------------------------------------------------------------------
# THE LEAK GATE.
#
# Two things must never reach docs/, and until 2026-07-28 only the first was
# checked:
#
#   1. A research FIELD rendered into the page. It carries a class, so it is
#      grepped for exactly.
#   2. The same information written as a SENTENCE. This is the realistic way
#      research escapes -- an apparatus note explaining *why* a reading was
#      made -- and it carries no class at all, so the marker check is blind to
#      it. A note reading "the 1914 census lists her as widowed with two
#      daughters" would have published cleanly.
#
# ALLOWED holds the phrases the edition legitimately publishes: they state the
# privacy boundary rather than crossing it. The gate fails CLOSED -- reword the
# FAQ and the build stops until the new phrasing is allowlisted here. That is
# the right direction for a rule whose breach cannot be undone: a stopped build
# costs a minute, a published census match cannot be recalled.
# ---------------------------------------------------------------------------
LEAK_MARKERS = ('class="eng"', 'class="census"')

RESEARCH_PROSE = re.compile(
    r"census|famil\w*\s*search|ancestry|findagrave|national\s*archives|"
    r"\bwidow(?:ed|er|s)?\b|\benumerat\w*", re.I)

RESEARCH_PROSE_ALLOWED = (
    "census matches or identifications of living people",
    "hard to match to census records",
    "spellings used by census takers",
    # A SECOND KIND of entry, added 2026-07-30. The three above are the FAQ
    # stating the privacy boundary. This one is a quotation from the edition's
    # own published source -- Parsons 1923, p. 195 -- in Genealogy II's
    # note-paternity, and "widow" is her word about a man dead before her
    # 1918-19 fieldwork. Quoting the 1923 paper is not research escaping; it is
    # the primary text, and an editorial attribution that cites its source is
    # worth more than one that gestures at it. The phrase is allowlisted in
    # full rather than by weakening the pattern, so the gate still fails closed
    # on every other use of the word, and it must not cross a source-line break
    # -- the check is an exact substring replace against the rendered HTML.
    "by the widow, not by the sister of the deceased or his brothers",
    # A THIRD KIND, added 2026-08-09 with the search page. This is the finding
    # aid stating the boundary in its own words -- the same category as the
    # three FAQ phrases above, arriving from vendor/search/ rather than from a
    # template here. It is allowlisted VERBATIM from that project's own
    # RESEARCH_PROSE_ALLOWED, and the two lists must stay in step: it runs this
    # same gate over everything it writes, so a phrase either project rewords
    # stops one build or the other.
    #
    # Note where the sentence actually lives. In `index.html` it does not --
    # it sits in `search.js`, which check_published_pages() never opens,
    # because that sweep globs *.html only. So this entry is what keeps the
    # gate honest if the script is ever inlined, not what makes today's build
    # pass. Don't delete it on the evidence that removing it changes nothing.
    "No census matches and no identifications of living people appear here",
)


def leak_report(html):
    """The first thing in `html` that must not be published, or None."""
    for marker in LEAK_MARKERS:
        if marker in html:
            return f"research chip {marker}"
    # CSS is not prose. The stylesheet ships `.eng` and `.census` rules to every
    # page -- unused in public output, which is why the marker check above looks
    # for class="census" and not for the selector -- and a selector name says
    # nothing about a person. Scripts stay in scope: a string in the card
    # builder could carry real text. Style blocks are dropped, nothing else.
    scrubbed = re.sub(r"<style\b[^>]*>.*?</style>", "", html, flags=re.S | re.I)
    for ok in RESEARCH_PROSE_ALLOWED:
        scrubbed = scrubbed.replace(ok, "")
    m = RESEARCH_PROSE.search(scrubbed)
    if not m:
        return None
    lo = max(0, m.start() - 70)
    quote = " ".join(scrubbed[lo:m.end() + 70].split())
    return f"research vocabulary {m.group(0)!r} in: ...{quote}..."


def check_published_pages():
    """
    Every HTML file in docs/, not just the table pages.

    The per-table check runs inside build_table and never saw the landing page
    or 404.html -- which is where the FAQ lives, i.e. the only public prose that
    discusses this vocabulary at all. Anything that fails is deleted rather than
    left on disk for a later `git add -A` to sweep up.
    """
    leaked = []
    pages = sorted(DOCS.rglob("*.html"))
    for f in pages:
        hit = leak_report(f.read_text(encoding="utf-8"))
        if hit:
            f.unlink()
            leaked.append(f"{f.relative_to(DOCS)} -- {hit}")
    if leaked:
        print("ABORTED: research data in published output; offending files deleted:")
        for line in leaked:
            print(f"  {line}")
        return False
    print(f"  no research chips or vocabulary in {len(pages)} published pages")
    return True


def build_table(spec, public, today):
    """Build one table. Returns (doc, stats) so the caller can assemble the site."""
    if public:
        persons, unions, ku, km, ub = load_baseline(spec)
        out = DOCS / spec["slug"] / "index.html"
    else:
        if not XLSX.exists():
            print(f"missing {XLSX}; run build_workbook.py first")
            return None, None
        persons, unions, ku, km, ub = load()
        out = OUT

    chart = Chart(persons, unions, ku, km, ub)
    # A root starts at generation 1 unless spec["root_columns"] sets it further
    # in -- see the Genealogy II entry in TABLES for why this is an indent and
    # not an UNATTACHED_BLOCKS splice. The offset is stated in the grid's own
    # tokens, so it is the same step the nesting produces and drift stays 0.
    rc = spec.get("root_columns", {})
    def _tree(r):
        col = rc.get(r, 1)
        style = ("" if col <= 1 else
                 ' style="margin-inline-start:calc((var(--col) + var(--stub))'
                 f' * {col - 1})"')
        return f'<div class="tree"{style}>{chart.render(r)[0]}</div>'
    trees = "".join(_tree(r) for r in spec["roots"])

    missing = sorted(set(persons) - chart.seen)
    links = sum(len(v) for v in ku.values()) + sum(len(v) for v in km.values())
    n_gens = max((int(p["generation"]) for p in persons.values() if p["generation"]),
                 default=0)
    gens = NUMBER_WORDS.get(n_gens, str(n_gens))

    status = [f"<li>{len(persons)} individuals across {n_gens} generation columns; "
              f"{len(unions)} marriages; {links} "
              "parent&ndash;child links.</li>"]
    if public:
        status.append("<li>Transcribed from the plate and verified against it. Clan "
                      "descent is matrilineal, which gives every bracket an independent "
                      f"check; all {len(unions)} marriages pass it.</li>")
    else:
        filled = sorted(p["id"] for p in persons.values() if p["english_name"])
        census_filled = sorted(p["id"] for p in persons.values() if p["census_name"])
        if filled:
            status.append("<li>English names filled in for: "
                          + ", ".join(str(i) for i in filled) + ".</li>")
        else:
            status.append("<li>No English names filled in yet. Add them in the "
                          "<code>english_name</code> column of PERSONS and re-run this script.</li>")
        if census_filled:
            status.append("<li>Census matches recorded for: "
                          + ", ".join(str(i) for i in census_filled) + ".</li>")
    # An undrawn person used to be a status line and a console warning, which
    # is how seven of Genealogy II's went unnoticed through a whole session:
    # nothing fails, the page just quietly holds fewer people than the plate.
    # It is now fatal on the published build, like a duplicate anchor. The
    # private build keeps reporting it, because a half-read plate legitimately
    # has people no bracket reaches yet.
    if missing:
        if public:
            raise SystemExit(
                f"ABORTED: {len(missing)} persons in PERSONS are not drawn in "
                f"{out.name}: {missing}. Every person the plate numbers must reach "
                "the page -- check roots, drawn_under and UNATTACHED_BLOCKS.")
        status.append(f"<li><strong>Not drawn:</strong> {missing} &mdash; these ids are in "
                      "PERSONS but reachable from neither founding couple.</li>")

    stats = {"persons": len(persons), "unions": len(unions),
             "links": links, "gens": gens}
    doc = build_doc(spec, describe(spec, len(persons), n_gens), gens, n_gens,
                    trees, status, public, today, stats, persons, unions, ku, km,
                    chart.seen)

    # Every drawn person must carry exactly one citable id="p{n}" -- its first
    # printed occurrence. A duplicate would silently break every deep link to
    # the later occurrence, so it aborts the build like the privacy grep does.
    ids = re.findall(r'id="(p\d+)"', doc)
    if len(ids) != len(set(ids)):
        dupes = sorted({i for i in ids if ids.count(i) > 1})
        raise SystemExit(f"ABORTED: duplicate person anchors {dupes} in {out.name}")
    if set(ids) != {f"p{i}" for i in chart.seen}:
        raise SystemExit(f"ABORTED: person anchors do not match drawn persons in {out.name}")

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(doc, encoding="utf-8")
    print(f"wrote {out.relative_to(ROOT) if out.is_relative_to(ROOT) else out}")
    print(f"  {len(chart.seen)} of {len(persons)} persons drawn")
    if missing:
        print(f"  NOT DRAWN: {missing}")

    if public:
        # A published build that leaked a research chip would be the one failure
        # that cannot be walked back, so check the bytes rather than trusting the
        # data path that produced them.
        hit = leak_report(doc)
        if hit:
            out.unlink()
            where = out.relative_to(ROOT) if out.is_relative_to(ROOT) else out
            raise SystemExit(f"ABORTED: {hit} in {where}; output deleted")
        print("  no english/census chips in output")

    return doc, stats


def main():
    ap = argparse.ArgumentParser(description=__doc__.strip().splitlines()[0])
    ap.add_argument("--public", action="store_true",
                    help="build the published page(s) from the 1923 baseline into docs/")
    ap.add_argument("--table", default="i", choices=sorted(TABLES),
                    help="which plate to build (default: i)")
    args = ap.parse_args()
    today = dt.date.today()

    if not args.public:
        # The workbook holds one table, so the private build is always that one.
        doc, _ = build_table(TABLES[args.table], public=False, today=today)
        return 0 if doc else 1

    # Publishing rebuilds every table, so the landing page, sitemap and the
    # per-table counts can never describe a stale set.
    built = []
    for key in sorted(TABLES):
        spec = TABLES[key]
        _, stats = build_table(spec, public=True, today=today)
        built.append((spec, stats))
    write_site(today=today, built=built)

    # Before the sweep, so docs/search/index.html is swept like any other page.
    # It is the one page here whose prose this build did not write.
    if not write_search([(TABLES[k]["numeral"], TABLES[k]["slug"])
                         for k in sorted(TABLES)]):
        return 1

    # Sweep every published page before anything else passes judgement on it.
    if not check_published_pages():
        return 1

    pages = [DOCS / "index.html"] + [
        DOCS / spec["slug"] / "index.html" for spec, _ in built]
    if not check_structured_data(pages):
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
