"""
Build parsons_genealogy_I.xlsx -- the editable workbook -- from the 1923 baseline.

    python3 scripts/build_workbook.py

WARNING: this OVERWRITES the workbook and therefore discards any research data
you have typed into it. Run it once to create the file. After that, edit the
.xlsx directly and use make_chart.py to regenerate the chart.
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).parent))
import transcription as T  # noqa: E402

OUT = Path(__file__).parent.parent / "data" / "parsons_genealogy_I.xlsx"

SOURCE_HDR = PatternFill("solid", fgColor="D9D9D9")   # 1923 transcription
EDIT_HDR = PatternFill("solid", fgColor="C6E0B4")     # yours to fill in
DERIVED_HDR = PatternFill("solid", fgColor="DDEBF7")  # formula, don't type here
EDIT_BODY = PatternFill("solid", fgColor="F2F9EE")
HDR_FONT = Font(bold=True, size=10)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PERSONS_HDR = [
    ("id", 6, SOURCE_HDR),
    ("generation", 11, SOURCE_HDR),
    ("sex", 6, SOURCE_HDR),
    ("name_as_printed", 24, SOURCE_HDR),
    ("name_ascii", 20, SOURCE_HDR),
    ("alt_name", 14, SOURCE_HDR),
    ("age", 6, SOURCE_HDR),
    ("clan", 16, SOURCE_HDR),
    ("vital_note", 18, SOURCE_HDR),
    ("origin", 9, SOURCE_HDR),
    ("cross_ref", 34, SOURCE_HDR),
    ("plate_note", 34, SOURCE_HDR),
    ("verified", 9, SOURCE_HDR),
    ("english_name", 18, EDIT_HDR),
    ("census_name", 20, EDIT_HDR),
    ("census_year", 12, EDIT_HDR),
    ("match_confidence", 17, EDIT_HDR),
    ("notes", 44, EDIT_HDR),
]
FIRST_EDIT_COL = 14  # english_name


def style_header(ws, spec, freeze):
    for i, (title, width, fill) in enumerate(spec, start=1):
        c = ws.cell(row=1, column=i, value=title)
        c.font = HDR_FONT
        c.fill = fill
        c.border = BOX
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(spec))}1"


def build_persons(wb):
    ws = wb.create_sheet("PERSONS")
    style_header(ws, PERSONS_HDR, "E2")

    for r, (pid, gen, sex, name, alt, age, clan, vital, origin, xref, pnote) in enumerate(
        T.PERSONS, start=2
    ):
        row = [
            pid, gen, sex, name, T.fold(name), alt,
            int(age) if age.isdigit() else age,
            clan, vital, origin, xref, pnote, "yes",
        ]
        for i, v in enumerate(row, start=1):
            ws.cell(row=r, column=i, value=v).border = BOX
        for i in range(FIRST_EDIT_COL, len(PERSONS_HDR) + 1):
            c = ws.cell(row=r, column=i)
            c.fill = EDIT_BODY
            c.border = BOX
        ws.cell(row=r, column=11).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=12).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=18).alignment = Alignment(wrap_text=True, vertical="top")

    # 'Hazel' is printed on the plate itself, so it is a source fact, not a guess.
    hazel = next(r for r, p in enumerate(T.PERSONS, start=2) if p[0] == 90)
    ws.cell(row=hazel, column=14, value="Hazel")
    ws.cell(row=hazel, column=17, value="certain")
    ws.cell(row=hazel, column=18,
            value="Printed in the 1923 plate as 'F. Heʼsa (Hazel). Badger' -- "
                  "not a research addition.")

    last = len(T.PERSONS) + 1
    dv_sex = DataValidation(type="list", formula1='"F,M"', allow_blank=True)
    dv_clan = DataValidation(type="list", formula1=f'"{",".join(T.CLANS)}"', allow_blank=True)
    dv_conf = DataValidation(type="list", formula1='"certain,probable,possible,rejected"',
                             allow_blank=True)
    dv_ver = DataValidation(type="list", formula1='"yes,no,recheck"', allow_blank=True)
    for dv, col in ((dv_sex, "C"), (dv_clan, "H"), (dv_ver, "M"), (dv_conf, "Q")):
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{last}")
    return ws


def _lookup(id_cell, col):
    return f'=IF({id_cell}=0,"",IFERROR(VLOOKUP({id_cell},PERSONS!$A:$R,{col},FALSE),""))'


NAME_COL, ENGLISH_COL = 4, 14


def build_unions(wb):
    ws = wb.create_sheet("UNIONS")
    spec = [
        ("union_id", 10, SOURCE_HDR), ("wife_id", 8, SOURCE_HDR),
        ("wife_name", 24, DERIVED_HDR), ("wife_english", 16, DERIVED_HDR),
        ("husband_id", 11, SOURCE_HDR), ("husband_name", 24, DERIVED_HDR),
        ("husband_english", 16, DERIVED_HDR),
        ("wife_marriage_order", 12, SOURCE_HDR), ("husband_marriage_order", 12, SOURCE_HDR),
        ("note", 70, SOURCE_HDR),
    ]
    style_header(ws, spec, "B2")
    for r, (uid, wife, husb, worder, horder, note) in enumerate(T.UNIONS, start=2):
        vals = [
            uid, wife, _lookup(f"B{r}", NAME_COL), _lookup(f"B{r}", ENGLISH_COL),
            husb, _lookup(f"E{r}", NAME_COL), _lookup(f"E{r}", ENGLISH_COL),
            worder, horder, note,
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BOX
            if i in (3, 4, 6, 7):
                c.fill = PatternFill("solid", fgColor="F7FBFF")
        ws.cell(row=r, column=10).alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def build_children(wb):
    ws = wb.create_sheet("CHILDREN")
    spec = [
        ("union_id", 10, SOURCE_HDR),
        ("mother_id", 10, SOURCE_HDR), ("mother_name", 22, DERIVED_HDR),
        ("father_id", 10, SOURCE_HDR), ("father_name", 22, DERIVED_HDR),
        ("child_id", 9, SOURCE_HDR), ("child_name", 22, DERIVED_HDR),
        ("child_english", 16, DERIVED_HDR),
        ("note", 70, SOURCE_HDR),
    ]
    style_header(ws, spec, "B2")
    for r, (uid, mother, father, child, note) in enumerate(T.CHILDREN, start=2):
        vals = [
            uid, mother, _lookup(f"B{r}", NAME_COL),
            father, _lookup(f"D{r}", NAME_COL),
            child, _lookup(f"F{r}", NAME_COL), _lookup(f"F{r}", ENGLISH_COL),
            note,
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BOX
            if i in (3, 5, 7, 8):
                c.fill = PatternFill("solid", fgColor="F7FBFF")
        ws.cell(row=r, column=9).alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def build_plate_notes(wb):
    ws = wb.create_sheet("PLATE_NOTES")
    style_header(ws, [("location", 22, SOURCE_HDR), ("text_as_printed", 66, SOURCE_HDR),
                      ("context", 60, SOURCE_HDR)], "A2")
    for r, row in enumerate(T.PLATE_NOTES, start=2):
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BOX
            c.alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def build_reference(wb):
    ws = wb.create_sheet("REFERENCE")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 62
    r = 1

    def head(text):
        nonlocal r
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=True, size=12)
        r += 2

    def line(*vals):
        nonlocal r
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    head("Source")
    line("Publication", "", "", "Elsie Clews Parsons, 'Laguna Genealogies', "
         "Anthropological Papers of the American Museum of Natural History, vol. 19 (1923)")
    line("Plate", "", "", "Table 1, 'Genealogy I' (foldout)")
    line("Scan", "", "", "sources/parsons-1923-table-1.jpg, 16172 x 11798 px")
    line("Extent", "", "", "104 numbered individuals, 27 marriages, 80 parent-child links, "
         "5 generation columns, two founding couples (1+2 and 54+55)")
    r += 1

    head("How to use this workbook")
    line("Grey columns", "", "", "The 1923 transcription. Treat as read-only so the data stays "
         "auditable against the plate.")
    line("Green columns", "", "", "Yours: english_name, census_name, census_year, "
         "match_confidence, notes. Add research data only here.")
    line("Blue columns", "", "", "Formulas that look up names from PERSONS. Don't type in them.")
    line("Regenerate chart", "", "", "python3 scripts/make_chart.py")
    line("Do not re-run", "", "", "scripts/build_workbook.py overwrites this file and would "
         "discard everything in the green columns.")
    r += 1

    head("Conventions in the transcription")
    line("name_as_printed", "", "", "Unicode-exact, including diacritics. A blank means the plate "
         "printed a dash instead of a name.")
    line("name_ascii", "", "", "Diacritic-free lowercase key, for matching census spellings.")
    line("alt_name", "", "", "A second name printed in parentheses or braced with the first.")
    line("age", "", "", "Age in italics on the plate, i.e. as of Parsons' fieldwork.")
    line("vital_note", "", "", "'d.' alone means died, date unknown.")
    line("id = 0", "", "", "In UNIONS/CHILDREN, means that person is not shown on the plate.")
    line("father_id = 0", "", "", "The plate draws the sibling group off the mother's line only, "
         "so paternity cannot be assigned.")
    r += 1

    head("Clans on this plate")
    for clan in T.CLANS:
        line(clan)
    r += 1

    head("Orthography key")
    line("glyph", "codepoint", "name", "value")
    for glyph, code, name, note in T.ORTHOGRAPHY:
        line(glyph, code, name, note)
    r += 1

    head("Editorial observations")
    line("Matrilineal check", "", "", "Every sibling group's clan matches its mother's clan, "
         "as Laguna matrilineal descent requires. This was used to verify the bracket readings.")
    line("Misprint at 76", "", "", "The '+' line under 76 is numbered 68, but names "
         "Shuwaiʼᶦri, Turkey = person 67. 67's own cross-reference confirms 67. "
         "See the note on union U23.")
    line("Cross-link", "", "", "Person 8 (Yu˙si) appears twice: as husband of 7 and 73 in the "
         "upper half, and as a son of 58+59 in the lower half. He joins the two founding lines.")
    line("Outside Genealogy I", "", "", "Persons 12 and 73 have spouses and offspring recorded "
         "in Genealogy II and III (Tables 2 and 3 of the same publication), not transcribed here.")
    return ws


def main():
    problems = T.self_check()
    if problems:
        print("refusing to build; transcription self-check failed:")
        for p in problems:
            print("  -", p)
        return 1

    wb = Workbook()
    wb.remove(wb.active)
    build_persons(wb)
    build_unions(wb)
    build_children(wb)
    build_plate_notes(wb)
    build_reference(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  PERSONS  {len(T.PERSONS)} rows")
    print(f"  UNIONS   {len(T.UNIONS)} rows")
    print(f"  CHILDREN {len(T.CHILDREN)} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
